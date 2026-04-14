import anvil.server
import anvil.users
from anvil.tables import app_tables
import anvil.media
import mysql.connector
from mysql.connector import pooling
from dotenv import load_dotenv
import os
from contextlib import contextmanager
import re
import pdfkit
import decimal
import base64
import html as html_module
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import bcrypt


# Load environment variables from .env file
load_dotenv()

# Set your wkhtmltopdf path here (adjust for your system)
WKHTMLTOPDF_PATH = os.getenv("WKHTMLTOPDF_PATH")  # Windows path
config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)


# ---------------------------------------------------------------------------
# Connection pool — reuses TCP connections instead of opening one per request
# ---------------------------------------------------------------------------
_db_pool = pooling.MySQLConnectionPool(
    pool_name="bma_pool",
    pool_size=int(os.getenv("DB_POOL_SIZE", "5")),
    host=os.getenv("DB_HOST"),
    user=os.getenv("DB_USER"),
    password=str(os.getenv("DB_PASSWORD")),
    port=int(os.getenv("DB_PORT", "3306")),
    database=os.getenv("DB_NAME"),
    auth_plugin=os.getenv("DB_AUTH_PLUGIN", "caching_sha2_password"),
)


def get_db_connection():
    """Return a connection from the shared pool."""
    return _db_pool.get_connection()


# ---------------------------------------------------------------------------
# Authorization helpers
# ---------------------------------------------------------------------------

def _get_current_user():
    """Return the currently logged-in Anvil user, or raise if not logged in."""
    user = anvil.users.get_user()
    if user is None:
        raise Exception("Authentication required.")
    return user


def _require_role(*allowed_roles):
    """Raise an exception if the current user does not hold one of the allowed roles."""
    user = _get_current_user()
    role_id = user["role_id"]
    # Fetch the role name from MySQL so we can compare by name
    with db_cursor() as cursor:
        cursor.execute("SELECT Roles FROM tbl_roles WHERE ID = %s", (role_id,))
        row = cursor.fetchone()
    role_name = row[0] if row else None
    if role_name not in allowed_roles:
        raise Exception("You do not have permission to perform this action.")
    return user


# Context manager for DB cursor
@contextmanager
def db_cursor():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        yield cursor
        connection.commit()
    except Exception as e:
        connection.rollback()
        raise e
    finally:
        cursor.close()
        connection.close()


# ***************************************************FAQ Section ************************************


@anvil.server.callable()
def get_faq_html():
    _get_current_user()
    with db_cursor() as cursor:
        # Fetch questions and answers from table
        cursor.execute("SELECT Question, Answer FROM tbl_faqs ORDER BY id ASC")
        faqs = cursor.fetchall()

        # HTML Template start
        html = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>FAQs</title>
            <style>
                /* --- Embed Mozilla Headline font --- */
                @font-face {{
                    font-family: 'Mozilla Headline';
                    src: url(data:font/ttf;base64,{font_base64}) format('truetype');
                }}
                .faq-container {
                    max-width: 800px;
                    margin: auto;
                    padding: 20px;
                    font-family: Arial, sans-serif;
                }
                .faq-container h2 {
                    text-align: center;
                    margin-bottom: 20px;
                    font-family: 'Mozilla Headline';
                }
                .faq-item {
                    border-bottom: 1px solid #ccc;
                }
                .faq-question {
                    background-color: #f9f9f9;
                    color: #333;
                    padding: 15px;
                    width: 100%;
                    text-align: left;
                    border: none;
                    outline: none;
                    font-size: 16px;
                    font-weight: bold;
                    cursor: pointer;
                    transition: background-color 0.3s ease;
                    font-family: 'Mozilla Headline';
                }
                .faq-question:hover {
                    background-color: #eee;
                }
                .faq-answer {
                    display: none;
                    padding: 15px;
                    background-color: #fff;
                    color: #555;
                    font-family: 'Mozilla Headline';
                }
                .faq-answer p {
                    margin: 0;
                }
            </style>
        </head>
        <body>
            <div class="faq-container">
                <h2>Frequently Asked Questions</h2>
        """

        # Add FAQ items dynamically — escape DB content to prevent XSS
        for row in faqs:
            question = html_module.escape(str(row[0]))
            answer = html_module.escape(str(row[1]))
            html += f"""
            <div class="faq-item">
                <button class="faq-question">{question}</button>
                <div class="faq-answer">
                    <p>{answer}</p>
                </div>
            </div>
            """

        # Close HTML and add JavaScript
        html += """
            </div>
                <script>
                        (function(){
                            const questions = document.querySelectorAll(".faq-question");
                            questions.forEach((question) => {
                                question.addEventListener("click", function () {
                                    const answer = this.nextElementSibling;
                                    document.querySelectorAll(".faq-answer").forEach((a) => {
                                        if (a !== answer) a.style.display = "none";
                                    });
                                    answer.style.display =
                                        answer.style.display === "block" ? "none" : "block";
                                });
                            });
                        })();
                </script>
        </body>
        </html>
        """

        return html


# ***************************************************Client Section ************************************


# Get client's name to be displayed in job card
@anvil.server.callable()
def getClientFullname():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Fullname, Phone, Address, Email, Narration FROM tbl_clientcontacts
            ORDER BY Fullname ASC  
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        result = [
            {
                "ID": row[0],
                "Fullname": row[1],
                "Phone": row[2],
                "Address": row[3],
                "Email": row[4],
                "Narration": row[5],
            }
            for row in rows
        ]
    return result


@anvil.server.callable()
def getClientFullnameFromSearchWord(valueSearch):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Fullname FROM tbl_clientcontacts
            WHERE Fullname LIKE %s
            ORDER BY Fullname ASC  
        """
        like_pattern = f"%{valueSearch}%"
        cursor.execute(query, (like_pattern,))
        rows = cursor.fetchall()
        return [(row[1], row[0]) for row in rows]  # (RegNo, ID)


# Get client details with ID
@anvil.server.callable()
def getClientNameWithID(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Fullname, Phone, Address, Email, Narration FROM tbl_clientcontacts
            WHERE ID = %s 
        """
        cursor.execute(query, (valueID,))
        rows = cursor.fetchall()
        result = [
            {
                "ID": row[0],
                "Fullname": row[1],
                "Phone": row[2],
                "Address": row[3],
                "Email": row[4],
                "Narration": row[5],
            }
            for row in rows
        ]

    return result[0]


# Save new client data
@anvil.server.callable()
def save_client_data(name, phone, address, email, narration):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_clientcontacts (Fullname, Phone, Address, Email, Narration)
            VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (name, phone, address, email, narration))


# Get Client report
@anvil.server.callable()
def getClientReport(clientID):
    _get_current_user()
    with db_cursor() as cursor:
        if clientID is None:
            query = (
                "SELECT ID , Fullname, Phone, Address, Email, Narration "
                "FROM tbl_clientcontacts "
                "ORDER BY Fullname ASC"
            )
            cursor.execute(query)
        else:
            query = (
                "SELECT ID , Fullname, Phone, Address, Email, Narration "
                "FROM tbl_clientcontacts "
                "WHERE ID = %s"
            )
            cursor.execute(query, (clientID,))

        rows = cursor.fetchall()
        result = [
            {
                "No": index + 1,
                "ID": row[0],
                "Fullname": row[1],
                "Phone": row[2],
                "Address": row[3],
                "Email": row[4],
                "Narration": row[5],
            }
            for index, row in enumerate(rows)
        ]

    return result


@anvil.server.callable()
def updateClientDetails(client_id, name, phone, address, email, narration):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_clientcontacts
            SET Fullname = %s,
                Phone = %s,
                Address = %s,
                Email = %s,
                Narration = %s
            WHERE ID = %s
        """
        cursor.execute(query, (name, phone, address, email, narration, client_id))


@anvil.server.callable()
def getClientNameAndPhoneNumber(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT ID, Fullname, Phone 
        FROM tbl_clientcontacts 
        WHERE Fullname LIKE %s OR Phone LIKE %s
        ORDER BY Fullname ASC
        """
        like_value = f"%{value}%"
        cursor.execute(query, (like_value, like_value))
        result = cursor.fetchall()
        # Return tuples: (Label shown, Value stored)
        return [(f"{r[1]} - {r[2]}", r[0]) for r in result]


# ***************************************************Technicians Details Section ************************************


@anvil.server.callable()
def getTechnicianReport(technicianName):
    _get_current_user()
    with db_cursor() as cursor:
        if technicianName is None:
            query = """
                SELECT tbl_technicians.ID, tbl_technicians.Fullname, tbl_technicians.Phone, tbl_toolkits.ToolkitName, tbl_technicians.IsArchived 
                FROM tbl_technicians 
                INNER JOIN 
                tbl_toolkits ON tbl_technicians.ToolkitID = tbl_toolkits.ID 
                 ORDER BY Fullname ASC
                """

            cursor.execute(query)
        else:
            query = """
                SELECT tbl_technicians.ID, tbl_technicians.Fullname, tbl_technicians.Phone, tbl_toolkits.ToolkitName,tbl_technicians.IsArchived 
                FROM tbl_technicians 
                INNER JOIN 
                tbl_toolkits ON tbl_technicians.ToolkitID = tbl_toolkits.ID 
                WHERE tbl_technicians.Fullname Like %s
                """
            cursor.execute(query, (f"%{technicianName}%",))
        # Fetch all rows
        rows = cursor.fetchall()
        result = [
            {
                "No": index + 1,
                "ID": row[0],
                "Fullname": row[1],
                "Phone": row[2],
                "Toolkit": row[3],
                "IsArchived": "Yes" if row[4] == 1 else "No",
            }
            for index, row in enumerate(rows)
        ]

    return result


@anvil.server.callable()
def getTechnicians():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Fullname
            FROM tbl_technicians
            WHERE IsArchived = FALSE
            ORDER BY Fullname ASC
        """
        cursor.execute(query)
        result = cursor.fetchall()
        return [{"ID": r[0], "Fullname": r[1]} for r in result]


@anvil.server.callable()
def getAssignedTechnician(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT tbl_technicians.Fullname 
            FROM tbl_pendingassignedjobs 
            INNER JOIN 
            tbl_technicians ON tbl_pendingassignedjobs.TechnicianID = tbl_technicians.ID 
            WHERE tbl_pendingassignedjobs.JobCardRefID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()
        if result is None:
            return None
        else:
            return result[0]


@anvil.server.callable()
def save_technician_data(fullname, phone, toolkit):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_technicians(Fullname,Phone,ToolkitID,IsArchived) VALUES (%s, %s, %s, %s)
        """
        cursor.execute(query, (fullname, phone, toolkit, 0))


@anvil.server.callable()
def get_technician_details(value):
    _get_current_user()
    with db_cursor() as cursor:
        if value is None:
            query = """
                SELECT ID, Fullname, Phone, ToolkitID, IsArchived 
                FROM tbl_technicians 
                ORDER BY Fullname ASC
            """
            cursor.execute(query)
        else:
            query2 = """
                            SELECT ID, Fullname, Phone, ToolkitID, IsArchived 
                            FROM tbl_technicians 
                            WHERE ID = %s
                        """
            cursor.execute(query2, (value,))

        rows = cursor.fetchall()
        result = [
            {
                "ID": row[0],
                "Fullname": row[1],
                "Phone": row[2],
                "ToolkitID": row[3],
                "Active": "Yes" if row[4] == 1 else "No",
            }
            for row in rows
        ]

        return result


@anvil.server.callable()
def update_technician_data(name, phone, toolkit, archived, technician_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_technicians
            SET Fullname = %s,
                Phone = %s,
                ToolkitID = %s,
                IsArchived = %s
            WHERE ID = %s
        """
        cursor.execute(query, (name, phone, toolkit, archived, technician_id))


@anvil.server.callable()
def getTechnicianDetailsFromName(valueName):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Fullname
            FROM tbl_technicians
            WHERE IsArchived = FALSE AND Fullname = %s
            ORDER BY Fullname ASC
        """
        cursor.execute(query, (valueName,))
        result = cursor.fetchall()
        return [{"ID": r[0], "Fullname": r[1]} for r in result]


@anvil.server.callable()
def getTechnicianNameAndID(valueName):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Fullname
            FROM tbl_technicians
            WHERE Fullname LIKE %s
            ORDER BY Fullname ASC
        """
        cursor.execute(query, (f"%{valueName}%",))
        result = cursor.fetchall()
        return [(f"{r[1]}", r[0]) for r in result]


@anvil.server.callable()
def getTechnicianInJobCard():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Fullname
            FROM tbl_technicians
            WHERE IsArchived = FALSE
            ORDER BY Fullname ASC
        """
        cursor.execute(query)
        result = cursor.fetchall()
        return [(r[1], r[0]) for r in result]


# ***************************************************Toolkit Section ************************************


@anvil.server.callable()
def get_toolkits(value):
    _get_current_user()
    with db_cursor() as cursor:
        if value is None:
            query = """
                SELECT ID, ToolkitName FROM tbl_toolkits ORDER BY ToolkitName ASC
            """
            cursor.execute(query)
        else:
            query2 = """
                        SELECT ID, ToolkitName FROM tbl_toolkits WHERE ID = %s
                    """
            cursor.execute(query2, (value,))

        result = cursor.fetchall()

        return [{"ID": row[0], "ToolkitName": row[1]} for row in result]


# Check for existing toolkits
@anvil.server.callable()
def check_duplicate_toolkit(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT ID 
                FROM tbl_toolkits 
                WHERE ToolkitName = %s
            """
        cursor.execute(query, (value,))

        result = cursor.fetchone()
        return result is not None


@anvil.server.callable()
def save_toolkit_data(name, amount):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                INSERT INTO tbl_toolkits(ToolkitName, Cost) 
                VALUES(%s, %s)
        """
        cursor.execute(query, (name, amount))


@anvil.server.callable()
def get_toolkit_details(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                            SELECT ID, ToolkitName, Cost
                            FROM tbl_toolkits 
                            WHERE ID = %s
                        """
        cursor.execute(query, (value,))

        rows = cursor.fetchall()
        result = [{"ID": row[0], "ToolkitName": row[1], "Cost": row[2]} for row in rows]

        return result


@anvil.server.callable()
def update_toolkit_data(name, amount, toolkit_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_toolkits
            SET ToolkitName = %s,
                Cost = %s
            WHERE ID = %s
        """
        cursor.execute(query, (name, amount, toolkit_id))


# ***************************************************Staff Details Section ************************************


@anvil.server.callable()
def save_staff_data(fullname, phone):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_checkstaff(Staff,Phone,IsArchived) VALUES (%s, %s, %s)
        """
        cursor.execute(query, (fullname, phone, 0))


@anvil.server.callable()
def get_staff_details(value):
    _get_current_user()
    with db_cursor() as cursor:
        if value is None:
            query = """
                SELECT ID, Staff, Phone, IsArchived 
                FROM tbl_checkstaff 
                ORDER BY Staff ASC
            """
            cursor.execute(query)
        else:
            query2 = """
                            SELECT ID, Staff, Phone, IsArchived
                            FROM tbl_checkstaff 
                            WHERE ID = %s
                        """
            cursor.execute(query2, (value,))

        rows = cursor.fetchall()
        result = [
            {
                "ID": row[0],
                "Fullname": row[1],
                "Phone": row[2],
                "Active": "Yes" if row[3] == 1 else "No",
            }
            for row in rows
        ]

        return result


@anvil.server.callable()
def update_staff_data(name, phone, archived, staff_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_checkstaff
            SET Staff = %s,
                Phone = %s,
                IsArchived = %s
            WHERE ID = %s
        """
        cursor.execute(query, (name, phone, archived, staff_id))


@anvil.server.callable()
def getStaffReport(staffID):
    _get_current_user()
    with db_cursor() as cursor:
        if staffID is None:
            query = (
                "SELECT ID, Staff, Phone, IsArchived"
                " FROM tbl_checkstaff "
                " ORDER BY Staff ASC"
            )
            cursor.execute(query)
        else:
            query = (
                "SELECT ID, Staff, Phone, IsArchived "
                "FROM tbl_checkstaff "
                "WHERE ID = %s"
            )
            cursor.execute(query, (staffID,))

        rows = cursor.fetchall()
        result = [
            {
                "No": index + 1,
                "ID": row[0],
                "Fullname": row[1],
                "Phone": row[2],
                "Active": "Yes" if row[3] == 1 else "No",
            }
            for index, row in enumerate(rows)
        ]

    return result


# Get Staff Details
@anvil.server.callable()
def getStaff():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT ID, Staff
                FROM tbl_checkstaff
                WHERE IsArchived = FALSE
                ORDER BY Staff ASC
            """
        cursor.execute(query)
        result = cursor.fetchall()
        return [{"ID": r[0], "Staff": r[1]} for r in result]


@anvil.server.callable()
def getStaffByName(valueName):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT ID, Staff
                FROM tbl_checkstaff
                WHERE Staff LIKE %s
                 ORDER BY Staff ASC
            """
        cursor.execute(query, (f"%{valueName}%",))
        result = cursor.fetchall()
        return [(f"{r[1]}", r[0]) for r in result]


# Get Staff Details By ID
@anvil.server.callable()
def getStaffByID(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT ID, Staff
                FROM tbl_checkstaff
                WHERE ID = %s
                ORDER BY Staff ASC
            """
        cursor.execute(query, (valueID,))
        result = cursor.fetchall()
        return [{"ID": r[0], "Staff": r[1]} for r in result]


@anvil.server.callable()
def getStaffAndTechnicianNames():
    _get_current_user()
    with db_cursor() as cursor:
        # Use 'AS combined_name' to ensure both sides of the UNION have the same key
        query = """
                SELECT Staff AS combined_name FROM tbl_checkstaff WHERE IsArchived = FALSE
                UNION ALL
                SELECT Fullname AS combined_name FROM tbl_technicians WHERE IsArchived = FALSE
                ORDER BY combined_name ASC
            """
        cursor.execute(query)
        result = cursor.fetchall()
        
        # Modern list comprehension that handles both DictCursor and TupleCursor
        names = []
        for r in result:
            # If r is a dict, get 'combined_name'. If it's a tuple/list, get index 0.
            val = r['combined_name'] if isinstance(r, dict) else r[0]
            names.append((val, val))
            
        return names
    
# ***************************************************Car Details Section ************************************


@anvil.server.callable()
def getCarRegistration():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT MAX(ID) as LastOfID, RegNo
            FROM tbl_jobcarddetails
            GROUP BY RegNo
            ORDER BY RegNo ASC
        """

        cursor.execute(query)
        result = cursor.fetchall()
        return [{"ID": r[0], "RegNo": r[1]} for r in result]


@anvil.server.callable()
def getCarRegistrationWithPartlyDetails(valueReg):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT MAX(ID) AS LastOfID, RegNo
            FROM tbl_jobcarddetails
            WHERE RegNo LIKE %s
            GROUP BY RegNo
            ORDER BY RegNo ASC
        """
        cursor.execute(query, (f"%{valueReg}%",))
        result = cursor.fetchall()

        # Return a list formatted for a DropDown in Anvil
        return [(row[1], row[0]) for row in result]  # (RegNo, ID)


@anvil.server.callable()
def get_car_details(search_term=None, limit=200):
    # Use table aliases (c and j) to make the query cleaner
    _get_current_user()
    query = """
        SELECT 
            c.Fullname, 
            c.Phone, 
            j.RegNo, 
            j.ChassisNo,
            j.ReceivedDate,
            j.JobCardRef,
            j.MakeAndModel
        FROM tbl_clientcontacts c
        INNER JOIN tbl_jobcarddetails j 
            ON c.ID = j.ClientDetails
    """
    params = []
    
    # Dynamically add the WHERE clause only if a search term is provided
    if search_term:
        query += """
            WHERE c.Fullname LIKE %s 
               OR c.Phone LIKE %s 
               OR j.RegNo LIKE %s 
               OR j.ChassisNo LIKE %s
               OR j.MakeAndModel LIKE %s
        """
        like_pattern = f"%{search_term}%"
        # Add the pattern 5 times to the parameters list
        params.extend([like_pattern] * 5)
        
    # Order directly by the date (Removed the expensive GROUP BY)
    query += " ORDER BY j.ReceivedDate DESC"
    
    # Add a LIMIT to prevent fetching massive amounts of data at once
    if limit:
        query += " LIMIT %s"
        params.append(limit)
        
    with db_cursor() as cursor:
        cursor.execute(query, tuple(params))
        results = cursor.fetchall()

    # Process and return the results
    return [
        {
            "No": i + 1,
            "Fullname": row[0],
            "Phone": row[1],
            "RegNo": row[2],
            "ChassisNo": row[3],
            "ReceivedDate": row[4],
            "JobCardRef": row[5],
            "MakeAndModel": row[6],
        }
        for i, row in enumerate(results)
    ]

@anvil.server.callable()
def get_car_details_and_parts(search_term, part_name):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT 
                    tbl_clientcontacts.Fullname, 
                    tbl_clientcontacts.Phone, 
                    tbl_jobcarddetails.RegNo, 
                    tbl_jobcarddetails.ChassisNo,
                    tbl_jobcarddetails.ReceivedDate,
                    tbl_jobcarddetails.JobCardRef,
                    tbl_jobcarddetails.MakeAndModel
                FROM 
                    tbl_clientcontacts 
                INNER JOIN 
                    tbl_jobcarddetails 
                ON 
                    tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails 
                INNER JOIN 
                    tbl_invoices
                ON
                    tbl_invoices.AssignedJobID = tbl_jobcarddetails.ID
                WHERE 
                    (
                        tbl_clientcontacts.Fullname LIKE %s 
                        OR tbl_clientcontacts.Phone LIKE %s 
                        OR tbl_jobcarddetails.RegNo LIKE %s 
                        OR tbl_jobcarddetails.ChassisNo LIKE %s
                        OR tbl_jobcarddetails.MakeAndModel LIKE %s
                    )
                    AND tbl_invoices.Item LIKE %s
                GROUP BY 
                    tbl_clientcontacts.Fullname, 
                    tbl_clientcontacts.Phone, 
                    tbl_jobcarddetails.RegNo, 
                    tbl_jobcarddetails.ChassisNo,
                    tbl_jobcarddetails.ReceivedDate,
                    tbl_jobcarddetails.JobCardRef,
                    tbl_jobcarddetails.MakeAndModel
                ORDER BY 
                    MAX(tbl_jobcarddetails.ReceivedDate) DESC
            """
        like_pattern = f"%{search_term}%"
        part_pattern = f"%{part_name}%"
        cursor.execute(
            query,
            (like_pattern, like_pattern, like_pattern, like_pattern, like_pattern, part_pattern),
        )

        results = cursor.fetchall()

        # Add row numbers using enumerate starting from 1
        return [
            {
                "No": i + 1,
                "Fullname": row[0],
                "Phone": row[1],
                "RegNo": row[2],
                "ChassisNo": row[3],
                "ReceivedDate": row[4],
                "JobCardRef": row[5],
                "MakeAndModel": row[6],
            }
            for i, row in enumerate(results)
        ]


# ***************************************************Job Card Details Section ************************************


@anvil.server.callable()
def getJobCardInstructions(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT ClientInstruction 
                FROM tbl_jobcarddetails 
                WHERE ID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()
        if result:
            raw_instructions = result[0]

            # Step 1: Remove all HTML tags like <div>, <br>, etc.
            text_only = re.sub(r"<[^>]+>", "", raw_instructions)

            # Step 2: Split into lines, strip extra whitespace
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Return as clean multiline text
            return "\n".join(lines)
        else:
            return ""


@anvil.server.callable()
def getJobCardTechNotes(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT Notes 
                FROM tbl_jobcarddetails 
                WHERE ID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()
        if result:
            raw_instructions = result[0]

            # Step 1: Remove all HTML tags like <div>, <br>, etc.
            text_only = re.sub(r"<[^>]+>", "", raw_instructions)

            # Step 2: Split into lines, strip extra whitespace
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Return as clean multiline text
            return "\n".join(lines)
        else:
            return ""


# Return a specific row from job card table
@anvil.server.callable()
def getJobCardRow(IdValue):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                ID, ClientDetails, RegNo, MakeAndModel, ChassisNo,
                EngineCC, EngineNo, EngineCode, Manual, Auto,
                PaintCode, Comp, TPO, JobCardRef,ReceivedDate, DueDate, ExpDate,
                CheckedInBy, Spare, Jack, Brace, Mileage, `Empty`, `Quarter`, Half,
                ThreeQuarter, `Full`, ClientInstruction, Notes, Status  
            FROM tbl_jobcarddetails 
            WHERE ID = %s 
        """
        cursor.execute(query, (IdValue,))
        result = cursor.fetchone()
        if result:
            x = {
                "ID": result[0],
                "ClientDetails": result[1],
                "RegNo": result[2],
                "MakeAndModel": result[3],
                "ChassisNo": result[4],
                "EngineCC": result[5],
                "EngineNo": result[6],
                "EngineCode": result[7],
                "Manual": result[8],
                "Auto": result[9],
                "PaintCode": result[10],
                "Comprehensive": result[11],
                "ThirdParty": result[12],
                "JobCardRef": result[13],
                "ReceivedDate": result[14],
                "DueDate": result[15],
                "ExpDate": result[16],
                "CheckedInBy": result[17],
                "Spare": result[18],
                "Jack": result[19],
                "Brace": result[20],
                "Mileage": result[21],
                "Empty": result[22],
                "Quarter": result[23],
                "Half": result[24],
                "ThreeQuarter": result[25],
                "Full": result[26],
                "ClientInstruction": result[27],
                "Notes": result[28],
                "Status": result[29],
            }

            return x
        else:
            return None


@anvil.server.callable()
def getJobCardRowWithClientID(IdValue):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                ID, ClientDetails, RegNo, MakeAndModel, ChassisNo, EngineCode  
            FROM tbl_jobcarddetails 
            WHERE ID = (
                SELECT MAX(ID)
                FROM tbl_jobcarddetails
                WHERE ClientDetails = %s
            )
        """
        cursor.execute(query, (IdValue,))
        result = cursor.fetchone()
        if result:
            return {
                "ID": result[0],
                "ClientDetails": result[1],
                "RegNo": result[2],
                "MakeAndModel": result[3],
                "ChassisNo": result[4],
                "EngineCode": result[5],
            }
        else:
            return None

# Check for job card ref duplicate
@anvil.server.callable()
def check_job_card_duplicate(JobCardRef):
    """Check if a job card reference already exists in the database."""
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT COUNT(*) FROM tbl_jobcarddetails 
            WHERE JobCardRef = %s 
        """
        cursor.execute(query, (JobCardRef,))
        result = cursor.fetchone()
        return result[0] > 0


# Save new job card data
@anvil.server.callable()
def save_job_card_details(
    technicianDetails,
    ClientDetails,
    JobCardRef,
    ReceivedDate,
    DueDate,
    ExpDate,
    CheckedInBy,
    Ins,
    Comp,
    TPO,
    Spare,
    Jack,
    Brace,
    RegNo,
    MakeAndModel,
    ChassisNo,
    EngineCC,
    Mileage,
    EngineNo,
    EngineCode,
    Manual,
    Auto,
    Empty,
    Quarter,
    Half,
    ThreeQuarter,
    Full,
    PaintCode,
    ClientInstruction,
    Notes,
    IsComplete,
    Status,
):
    _get_current_user()
    with db_cursor() as cursor:
        # Insert into tbl_jobcarddetails
        query = """
            INSERT INTO tbl_jobcarddetails (
                ClientDetails, JobCardRef, ReceivedDate, DueDate, ExpDate, CheckedInBy,
                Ins, Comp, TPO, Spare, Jack, Brace, RegNo, MakeAndModel, ChassisNo,
                EngineCC, Mileage, EngineNo, EngineCode, Manual, Auto, `Empty`, `Quarter`,
                `Half`, `ThreeQuarter`, `Full`, PaintCode, ClientInstruction, Notes, IsComplete, Status
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
        """
        values = (
            ClientDetails,
            JobCardRef,
            ReceivedDate,
            DueDate,
            ExpDate,
            CheckedInBy,
            Ins,
            Comp,
            TPO,
            Spare,
            Jack,
            Brace,
            RegNo,
            MakeAndModel,
            ChassisNo,
            EngineCC,
            Mileage,
            EngineNo,
            EngineCode,
            Manual,
            Auto,
            Empty,
            Quarter,
            Half,
            ThreeQuarter,
            Full,
            PaintCode,
            ClientInstruction,
            Notes,
            IsComplete,
            Status,
        )
        cursor.execute(query, values)

        # Get the ID of the newly inserted job card
        job_card_id = cursor.lastrowid

        # Insert into tbl_pendingassignedjobs using the returned ID
        query2 = """
            INSERT INTO tbl_pendingassignedjobs (
                JobCardRefID, TechnicianID, DateAssigned
            ) 
            VALUES (%s, %s, NOW())
        """
        values2 = (
            job_card_id,  # The ID of the inserted job card
            technicianDetails,  # The technician
        )
        cursor.execute(query2, values2)


# Get job card id to assign in tbl_pendingassignedjobs
@anvil.server.callable()
def getJobCardID(jobCardRef):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID FROM tbl_jobcarddetails 
            WHERE JobCardRef = %s 
        """
        cursor.execute(query, (jobCardRef,))
        result = cursor.fetchone()
        return result[0] if result else None


@anvil.server.callable()
def getJobCardRef(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, JobCardRef
            FROM tbl_jobcarddetails
            WHERE ID = %s
            ORDER BY RegNo ASC
        """

        cursor.execute(query, (valueID,))
        result = cursor.fetchall()
        return [{"ID": r[0], "JobCardRef": r[1]} for r in result]


@anvil.server.callable()
def getQuotationInvoiceName(jobCardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT JobCardRef
            FROM tbl_jobcarddetails
            WHERE ID = %s
        """

        cursor.execute(query, (jobCardID,))
        result = cursor.fetchone()
        return result[0]


@anvil.server.callable()
def get_all_jobcards_by_status():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT ID, MakeAndModel, JobCardRef, RegNo, DueDate, ClientInstruction, Status, ChassisNo
        FROM tbl_jobcarddetails 
        WHERE Status IN (
            'Checked In', 'Create Quote', 'Confirm Quote',
            'In Service', 'Verify Task', 'Issue Invoice', 'Ready for Pickup'
        )
        ORDER BY DueDate DESC
        """
        cursor.execute(query)
        result = cursor.fetchall()

        jobcards = {}
        for r in result:
            # Clean up the ClientInstruction field
            instruction = r[5] or ""
            text_only = re.sub(r"<[^>]+>", "", instruction)
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]
            clean_instruction = "\n".join(lines)

            card = {
                "id": r[0],
                "make": r[1],
                "jobcardref": r[2],
                "plate": r[3],
                "date": r[4],
                "instruction": clean_instruction,
                "status": r[6],
                "chassis": r[7],
            }

            jobcards.setdefault(r[6], []).append(card)

        return jobcards


@anvil.server.callable()
def getTechnicianJobCards(status=None, regNo=None):
    _get_current_user()
    with db_cursor() as cursor:
        # If both filters are missing, return empty list
        if status is None and regNo is None:
            return []

        base_query = """
            SELECT 
                tbl_jobcarddetails.ID, 
                tbl_jobcarddetails.MakeAndModel, 
                tbl_jobcarddetails.JobCardRef, 
                tbl_jobcarddetails.RegNo, 
                tbl_jobcarddetails.DueDate, 
                tbl_jobcarddetails.ClientInstruction, 
                tbl_jobcarddetails.Status, 
                tbl_technicians.Fullname, 
                tbl_jobcarddetails.ChassisNo
            FROM 
                tbl_jobcarddetails 
            INNER JOIN 
                tbl_pendingassignedjobs 
            ON 
                tbl_jobcarddetails.ID = tbl_pendingassignedjobs.JobCardRefID
            INNER JOIN 
                tbl_technicians 
            ON 
                tbl_technicians.ID = tbl_pendingassignedjobs.TechnicianID
        """

        conditions = []
        params = []

        # Add filters dynamically
        if status is not None:
            conditions.append("tbl_jobcarddetails.Status = %s")
            params.append(status)

        if regNo is not None:
            conditions.append("tbl_jobcarddetails.RegNo = %s")
            params.append(regNo)

        # Combine WHERE conditions
        if conditions:
            base_query += " WHERE " + " AND ".join(conditions)

        base_query += " ORDER BY tbl_jobcarddetails.DueDate DESC"

        cursor.execute(base_query, tuple(params))
        result = cursor.fetchall()

        # Clean and return
        jobcards = []
        for r in result:
            instruction = r[5] or ""
            text_only = re.sub(r"<[^>]+>", "", instruction)
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]
            clean_instruction = "\n".join(lines)

            card = {
                "id": r[0],
                "make": r[1],
                "jobcardref": r[2],
                "plate": r[3],
                "date": r[4],
                "instruction": clean_instruction,
                "status": r[6],
                "technician": r[7],
                "chassis": r[8],
            }
            jobcards.append(card)

        return jobcards


@anvil.server.callable()
def getJobCardRefEditDetails():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT
                ID,
                JobCardRef
            FROM
                tbl_jobcarddetails
            WHERE
                (Status = 'Checked In' OR Status IS NULL)
                AND IsComplete = 0
            ORDER BY
                JobCardRef DESC;
        """
        cursor.execute(query)
        result = cursor.fetchall()
        return [(r[1], r[0]) for r in result]  # (Display Text, Value)


@anvil.server.callable()
def getJobCardDetailsWithNameSearch(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, JobCardRef
                FROM
                    tbl_jobcarddetails
                WHERE
                     JobCardRef Like %s
                ORDER BY
                    Mileage DESC;
        """
        like_value = f"%{value}%"
        cursor.execute(query, (like_value,))
        results = cursor.fetchall()

        # Return in format suitable for dropdown: [(display_text, value), ...]
        return [(row[1], row[0]) for row in results]  # (JobCardRef, ID)


@anvil.server.callable()
def getJobCardDetailsWithRefOrFullnameSearch(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_jobcarddetails.ID, 
                tbl_jobcarddetails.JobCardRef,
                tbl_clientcontacts.Fullname
            FROM
                tbl_jobcarddetails
            JOIN
                tbl_clientcontacts
                ON tbl_jobcarddetails.ClientDetails = tbl_clientcontacts.ID
            WHERE
                LOWER(tbl_jobcarddetails.JobCardRef) LIKE LOWER(%s)
                OR LOWER(tbl_clientcontacts.Fullname) LIKE LOWER(%s)
            ORDER BY
                tbl_jobcarddetails.Mileage DESC;
        """
        like_value = f"%{value}%"
        cursor.execute(query, (like_value, like_value))
        results = cursor.fetchall()

        # Return format: ("Fullname - JobCardRef", ID)
        return [(f"{row[2]} - {row[1]}", row[0]) for row in results]


@anvil.server.callable()
def update_job_card_details(
    jobcardID,
    technicianDetails,
    ClientDetails,
    JobCardRef,
    ReceivedDate,
    DueDate,
    ExpDate,
    CheckedInBy,
    Ins,
    Comp,
    TPO,
    Spare,
    Jack,
    Brace,
    RegNo,
    Chassis,
    MakeAndModel,
    EngineCC,
    Mileage,
    EngineNo,
    EngineCode,
    Manual,
    Auto,
    Empty,
    Quarter,
    Half,
    ThreeQuarter,
    Full,
    PaintCode,
    ClientInstruction,
    Notes,
    workDone,
):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_jobcarddetails
            SET
                ClientDetails = %s, JobCardRef = %s, ReceivedDate = %s,DueDate = %s,ExpDate = %s,CheckedInBy = %s,
                TPO = %s,Comp = %s,Spare = %s,Jack = %s, Brace = %s, RegNo = %s, ChassisNo = %s, MakeAndModel = %s,
                EngineCC = %s, Mileage = %s,EngineNo = %s,EngineCode = %s,Manual = %s, Auto = %s, `Empty` = %s,
                `Quarter` = %s,Half = %s, ThreeQuarter = %s, `Full` = %s, PaintCode = %s, ClientInstruction = %s,Notes = %s
            WHERE ID = %s
        """
        values = (
            ClientDetails,
            JobCardRef,
            ReceivedDate,
            DueDate,
            ExpDate,
            CheckedInBy,
            TPO,
            Comp,
            Spare,
            Jack,
            Brace,
            RegNo,
            Chassis,
            MakeAndModel,
            EngineCC,
            Mileage,
            EngineNo,
            EngineCode,
            Manual,
            Auto,
            Empty,
            Quarter,
            Half,
            ThreeQuarter,
            Full,
            PaintCode,
            ClientInstruction,
            Notes,
            jobcardID,
        )

        cursor.execute(query, values)

        query2 = """
            UPDATE tbl_pendingassignedjobs 
            SET 
                TechnicianID=%s, DateAssigned=Now()
            WHERE JobCardRefID = %s
        """
        values2 = (technicianDetails, jobcardID)
        cursor.execute(query2, values2)

        query3 = """
            UPDATE tbl_workdoneinjobcard 
            SET 
                WorkDone=%s
            WHERE JobCardRefID = %s
        """
        values3 = (workDone, jobcardID)
        cursor.execute(query3, values3)


@anvil.server.callable()
def saveSignedJobCardDetails(jobCardId, signature, CreatedAt):
    _get_current_user()
    with db_cursor() as cursor:
        # Read binary content from Anvil media object
        signature_bytes = signature.get_bytes()

        query = """
            INSERT INTO tbl_signedjobcards (AssignedJobID, Signature, CreatedAt)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE
                Signature = VALUES(Signature),
                CreatedAt = VALUES(CreatedAt)
        """
        cursor.execute(query, (jobCardId, signature_bytes, CreatedAt))


@anvil.server.callable()
def getCheckedInJobcards(jobCardRef):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, JobCardRef
            FROM tbl_jobcarddetails
            WHERE Status = 'Checked In' AND IsComplete = 0 and JobCardRef LIKE %s
            ORDER BY JobCardRef DESC;
        """
        cursor.execute(query, (f"%{jobCardRef}%",))
        result = cursor.fetchall()
        return [(r[1], r[0]) for r in result]


@anvil.server.callable()
def saveCancellationReason(jobcardid, reason):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_cancelledjobcards (AssignedJobID, Reason, CreatedAt)
            VALUES (%s, %s, NOW())
        """
        cursor.execute(query, (jobcardid, reason))


@anvil.server.callable()
def getCancelledJobcardReason(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT Reason 
                FROM tbl_cancelledjobcards 
                WHERE AssignedJobID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return ""


@anvil.server.callable()
def getRegNoUsingStatus(status=None, regNo=None):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT RegNo
            FROM tbl_jobcarddetails
            WHERE Status = %s OR RegNo LIKE %s
            GROUP BY RegNo
            ORDER BY RegNo ASC
            
        """
        cursor.execute(query, (status, f"%{regNo}%"))
        result = cursor.fetchall()
        return [r[0] for r in result]


@anvil.server.callable()
def getWorkDoneInJobCard(jobcardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT WorkDone 
            FROM tbl_workdoneinjobcard
            WHERE JobCardRefID = %s
        """
        cursor.execute(query, (jobcardID,))
        result = cursor.fetchone()
        if result:
            raw_workDone = result[0]

            # Step 1: Remove all HTML tags like <div>, <br>, etc.
            text_only = re.sub(r"<[^>]+>", "", raw_workDone)

            # Step 2: Split into lines, strip extra whitespace
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Return as clean multiline text
            return "\n".join(lines)
        else:
            return ""

# ********************************************Tech Notes Document Section ************************************
@anvil.server.callable()
def get_tech_notes_details_by_job_id(jobCardID):
    """
    Returns detailed job card information including:
      - Client name
      - Vehicle details
      - Technician name
      - Cleaned and enumerated technician's notes
    """
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_clientcontacts.Fullname AS ClientName,
                tbl_jobcarddetails.RegNo,
                tbl_jobcarddetails.MakeAndModel,
                tbl_jobcarddetails.EngineCode,
                tbl_jobcarddetails.ChassisNo,
                tbl_jobcarddetails.ReceivedDate,
                tbl_technicians.Fullname AS TechnicianName,
                tbl_jobcarddetails.Notes
            FROM tbl_jobcarddetails
            JOIN tbl_clientcontacts 
                ON tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails
            JOIN tbl_pendingassignedjobs 
                ON tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_technicians 
                ON tbl_technicians.ID = tbl_pendingassignedjobs.TechnicianID
            WHERE tbl_jobcarddetails.ID = %s;

        """
        cursor.execute(query, (jobCardID,))
        result = cursor.fetchone()

        if not result:
            return None

        # Unpack query result
        (
            client_name,
            reg_no,
            make_model,
            engine_code,
            chassis_no,
            received_date,
            technician_name,
            notes,
        ) = result

        # --- Clean and process notes text ---
        if notes:
            # Step 1: Remove HTML tags (<div>, <br>, etc.)
            text_only = re.sub(r"<[^>]+>", "", notes)

            # Step 2: Split into lines, remove empty ones
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Enumerate notes
            numbered_notes = [
                {"No": i + 1, "Notes": line} for i, line in enumerate(lines)
            ]
        else:
            numbered_notes = []


        # --- Return combined data ---
        return {
            "ClientName": client_name,
            "RegNo": reg_no,
            "MakeAndModel": make_model,
            "EngineCode": engine_code,
            "ChassisNo": chassis_no,
            "ReceivedDate": str(received_date),
            "TechnicianName": technician_name,
            "Notes": numbered_notes,
        }
    
@anvil.server.callable()
def fillTechNotesFormData(jobCardID, docType, logo_path: str = os.getenv("LOGO"),font_path: str = os.getenv("FONT_PATH")) -> str:
    _get_current_user()
    if docType == "TechNotes":
        docTitle = "Technician Notes"
        technotes = get_tech_notes_details_by_job_id(jobCardID)

    # === Embed MozillaHeadline font as base64 ===
    font_base64 = ""
    if font_path and os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # === Handle company logo ===
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
        logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" alt="Company Logo" style="width: 100%; height: 100%; border-radius: 2px;">'
    else:
        logo_img_tag = "LOGO"

    # === Generate defects rows — escape DB content to prevent XSS ===
    items_html = ""
    for item in technotes["Notes"]:
        items_html += f"""
        <tr class="item-row">
            <td>{html_module.escape(str(item["No"]))}</td>
            <td>{html_module.escape(str(item["Notes"]))}</td>
        </tr>"""

    # === Full HTML ===
    html_content = f"""<!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>{html_module.escape(str(docTitle))}</title>
            <style>
                /* --- Embed Mozilla Headline font --- */
                @font-face {{
                    font-family: 'Mozilla Headline';
                    src: url(data:font/ttf;base64,{font_base64}) format('truetype');
                }}

                body {{
                    font-family: 'Roboto', 'Noto', Arial, sans-serif;
                    font-size: 14px;
                    line-height: 1.4286;
                    background-color: #fafafa;
                    margin: 0;
                    padding: 16px;
                }}

                .quotation-container {{
                    background-color: white;
                    border-radius: 2px;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                    max-width: 800px;
                    margin: 0 auto;
                    overflow: hidden;
                }}

                .logo-section {{
                    text-align: center;
                    padding: 24px;
                    background-color: white;
                    border-bottom: 1px solid #e0e0e0;
                }}

                .logo-container {{
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    gap: 20px;
                    margin-bottom: 16px;
                }}

                .logo-image {{
                    width: 725px;
                    height: 100px;
                    background: linear-gradient(135deg, #228B22, #90EE90, #FFD700, #FF6347);
                    border-radius: 2px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-size: 24px;
                    color: white;
                    font-weight: 500;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                }}

                .header {{
                    background-color: #000;
                    color: white;
                    text-align: center;
                    padding: 16px 24px;
                    font-size: 16px;
                    font-weight: 300;
                    letter-spacing: .5px;
                    box-shadow: 0 4px 5px 0 rgba(0, 0, 0, 0.14),
                                0 1px 10px 0 rgba(0, 0, 0, 0.12),
                                0 2px 4px -1px rgba(0, 0, 0, 0.2);
                }}

                .detail-row {{
                    display: grid;
                    grid-template-columns: 140px 1fr;
                    column-gap: 8px;
                    margin-bottom: 12px;
                }}

                .detail-label {{
                    font-weight: bold;
                    font-size: 16px;
                    color: rgba(0,0,0,0.87);
                }}

                .detail-value {{
                    font-size: 16px;
                    color: rgba(0,0,0,0.87);
                    text-align: left;
                }}

                .items-table {{
                    border-collapse: collapse;
                    width: 100%;
                    margin: 0 24px 24px 0;
                    background-color: white;
                    border-radius: 2px;
                    overflow: hidden;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                }}

                .items-table th {{
                    background-color: #f5f5f5;
                    border-bottom: 1px solid #e0e0e0;
                    padding: 16px;
                    text-align: left;
                    font-weight: bold;
                    font-size: 14px;
                    color: rgba(0,0,0,0.87);
                    text-transform: uppercase;
                    letter-spacing: .5px;
                }}

                .items-table td {{
                    border-bottom: 1px solid rgba(0,0,0,0.12);
                    padding: 16px;
                    font-size: 14px;
                    color: rgba(0,0,0,0.87);
                }}

                .items-table .item-row:hover {{
                    background-color: rgba(0,0,0,0.04);
                }}

                .notes-title, .notes-list li, #footer div {{
                    font-family: 'Mozilla Headline', Arial, sans-serif;
                }}

                #footer div {{
                    width: 80%;
                    margin: 0 auto;
                    text-align: center;
                    font-size: 12px;
                }}
            </style>
        </head>
        <body>
            <div class="quotation-container">
                <div class="logo-section">
                    <div class="logo-container">
                        <div class="logo-image">
                            {logo_img_tag}
                        </div>
                    </div>
                </div>

                <div class="header">{docTitle.upper()}</div>

                <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
                    <tr>
                        <td style="width: 50%; vertical-align: top; padding-left: 24px; padding-right: 32px;">
                            <div class="detail-row"><span class="detail-label">Customer Name:</span><div><span class="detail-value">{html_module.escape(str(technotes['ClientName']))}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Make And Model:</span><div><span class="detail-value">{html_module.escape(str(technotes['MakeAndModel']))}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Reg No:</span><div><span class="detail-value">{html_module.escape(str(technotes['RegNo']))}</span></div></div>
                        </td>
                        <td style="width: 50%; vertical-align: top; padding-left: 32px;">
                            <div class="detail-row"><span class="detail-label">Engine:</span><div><span class="detail-value">{html_module.escape(str(technotes['EngineCode']))}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Chassis:</span><div><span class="detail-value">{html_module.escape(str(technotes['ChassisNo']))}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Date:</span><div><span class="detail-value">{html_module.escape(str(technotes['ReceivedDate']))}</span></div></div>
                        </td>
                    </tr>
                </table>

                <table class="items-table">
                    <thead><tr><th>No.</th><th>Technician Notes</th></tr></thead>
                    <tbody>{items_html}</tbody>
                </table>


                <footer id="footer">
                    <div><p>Joy Is The Feeling Of Being Looked After By The Best - BMW CENTER For Your BMW.</p></div>
                </footer>
            </div>
        </body>
        </html>
    """

    return html_content


@anvil.server.callable()
def downloadTechNotesPdfForm(jobcardID, docType):
    _get_current_user()
    try:
        result = anvil.server.call("getJobCardRow", jobcardID)
        docName = result["JobCardRef"]

        if docType == "TechNotes":
            fileName = str(docName) + " Technician Notes"

        setting_options = {
            "encoding": "UTF-8",
            "custom-header": [("Accept-Encoding", "gzip")],
            "page-size": "A4",
            "orientation": "Portrait",
            "margin-top": "0.75in",
            "margin-right": "0.75in",
            "margin-bottom": "0.75in",
            "margin-left": "0.75in",
            "no-outline": False,
            "enable-local-file-access": None,
        }

        html_string = fillTechNotesFormData(jobcardID, docType)

        pdfkit.from_string(
            html_string,
            fileName,
            options={**setting_options, "debug-javascript": ""},
            configuration=config,
        )

        media_object = anvil.media.from_file(fileName, "application/pdf", name=fileName)
        return media_object

    except Exception as e:
        print("PDF generation failed:", str(e))
        raise

@anvil.server.callable()
def getTechNotes(jobcardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Notes
            FROM tbl_jobcarddetails
            WHERE ID = %s
        """
        cursor.execute(query, (jobcardID,))
        result = cursor.fetchone()
        text_only = re.sub(r"<[^>]+>", "", result[0])
        return text_only if result else ""  

# ***************************************************Technician Details Section ************************************


@anvil.server.callable()
def saveTecnicianDefectsAndRequestedParts(
    jobcardref, defects, priceddefects, requiredparts, staffID, signature
):
    _get_current_user()
    with db_cursor() as cursor:
        # Read binary content from Anvil media object
        signature_bytes = signature.get_bytes()

        query = """
            INSERT INTO tbl_techniciandefectsandrequestedparts (JobCardRefID, Defects, PricedDefectsList, RequestedParts,PreparedByStaff,Signature)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query, (jobcardref, defects, priceddefects, requiredparts, staffID, signature_bytes)
        )

        
@anvil.server.callable()
def getJobCardDefects(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT Defects 
                FROM tbl_techniciandefectsandrequestedparts 
                WHERE JobCardRefID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()

        if result:
            raw_instructions = result[0]

            # Step 1: Remove all HTML tags like <div>, <br>, etc.
            text_only = re.sub(r"<[^>]+>", "", raw_instructions)

            # Step 2: Split into lines, strip extra whitespace
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Return as clean multiline text
            return "\n".join(lines)
        else:
            return ""

@anvil.server.callable()
def getJobCardPricedDefects(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT PricedDefectsList 
                FROM tbl_techniciandefectsandrequestedparts 
                WHERE JobCardRefID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()

        if result:
            raw_instructions = result[0]

            # Step 1: Remove all HTML tags like <div>, <br>, etc.
            text_only = re.sub(r"<[^>]+>", "", raw_instructions)

            # Step 2: Split into lines, strip extra whitespace
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Return as clean multiline text
            return "\n".join(lines)
        else:
            return ""


@anvil.server.callable()
def deleteDefects(JobCardID):
    _get_current_user()
    with db_cursor() as cursor:
        # Delete existing records in tbl_techniciandefectsandrequestedparts
        query_delete = """
            DELETE FROM tbl_techniciandefectsandrequestedparts
            WHERE JobCardRefID = %s
        """
        cursor.execute(query_delete, (JobCardID,))

@anvil.server.callable()
def getRequestedParts(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT RequestedParts 
                FROM tbl_techniciandefectsandrequestedparts 
                WHERE JobCardRefID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()
        if result:
            raw_instructions = result[0]

            # Step 1: Remove all HTML tags like <div>, <br>, etc.
            text_only = re.sub(r"<[^>]+>", "", raw_instructions)

            # Step 2: Split into lines, strip extra whitespace
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Return as clean multiline text
            return "\n".join(lines)
        else:
            return ""


@anvil.server.callable()
def getDefectsStaffAndSignature(jobcardref_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT PreparedByStaff, Signature 
            FROM tbl_techniciandefectsandrequestedparts
            WHERE JobCardRefID = %s
            ORDER BY ID DESC
        """
        cursor.execute(query, (jobcardref_id,))
        results = cursor.fetchall()

        # Convert signatures to BlobMedia for Anvil use
        formatted_results = []
        for row in results:
            prepared_by_id = row[0]
            signature_bytes = row[1]

            signature_media = None
            if signature_bytes:
                signature_media = anvil.BlobMedia("image/png", signature_bytes)

                formatted_results.append(
                    {"PreparedByStaff": prepared_by_id, "Signature": signature_media}
                )

        return formatted_results


@anvil.server.callable()
def updateJobCardStatus(jobcardid, status):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                        UPDATE tbl_jobcarddetails 
                        SET Status = %s 
                        WHERE ID = %s;
                    """
        cursor.execute(query, (status, jobcardid))


@anvil.server.callable()
def saveWorkDoneInJobCard(jobcardID, workDone):
    _get_current_user()
    with db_cursor() as cursor:
        # Insert work done into tbl_workdoneinjobcard
        query = """
                            INSERT INTO tbl_workdoneinjobcard (JobCardRefID, WorkDone) 
                            VALUES (%s, %s);
                        """
        cursor.execute(query, (jobcardID, workDone))
        #New technician portal notification
        publish_technician_portal_notification(jobcardID, "Jobcard transitioned to Verify Task from Technician Portal")
            


@anvil.server.callable()
def updateDefectsList(
    jobcardID, instructions, notes, defects, priceddefects, parts, staffId, signature
):
    _get_current_user()
    with db_cursor() as cursor:
        
        query1 = """
            UPDATE tbl_jobcarddetails
            SET ClientInstruction = %s, Notes = %s
            WHERE ID = %s
        """
        cursor.execute(query1, (instructions, notes, jobcardID))

        if signature:
            # Read binary content from Anvil media object
            signature_bytes = signature.get_bytes()

            query2 = """
                UPDATE tbl_techniciandefectsandrequestedparts
                SET Defects = %s, PricedDefectsList=%s, RequestedParts = %s, PreparedByStaff = %s, Signature = %s
                WHERE JobCardRefID = %s
            """
            cursor.execute(
                query2, (defects, priceddefects, parts, staffId, signature_bytes, jobcardID)
            )

        else:
            query3 = """
                UPDATE tbl_techniciandefectsandrequestedparts
                SET Defects = %s, PricedDefectsList=%s, RequestedParts = %s, PreparedByStaff = %s
                WHERE JobCardRefID = %s
            """
            cursor.execute(query3, (defects, priceddefects, parts, staffId, jobcardID))

@anvil.server.callable()
def transitionCheckedInToCreateQuote():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                UPDATE 
                    tbl_jobcarddetails AS jobs
                INNER JOIN 
                    tbl_techniciandefectsandrequestedparts AS defects 
                ON 
                    jobs.ID = defects.JobCardRefID
                SET 
                    jobs.Status = 'Create Quote'
                WHERE 
                    jobs.Status = 'Checked In';
            """
        cursor.execute(query)

@anvil.server.callable()
def updateBlankDefectsAndRequestedParts():
    _get_current_user()
    with db_cursor() as cursor:
        query1 = """
                UPDATE tbl_techniciandefectsandrequestedparts SET Defects="None" WHERE Defects ="";
            """
        cursor.execute(query1)
        query2 = """
                UPDATE tbl_techniciandefectsandrequestedparts SET PricedDefectsList="None" WHERE PricedDefectsList ="";
            """
        cursor.execute(query2)
        query3 = """
                UPDATE tbl_techniciandefectsandrequestedparts SET RequestedParts="None" WHERE RequestedParts ="";
            """
        cursor.execute(query3)


@anvil.server.callable()
def getDefectsList(jobcardID):
    """
    Joins job card details with technician defects and returns 
    data including a base64 encoded signature.
    """
    _get_current_user()
    query = """
        SELECT
            j.ClientInstruction, 
            j.Notes, 
            t.Defects,
            t.PricedDefectsList,
            t.TechnicianPortalRequestedParts,
            t.RequestedParts, 
            t.PreparedByStaff, 
            t.Signature
        FROM tbl_jobcarddetails j
        JOIN tbl_techniciandefectsandrequestedparts t ON t.JobCardRefID = j.ID
        WHERE j.ID = %s
        ORDER BY t.ID DESC
    """
    
    with db_cursor() as cursor:
        cursor.execute(query, (jobcardID,))
        results = cursor.fetchall()

    formatted_data = []
    
    for row in results:
        instruction, notes, defects, pricedDefects, portalParts, parts, staff, sig_bytes = row
        
        signature_media = None
        if sig_bytes:
            signature_media = anvil.BlobMedia("image/png", sig_bytes)

        formatted_data.append({
            "Instruction": instruction,
            "Notes": notes,
            "Defects": defects,
            "PricedDefects": pricedDefects,
            "TechnicianPortalRequestedParts": portalParts if portalParts else "None",
            "RequestedParts": parts,
            "PreparedByStaff": staff,
            "Signature": signature_media  # This is now an anvil.BlobMedia object
        })

    return formatted_data

# ***************************************************Car Parts Details Section ************************************
@anvil.server.callable()
def getCarPartNames():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Name 
            FROM tbl_carpartnames
            GROUP BY Name 
            ORDER BY Name ASC
        """
        cursor.execute(query)
        result = cursor.fetchall()
        return [{"Name": r[0]} for r in result]


@anvil.server.callable()
def getCarPartNumber(name):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT ID, PartNo 
        FROM tbl_carpartnames 
        WHERE Name = %s
        ORDER BY PartNo ASC
        """
        cursor.execute(query, (name,))
        result = cursor.fetchall()
        return [{"ID": r[0], "PartNo": r[1]} for r in result]


@anvil.server.callable()
def getCarPartNameAndNumber(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT ID, PartNo, Name 
        FROM tbl_carpartnames 
        WHERE PartNo LIKE %s OR Name LIKE %s
        ORDER BY Name ASC
        """
        like_value = f"%{value}%"
        cursor.execute(query, (like_value, like_value))
        result = cursor.fetchall()
        # Return tuples: (Label shown, Value stored)
        return [(f"{r[1]} - {r[2]}", r[0]) for r in result]


@anvil.server.callable()
def getCarPartNamesWithId(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Name 
            FROM tbl_carpartnames
            WHERE ID = %s
        """
        cursor.execute(query, (valueID,))
        result = cursor.fetchall()
        return [{"Name": r[0]} for r in result]


@anvil.server.callable()
def getCarPartIDWithNumber(valueNumber):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID 
            FROM tbl_carpartnames
            WHERE PartNo = %s
        """
        cursor.execute(query, (valueNumber,))
        result = cursor.fetchall()
        # Check if a result was returned
        if result:
            return result[0]  # Return the first (and only) column from the row
        else:
            return None  # Return None if no part was found


@anvil.server.callable()
def getCarPartNumberWithID(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT ID, PartNo 
        FROM tbl_carpartnames 
        WHERE ID = %s
        ORDER BY PartNo ASC
        """
        cursor.execute(query, (valueID,))
        result = cursor.fetchall()
        return [{"ID": r[0], "PartNo": r[1]} for r in result]


@anvil.server.callable()
def getSellingPrice(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Amount 
            FROM tbl_partssellingprice
            WHERE CarPartsNamesID = %s
        """
        cursor.execute(query, (id,))
        result = cursor.fetchone()
        return result

@anvil.server.callable()
def insertMissingSellingPrice(part_id, amount):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_partssellingprice 
                (CarPartsNamesID, Amount, SaleDiscount, SetPriceDate)
            VALUES 
                (%s, %s, 0, CURDATE())
        """
        cursor.execute(query, (part_id, amount))

# ***************************************************Quotation Details Section ************************************

@anvil.server.callable()
def getDuplicateQuotationEntries(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                COUNT(*) as Entries
            FROM 
                tbl_techniciandefectsandrequestedparts 
            WHERE 
                tbl_techniciandefectsandrequestedparts.JobCardRefID=%s;
        """
        cursor.execute(query, (valueID,))
        result = cursor.fetchone()
        return result[0] if result else 0  # Return the count of entries or 0 if no result
    
@anvil.server.callable()
def displayDuplicateQuotationEntries(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_techniciandefectsandrequestedparts.ID,
                tbl_jobcarddetails.JobCardRef,
                tbl_techniciandefectsandrequestedparts.TechnicianPortalRequestedParts,
                tbl_techniciandefectsandrequestedparts.Defects,
                tbl_techniciandefectsandrequestedparts.PricedDefectsList,
                tbl_techniciandefectsandrequestedparts.RequestedParts,
                tbl_techniciandefectsandrequestedparts.PreparedByStaff
            FROM
                tbl_techniciandefectsandrequestedparts
            INNER JOIN tbl_jobcarddetails
                ON tbl_techniciandefectsandrequestedparts.JobCardRefID = tbl_jobcarddetails.ID
            WHERE
                tbl_techniciandefectsandrequestedparts.JobCardRefID = %s;
        """
        cursor.execute(query, (valueID,))
        results = cursor.fetchall()
        return [
            {
                "ID": r[0],
                "JobCardRef": r[1],
                "TechnicianRequestedParts": r[2],
                "Defects": r[3],
                "PricedDefects": r[4],
                "RequestedParts": str(r[5]),
                "CreatedBy": r[6],
            }
            for r in results
        ]

@anvil.server.callable()
def deleteDuplicateQuotationEntry(entryID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            DELETE FROM tbl_techniciandefectsandrequestedparts
            WHERE ID = %s
        """
        cursor.execute(query, (entryID,))

@anvil.server.callable()
def getQuotationJobCardDetails():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT DISTINCT tbl_jobcarddetails.JobCardRef, tbl_quotation.AssignedJobID
            FROM tbl_quotation 
            INNER JOIN tbl_jobcarddetails 
            ON tbl_quotation.AssignedJobID = tbl_jobcarddetails.ID
            ORDER BY tbl_jobcarddetails.JobCardRef;
        """

        cursor.execute(query)
        result = cursor.fetchall()
        return [{"JobCardRef": r[0], "ID": r[1]} for r in result]


@anvil.server.callable()
def saveQuotationPartsAndServices(
    assignedDate, jobCardID, name, number, quantity, amount
):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_quotation (Date, AssignedJobID, Item, Part_No, QuantityIssued, Amount)
            VALUES (%s,%s,%s,%s, %s, %s)
        """
        cursor.execute(query, (assignedDate, jobCardID, name, number, quantity, amount))


@anvil.server.callable()
def get_quote_details_by_job_id(job_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT
            tbl_clientcontacts.Fullname,
            tbl_jobcarddetails.MakeAndModel,
            tbl_jobcarddetails.RegNo,
            tbl_quotation.Date,
            tbl_jobcarddetails.ChassisNo,
            tbl_jobcarddetails.EngineCode,
            tbl_jobcarddetails.Mileage,
            tbl_quotation.Item,
            tbl_quotation.QuantityIssued,
            tbl_quotation.Amount,
            tbl_quotation.AssignedJobID
        FROM
            (tbl_jobcarddetails
            INNER JOIN tbl_clientcontacts
            ON tbl_jobcarddetails.ClientDetails = tbl_clientcontacts.ID)
        INNER JOIN tbl_quotation
            ON tbl_jobcarddetails.ID = tbl_quotation.AssignedJobID
        WHERE tbl_quotation.AssignedJobID = %s
        """
        cursor.execute(query, (job_id,))
        rows = cursor.fetchall()

        result = []
        for r in rows:
            # Helper function to ensure we get a float from various types
            def safe_float_convert(value):
                if value is None:
                    return None  # Or 0.0, depending on desired behavior for NULLs
                if isinstance(value, (int, float, decimal.Decimal)):
                    return float(value)
                # If it's a string, try to clean it and convert
                if isinstance(value, str):
                    try:
                        return float(value.replace(",", ""))
                    except ValueError:
                        # Handle cases where string cannot be converted (e.g., non-numeric string)
                        print(f"Warning: Could not convert string '{value}' to float.")
                        return None  # Or raise an error
                return None  # Default if type is unexpected

            quantity_issued_val = safe_float_convert(r[8])  # QuantityIssued
            amount_val = safe_float_convert(r[9])  # Amount

            # Ensure 'QuantityIssued' for the dictionary can be "" if originally None or non-numeric string
            # If your front-end expects an empty string, keep this logic
            display_quantity_issued = (
                ""
                if r[8] is None or not isinstance(r[8], (int, float, decimal.Decimal))
                else quantity_issued_val
            )

            # Calculate Total - ensuring we use the float versions for calculation
            total_calc = None
            if quantity_issued_val is None:  # If QuantityIssued was NULL/None from DB
                total_calc = amount_val
            elif amount_val is not None:  # Ensure Amount is not None for multiplication
                total_calc = round(quantity_issued_val * amount_val, 2)
            # If amount_val is None and quantity_issued_val is not None, total_calc would remain None

            result.append(
                {
                    "Fullname": r[0],
                    "MakeAndModel": r[1],
                    "RegNo": r[2],
                    "Date": r[3],
                    "ChassisNo": r[4],
                    "EngineCode": r[5],
                    "Mileage": r[6],
                    "Item": r[7],
                    "QuantityIssued": display_quantity_issued,  # Use the potentially "" value for display
                    "Amount": amount_val,  # This is now always a float or None
                    "AssignedJobID": r[10],
                    "Total": total_calc,  # This is now always a float or None
                }
            )

        return result


@anvil.server.callable()
def get_quote_confirmation_details_by_job_id(job_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT
            tbl_clientcontacts.Fullname,
            tbl_jobcarddetails.MakeAndModel,
            tbl_jobcarddetails.RegNo,
            tbl_quotationpartsandservicesfeedback.Date,
            tbl_jobcarddetails.ChassisNo,
            tbl_jobcarddetails.EngineCode,
            tbl_jobcarddetails.Mileage,
            tbl_quotationpartsandservicesfeedback.Item,
            tbl_quotationpartsandservicesfeedback.QuantityIssued,
            tbl_quotationpartsandservicesfeedback.Amount,
            tbl_quotationpartsandservicesfeedback.AssignedJobID
        FROM
            (tbl_jobcarddetails
            INNER JOIN tbl_clientcontacts
            ON tbl_jobcarddetails.ClientDetails = tbl_clientcontacts.ID)
        INNER JOIN tbl_quotationpartsandservicesfeedback
            ON tbl_jobcarddetails.ID = tbl_quotationpartsandservicesfeedback.AssignedJobID
        WHERE tbl_quotationpartsandservicesfeedback.AssignedJobID = %s
        """
        cursor.execute(query, (job_id,))
        rows = cursor.fetchall()

        result = []
        for r in rows:
            # Helper function to ensure we get a float from various types
            def safe_float_convert(value):
                if value is None:
                    return None  # Or 0.0, depending on desired behavior for NULLs
                if isinstance(value, (int, float, decimal.Decimal)):
                    return float(value)
                # If it's a string, try to clean it and convert
                if isinstance(value, str):
                    try:
                        return float(value.replace(",", ""))
                    except ValueError:
                        # Handle cases where string cannot be converted (e.g., non-numeric string)
                        print(f"Warning: Could not convert string '{value}' to float.")
                        return None  # Or raise an error
                return None  # Default if type is unexpected

            quantity_issued_val = safe_float_convert(r[8])  # QuantityIssued
            amount_val = safe_float_convert(r[9])  # Amount

            # Ensure 'QuantityIssued' for the dictionary can be "" if originally None or non-numeric string
            # If your front-end expects an empty string, keep this logic
            display_quantity_issued = (
                ""
                if r[8] is None or not isinstance(r[8], (int, float, decimal.Decimal))
                else quantity_issued_val
            )

            # Calculate Total - ensuring we use the float versions for calculation
            total_calc = None
            if quantity_issued_val is None:  # If QuantityIssued was NULL/None from DB
                total_calc = amount_val
            elif amount_val is not None:  # Ensure Amount is not None for multiplication
                total_calc = round(quantity_issued_val * amount_val, 2)
            # If amount_val is None and quantity_issued_val is not None, total_calc would remain None

            result.append(
                {
                    "Fullname": r[0],
                    "MakeAndModel": r[1],
                    "RegNo": r[2],
                    "Date": r[3],
                    "ChassisNo": r[4],
                    "EngineCode": r[5],
                    "Mileage": r[6],
                    "Item": r[7],
                    "QuantityIssued": display_quantity_issued,  # Use the potentially "" value for display
                    "Amount": amount_val,  # This is now always a float or None
                    "AssignedJobID": r[10],
                    "Total": total_calc,  # This is now always a float or None
                }
            )

        return result


@anvil.server.callable()
def get_defects_list_details_by_job_id(jobCardID):
    """
    Returns detailed job card information including:
      - Client name
      - Vehicle details
      - Technician name
      - Cleaned and enumerated defects list
    """
    _get_current_user()
    with db_cursor() as cursor:
        query1 = """
            SELECT 
                tbl_clientcontacts.Fullname AS ClientName,
                tbl_jobcarddetails.RegNo,
                tbl_jobcarddetails.MakeAndModel,
                tbl_jobcarddetails.EngineCode,
                tbl_jobcarddetails.ChassisNo,
                tbl_jobcarddetails.ReceivedDate,
                tbl_technicians.Fullname AS TechnicianName,
                tbl_techniciandefectsandrequestedparts.Defects,
                tbl_checkstaff.Staff AS PreparedByStaff,
                tbl_techniciandefectsandrequestedparts.Signature
            FROM tbl_jobcarddetails
            JOIN tbl_clientcontacts 
                ON tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails
            JOIN tbl_pendingassignedjobs 
                ON tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_technicians 
                ON tbl_technicians.ID = tbl_pendingassignedjobs.TechnicianID
            JOIN tbl_techniciandefectsandrequestedparts 
                ON tbl_techniciandefectsandrequestedparts.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_checkstaff
                ON tbl_checkstaff.Staff = tbl_techniciandefectsandrequestedparts.PreparedByStaff
            WHERE tbl_jobcarddetails.ID = %s;

        """
        cursor.execute(query1, (jobCardID,))
        result1 = cursor.fetchone()

        query2 = """
            SELECT 
                tbl_clientcontacts.Fullname AS ClientName,
                tbl_jobcarddetails.RegNo,
                tbl_jobcarddetails.MakeAndModel,
                tbl_jobcarddetails.EngineCode,
                tbl_jobcarddetails.ChassisNo,
                tbl_jobcarddetails.ReceivedDate,
                tbl_technicians.Fullname AS TechnicianName,
                tbl_techniciandefectsandrequestedparts.Defects,
                tbl_technicians.Fullname AS PreparedByStaff,
                tbl_techniciandefectsandrequestedparts.Signature
            FROM tbl_jobcarddetails
            JOIN tbl_clientcontacts 
                ON tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails
            JOIN tbl_pendingassignedjobs 
                ON tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_techniciandefectsandrequestedparts 
                ON tbl_techniciandefectsandrequestedparts.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_technicians
                ON tbl_technicians.Fullname = tbl_techniciandefectsandrequestedparts.PreparedByStaff
            WHERE tbl_jobcarddetails.ID = %s;
            """
        cursor.execute(query2, (jobCardID,))
        result2 = cursor.fetchone()
        if result1 is None and result2 is None:
            return None  # No data found for the given JobCardID
        elif result1 is not None:   
            result = result1
        else:
            result = result2    

        # Unpack query result
        (
            client_name,
            reg_no,
            make_model,
            engine_code,
            chassis_no,
            received_date,
            technician_name,
            raw_defects,
            staff,
            signature,
        ) = result

        # --- Clean and process defects text ---
        if raw_defects:
            # Step 1: Remove HTML tags (<div>, <br>, etc.)
            text_only = re.sub(r"<[^>]+>", "", raw_defects)

            # Step 2: Split into lines, remove empty ones
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Enumerate defects
            numbered_defects = [
                {"No": i + 1, "Defect": line} for i, line in enumerate(lines)
            ]
        else:
            numbered_defects = []

        # If Signature is stored as BLOB → convert to base64
        if isinstance(signature, (bytes, bytearray)):
            signature_b64 = base64.b64encode(signature).decode("utf-8")

        # --- Return combined data ---
        return {
            "ClientName": client_name,
            "RegNo": reg_no,
            "MakeAndModel": make_model,
            "EngineCode": engine_code,
            "ChassisNo": chassis_no,
            "ReceivedDate": str(received_date),
            "TechnicianName": technician_name,
            "Defects": numbered_defects,
            "Staff": staff,
            "Signature": signature_b64,
        }


@anvil.server.callable()
def get_priced_defects_list_details_by_job_id(jobCardID):
    """
    Returns detailed job card information including:
      - Client name
      - Vehicle details
      - Technician name
      - Cleaned and enumerated priced defects list
    """
    _get_current_user()
    with db_cursor() as cursor:
        query1 = """
            SELECT 
                tbl_clientcontacts.Fullname AS ClientName,
                tbl_jobcarddetails.RegNo,
                tbl_jobcarddetails.MakeAndModel,
                tbl_jobcarddetails.EngineCode,
                tbl_jobcarddetails.ChassisNo,
                tbl_jobcarddetails.ReceivedDate,
                tbl_technicians.Fullname AS TechnicianName,
                tbl_techniciandefectsandrequestedparts.PricedDefectsList,
                tbl_checkstaff.Staff AS PreparedByStaff,
                tbl_techniciandefectsandrequestedparts.Signature
            FROM tbl_jobcarddetails
            JOIN tbl_clientcontacts 
                ON tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails
            JOIN tbl_pendingassignedjobs 
                ON tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_technicians 
                ON tbl_technicians.ID = tbl_pendingassignedjobs.TechnicianID
            JOIN tbl_techniciandefectsandrequestedparts 
                ON tbl_techniciandefectsandrequestedparts.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_checkstaff
                ON tbl_checkstaff.Staff = tbl_techniciandefectsandrequestedparts.PreparedByStaff
            WHERE tbl_jobcarddetails.ID = %s;

        """
        cursor.execute(query1, (jobCardID,))
        result1 = cursor.fetchone()

        query2 = """
            SELECT 
                tbl_clientcontacts.Fullname AS ClientName,
                tbl_jobcarddetails.RegNo,
                tbl_jobcarddetails.MakeAndModel,
                tbl_jobcarddetails.EngineCode,
                tbl_jobcarddetails.ChassisNo,
                tbl_jobcarddetails.ReceivedDate,
                tbl_technicians.Fullname AS TechnicianName,
                tbl_techniciandefectsandrequestedparts.PricedDefectsList,
                tbl_technicians.Fullname AS PreparedByStaff,
                tbl_techniciandefectsandrequestedparts.Signature
            FROM tbl_jobcarddetails
            JOIN tbl_clientcontacts 
                ON tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails
            JOIN tbl_pendingassignedjobs 
                ON tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_techniciandefectsandrequestedparts 
                ON tbl_techniciandefectsandrequestedparts.JobCardRefID = tbl_jobcarddetails.ID
            JOIN tbl_technicians
                ON tbl_technicians.Fullname = tbl_techniciandefectsandrequestedparts.PreparedByStaff
            WHERE tbl_jobcarddetails.ID = %s;
            """
        cursor.execute(query2, (jobCardID,))
        result2 = cursor.fetchone()
        if result1 is None and result2 is None:
            return None  # No data found for the given JobCardID
        elif result1 is not None:   
            result = result1
        else:
            result = result2    

        # Unpack query result
        (
            client_name,
            reg_no,
            make_model,
            engine_code,
            chassis_no,
            received_date,
            technician_name,
            raw_defects,
            staff,
            signature,
        ) = result

        # --- Clean and process defects text ---
        if raw_defects:
            # Step 1: Remove HTML tags (<div>, <br>, etc.)
            text_only = re.sub(r"<[^>]+>", "", raw_defects)

            # Step 2: Split into lines, remove empty ones
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]

            # Step 3: Enumerate defects
            numbered_defects = [
                {"No": i + 1, "Defect": line} for i, line in enumerate(lines)
            ]
        else:
            numbered_defects = []

        # If Signature is stored as BLOB → convert to base64
        if isinstance(signature, (bytes, bytearray)):
            signature_b64 = base64.b64encode(signature).decode("utf-8")

        # --- Return combined data ---
        return {
            "ClientName": client_name,
            "RegNo": reg_no,
            "MakeAndModel": make_model,
            "EngineCode": engine_code,
            "ChassisNo": chassis_no,
            "ReceivedDate": str(received_date),
            "TechnicianName": technician_name,
            "Defects": numbered_defects,
            "Staff": staff,
            "Signature": signature_b64,
        }

@anvil.server.callable()
def fillQuotationInvoiceData(
    jobCardID, docType, logo_path: str = os.getenv("LOGO"), font_path: str = os.getenv("FONT")
) -> str:
    _get_current_user()
    if docType == "Quotation":
        docTitle = "Quotation"
        vehicledetails = get_quote_details_by_job_id(jobCardID)
    elif docType == "InterimQuotation":
        docTitle = "Interim Quotation"
        vehicledetails = get_quote_details_by_job_id(jobCardID)
    elif docType == "Invoice":
        docTitle = "Invoice"
        vehicledetails = get_invoice_details_by_job_id(jobCardID)
    elif docType == "Confirm Quotation":
        docTitle = "Confirm Quotation"
        vehicledetails = get_quote_confirmation_details_by_job_id(jobCardID)

    # === Embed MozillaHeadline font as base64 ===
    font_base64 = ""
    if font_path and os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # === Handle company logo ===
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
        logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" alt="Company Logo" style="width: 100%; height: 100%; border-radius: 2px;">'
    else:
        logo_img_tag = "LOGO"

    # Calculate grand total
    sub_total = sum(float(item["Total"]) for item in vehicledetails)
    for item in vehicledetails:
        if item["Item"] == "Previous Balance":
            previous_balance = item["Amount"]
            sub_total = (
                sub_total - item["Amount"]
            )  # Get sub total without previous balance
            footer_total_details = f"""
                        <tr class="total-row">
                            <td colspan="4" style="text-align: right; font-weight: 500;">Sub Total</td>
                            <td style="font-weight: 500;">{sub_total:,.2f}</td>                    
                        </tr>
                        <tr class="total-row">
                            <td colspan="4" style="text-align: right; font-weight: 500;">Previous Balance</td>
                            <td style="font-weight: 500;">{previous_balance:,.2f}</td>
                        </tr>       
                """

        else:
            previous_balance = 0
            footer_total_details = ""

    grand_total = sub_total + float(previous_balance)

    # Generate items table rows
    items_html = ""
    counter = 0
    for item in vehicledetails:
        counter = counter + 1
        if item["Amount"] == 0:  # Implies To Be Confirmed
            textAmount = "TO BE CONFIRMED"
        else:
            textAmount = f"{item['Amount']:,.2f}"
        if item["Total"] == 0:  # Implies To Be Confirmed
            textTotal = "TO BE CONFIRMED"
        else:
            textTotal = f"{item['Total']:,.2f}"

        # Do not display Previous balance in the table
        if item["Item"] != "Previous Balance":

            items_html += f"""
                    <tr class="item-row">
                        <td>{counter}</td>
                        <td>{html_module.escape(str(item['Item']))}</td>
                        <td>{html_module.escape(str(item['QuantityIssued']))}</td>
                        <td>{textAmount}</td>
                        <td>{textTotal}</td>
                    </tr>"""

    if docType in ("Quotation", "InterimQuotation"):
        quotationNotes = """
                    <div class="notes-section">
                        <div class="notes-title">NOTE: THE ABOVE ESTIMATE IS SUBJECT TO REVIEW DUE TO:</div>
                        <ol class="notes-list">
                            <li>Price change at the time of actual repair</li>
                            <li>Further damages found during repairs</li>
                            <li>100% Deposit on imported parts</li>
                            <li>70% deposit on local parts on commencement</li>
                        </ol>
                    </div>
        """
    else:
        quotationNotes = """
                    <div class="notes-section">
                        <div class="notes-title">NOTES: </div>
                        <ol class="notes-list">
                            <li>Thank you for choosing BMW CENTER LIMITED</li>
                            <li>Cheque Address to: BMW CENTER LIMITED</li>
                         </ol>
                    </div>
            """

        # Complete HTML template with fixed structure
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{docTitle}</title>
    <style>
        /* --- Embed Mozilla Headline font --- */
        @font-face {{
            font-family: 'Mozilla Headline';
            src: url(data:font/ttf;base64,{font_base64}) format('truetype');
        }}
        body {{
            font-family: Roboto, Noto, Arial, sans-serif;
            font-size: 14px;
            line-height: 1.4286;
            background-color: #fafafa;
            margin: 0;
            padding: 16px;
        }}

        .quotation-container {{
            background-color: white;
            border-radius: 2px;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
            max-width: 800px;
            margin: 0 auto;
            overflow: hidden;
        }}

        .logo-section {{
            text-align: center;
            padding: 24px;
            background-color: white;
            border-bottom: 1px solid #e0e0e0;
        }}

        .logo-container {{
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 20px;
            margin-bottom: 16px;
        }}

        .logo-image {{
            width: 725px;
            height: 100px;
            background: linear-gradient(135deg, #228B22, #90EE90, #FFD700, #FF6347);
            border-radius: 2px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
            color: white;
            font-weight: 500;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}

        .header {{
            background-color: #000;
            color: white;
            text-align: center;
            padding: 16px 24px;
            font-size: 16px;
            font-weight: 300;
            letter-spacing: .5px;
            box-shadow: 0 4px 5px 0 rgba(0, 0, 0, 0.14),
                        0 1px 10px 0 rgba(0, 0, 0, 0.12),
                        0 2px 4px -1px rgba(0, 0, 0, 0.2);
        }}

        .detail-row {{
            display: grid;
            grid-template-columns: 140px 1fr; /* label column, value column */
            column-gap: 8px;
            margin-bottom: 12px;
        }}

        .detail-label {{
                font-weight: bold;
                font-size: 16px;
                color: rgba(0,0,0,0.87);
                
        }}

        .detail-value {{
            font-size: 16px;
            color: rgba(0,0,0,0.87);
            text-align: left;
            
        }}

        .items-table {{
            border-collapse: collapse;
            width: 100%;
            margin: 0 24px 24px 0;
            background-color: white;
            border-radius: 2px;
            overflow: hidden;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}

        .items-table th {{
            background-color: #f5f5f5;
            border-bottom: 1px solid #e0e0e0;
            padding: 16px;
            text-align: left;
            font-weight: bold;
            font-size: 14px;
            color: rgba(0,0,0,0.87);
            text-transform: uppercase;
            letter-spacing: .5px;
           
        }}

        .items-table td {{
            border-bottom: 1px solid rgba(0,0,0,0.12);
            padding: 16px;
            font-size: 14px;
            color: rgba(0,0,0,0.87);
            
        }}

        .items-table .item-row:hover {{
            background-color: rgba(0,0,0,0.04);
        }}

        .total-row {{
            background-color: #000 !important;
            color: white !important;
        }}

        .total-row td {{
            border-bottom: none !important;
            font-weight: 300;
            font-size: 16px;
            color: white !important;
            padding: 16px;
            
        }}

        .notes-section {{
            padding: 24px;
            background-color: #f5f5f5;
            margin-top: 16px;
        }}

        .notes-title {{
            margin-bottom: 16px;
            font-weight: 500;
            font-size: 16px;
            color: rgba(0,0,0,0.87);
            font-family: 'Mozilla Headline';
        }}

        .notes-list {{
            margin: 0;
            padding-left: 24px;
            color: rgba(0,0,0,0.74);
        }}

        .notes-list li {{
            margin-bottom: 8px;
            line-height: 1.5;
            font-family: 'Mozilla Headline';
        }}
        #footer  div {{
                width: 80%;
                margin: 0 auto;
                text-align: center;
                font-size: 12px;
                font-family: 'Mozilla Headline';
            }}
    </style>
    
</head>
<body>
    <div class="quotation-container">
        <div class="logo-section">
            <div class="logo-container">
                <div class="logo-image">
                    {logo_img_tag}
                </div>
            </div>
        </div>

        <div class="header">
            {docTitle.upper()}
        </div>

        <!-- UPDATED: Replaced details-section div with table layout -->
        <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
            <tr>
                <!-- Left Column -->
                <td style="width: 50%; vertical-align: top; padding-left: 24px; padding-right: 32px;">
                    <div class="detail-row">
                        <span class="detail-label">Customer Name:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['Fullname']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Make And Model:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['MakeAndModel']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Reg No:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['RegNo']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Date:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['Date']}</span>
                        </div>
                    </div>
                </td>

                <!-- Right Column -->
                <td style="width: 50%; vertical-align: top; padding-left: 32px;">
                    <div class="detail-row">
                        <span class="detail-label">Chassis:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['ChassisNo']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Engine:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['EngineCode']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Mileage:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['Mileage']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">&nbsp;</span>
                        <span class="detail-value">&nbsp;</span>
                    </div>
                </td>
            </tr>
        </table>
        <!-- END UPDATED -->

        <table class="items-table">
            <thead>
                <tr>
                    <th>No.</th>
                    <th>Item</th>
                    <th>Quantity</th>
                    <th>Amount (Kshs)</th>
                    <th>Total (Kshs)</th>
                </tr>
            </thead>
            <tbody>
                {items_html}
                {footer_total_details}
                
                <tr class="total-row">
                    <td colspan="4" style="text-align: right; font-weight: 500;">Grand Total</td>
                    <td style="font-weight: 500;">{grand_total:,.2f}</td>
                </tr>
                
            </tbody>
        </table>
    {quotationNotes} 
   <footer id="footer">
        <div> 
            <p>Joy Is The Feeling Of Being Looked After By The Best - BMW CENTER For Your BMW.</p>
        </div>
    </footer>
    </div>
    </body>
    </html>"""

    return html_content


@anvil.server.callable()
def createQuotationInvoicePdf(jobCardID, docType):
    _get_current_user()
    try:
        docName = anvil.server.call("getQuotationInvoiceName", jobCardID)
        if docType == "Quotation":
            fileName = str(docName) + " Quotation"
        elif docType == "InterimQuotation":
            fileName = str(docName) + " Interim Quote"
        elif docType == "Invoice":
            fileName = str(docName) + " Invoice"
        elif docType == "Confirm Quotation":
            fileName = str(docName) + " Confirm Quotation"

        setting_options = {
            "encoding": "UTF-8",
            "custom-header": [("Accept-Encoding", "gzip")],
            "page-size": "A4",
            "orientation": "Portrait",
            "margin-top": "0.75in",
            "margin-right": "0.75in",
            "margin-bottom": "0.75in",
            "margin-left": "0.75in",
            "no-outline": False,
            "enable-local-file-access": None,
        }

        html_string = fillQuotationInvoiceData(jobCardID, docType)
        pdfkit.from_string(
            html_string, fileName, options=setting_options, configuration=config
        )
        media_object = anvil.media.from_file(fileName, "application/pdf", name=fileName)
        return media_object

    except Exception as e:
        print("PDF generation failed:", str(e))
        raise


@anvil.server.callable()
def deleteFile(jobCardID, docType):
    _get_current_user()
    docName = anvil.server.call("getQuotationInvoiceName", jobCardID)
    
    if docType == "Quotation":
        fileName = str(docName) + " Quotation"
    elif docType == "Confirm Quotation":
        fileName = str(docName) + " Confirm Quotation"
    elif docType == "InterimQuotation":
        fileName = str(docName) + " Interim Quote"
    elif docType == "Invoice":
        fileName = str(docName) + " Invoice"
    elif docType == "Payment":
        fileName = str(docName) + " Payment"
    elif docType == "Jobcard":
        fileName = str(docName) + " Jobcard"
    elif docType == "TechNotes":
        fileName = str(docName) + " Technician Notes"
    elif docType == "DefectsList":
        result = anvil.server.call("getJobCardRow", jobCardID)
        fileName = str(result["RegNo"]) + " Defects List"
    elif docType == "PricedDefectsList":
        result = anvil.server.call("getJobCardRow", jobCardID)
        fileName = str(result["RegNo"]) + " Priced Defects List"

    if os.path.exists(fileName):
        os.remove(fileName)
        print("File deleted successfully.")
    else:
        print("File does not exist.")


@anvil.server.callable()
def populate_confirmation_from_quote(jobcardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Item, Part_No, QuantityIssued, Amount 
            FROM tbl_quotation
            WHERE AssignedJobID = %s
            ORDER BY tbl_quotation.ID
        """
        cursor.execute(query, (jobcardID,))
        results = cursor.fetchall() or []

        # Return as a list of dictionaries
        return [
            {
                "Name": r[0],
                "Number": r[1],
                "Quantity": r[2],
                "Amount": f"{float(r[3]):,.2f}",
            }
            for r in results
        ]


# ***************************************************Quotation Feedback Details Section ************************************

@anvil.server.callable()
def saveFullQuotationPartsAndServicesFeedback(assignedDate, jobCardID, remarks, items):
    _get_current_user()
    with db_cursor() as cursor:

        # Step 1: Delete existing feedback records for the job card
        cursor.execute(
            "DELETE FROM tbl_quotationpartsandservicesfeedback WHERE AssignedJobID = %s",
            (jobCardID,),
        )
        cursor.execute(
            "DELETE FROM tbl_clientquotationfeedback WHERE AssignedJobID = %s",
            (jobCardID,),
        )

        # Step 2: Insert one feedback record
        cursor.execute(
            """
            INSERT INTO tbl_clientquotationfeedback (AssignedJobID, Remarks)
            VALUES (%s, %s)
        """,
            (jobCardID, remarks),
        )

        # Step 3: Get the last inserted ID
        feedback_id = cursor.lastrowid

        # Step 4: Insert all items linked to the single feedback ID
        insert_item = """
            INSERT INTO tbl_quotationpartsandservicesfeedback 
            (Date, AssignedJobID, Item, Part_No, QuantityIssued, Amount, ClientQuotationFeedbackID)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        for item in items:
            cursor.execute(
                insert_item,
                (
                    assignedDate,
                    jobCardID,
                    item["name"],
                    item["number"],
                    item["quantity"],
                    item["amount"],
                    feedback_id,
                ),
            )

        return {"status": "success", "feedback_id": feedback_id}

@anvil.server.callable()
def updateQuotationPartsAndServices(assignedDate, jobCardID, items):
    _get_current_user()
    with db_cursor() as cursor:
        # Step 1: Delete existing feedback records for the job card
        cursor.execute(
            "DELETE FROM tbl_quotationpartsandservicesfeedback WHERE AssignedJobID = %s",
            (jobCardID,),
        )
        # Step 2: Get Client Quotation Feedback ID
        cursor.execute(
            "SELECT ID FROM tbl_clientquotationfeedback WHERE AssignedJobID = %s",
            (jobCardID,),
        )

        row = cursor.fetchone()

        if not row:
            raise Exception("No client quotation feedback found for this JobCard.")

        feedback_id = row[0]
 
        # Step 3: Insert all items linked to the single feedback ID
        insert_item = """
            INSERT INTO tbl_quotationpartsandservicesfeedback 
            (Date, AssignedJobID, Item, Part_No, QuantityIssued, Amount, ClientQuotationFeedbackID)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        for item in items:
            cursor.execute(
                insert_item,
                (
                    assignedDate,
                    jobCardID,
                    item["name"],
                    item["number"],
                    item["quantity"],
                    item["amount"],
                    feedback_id,
                ),
            )

        return {"status": "success", "feedback_id": feedback_id}

@anvil.server.callable()
def getClientQuotationFeedback(JobCardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT ID FROM tbl_clientquotationfeedback WHERE AssignedJobID = %s
        """
        cursor.execute(query, (JobCardID,))
        result = cursor.fetchone()

        if not result:
            return "No feedback found"

        feedback_id = result[0]

        query2 = """
        SELECT Item, QuantityIssued AS Quantity 
        FROM tbl_quotationpartsandservicesfeedback 
        WHERE ClientQuotationFeedbackID = %s
        """
        cursor.execute(query2, (feedback_id,))
        rows = cursor.fetchall()

        return [{"Item": row[0], "Quantity": row[1]} for row in rows]


@anvil.server.callable()
def getQuotationConfirmationFeedback(JobCardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
        
            SELECT 
                tbl_quotationpartsandservicesfeedback.Item, tbl_quotationpartsandservicesfeedback.Part_No, 
                tbl_quotationpartsandservicesfeedback.QuantityIssued, tbl_quotationpartsandservicesfeedback.Amount, 
                tbl_carpartnames.ID AS CarPartID
            FROM 
                tbl_quotationpartsandservicesfeedback 
            LEFT JOIN 
                tbl_carpartnames ON tbl_quotationpartsandservicesfeedback.Part_No = tbl_carpartnames.PartNo
            WHERE 
                tbl_quotationpartsandservicesfeedback.AssignedJobID = %s
            ORDER BY
                tbl_quotationpartsandservicesfeedback.ID

        """
        cursor.execute(query, (JobCardID,))
        result = cursor.fetchall()

        if result:
            return [
                {
                    "Name": row[0],
                    "Number": row[1],
                    "Quantity": row[2],
                    "Amount": f"{float(row[3]):,.2f}",
                }
                for row in result
            ]
        else:
            return []


# ***************************************************Verify Task  Details Section ************************************


@anvil.server.callable()
def saveConfirmationDetails(jobCardId, remarks, signature, dateCompleted):
    _get_current_user()
    with db_cursor() as cursor:
        # Read binary content from Anvil media object
        signature_bytes = signature.get_bytes()

        query = """
            INSERT INTO tbl_completedjobcards (AssignedJobCardID , Remarks, Signature, DateCompleted)
            Values (%s, %s, %s, %s)
        """
        cursor.execute(query, (jobCardId, remarks, signature_bytes, dateCompleted))

        # Update job card to completed
        query2 = """
            UPDATE tbl_jobcarddetails SET tbl_jobcarddetails.IsComplete = 1
            WHERE ID = %s
        """
        cursor.execute(query2, (jobCardId,))


# ***************************************************Invoice Details Section ************************************
@anvil.server.callable()
def getInvoiceJobCardDetails():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT DISTINCT tbl_jobcarddetails.JobCardRef, tbl_invoices.AssignedJobID
            FROM tbl_invoices 
            INNER JOIN tbl_jobcarddetails 
            ON tbl_invoices.AssignedJobID = tbl_jobcarddetails.ID
            ORDER BY tbl_jobcarddetails.JobCardRef;
        """

        cursor.execute(query)
        result = cursor.fetchall()
        return [{"JobCardRef": r[0], "ID": r[1]} for r in result]


@anvil.server.callable()
def saveInvoice(assignedDate, jobCardID, items):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_invoices 
            (Date, AssignedJobID, Item, Part_No, QuantityIssued, Amount, Status)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        for item in items:
            cursor.execute(
                query,
                (
                    assignedDate,
                    jobCardID,
                    item["name"],
                    item["number"],
                    item["quantity"],
                    item["amount"],
                    "Pending",
                ),
            )
        query2 = """
            INSERT INTO tbl_assignedcarparts ( Date, AssignedJobID, CarPartID, QuantityIssued )
            VALUES (%s, %s, %s, %s)
        """
        for item2 in items:
            if item2["CarPartID"] is not None:
                cursor.execute(
                    query2,
                    (assignedDate, jobCardID, item2["CarPartID"], item2["quantity"]),
                )


@anvil.server.callable()
def get_invoice_details_by_job_id(job_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT 
                    tbl_clientcontacts.Fullname, 
                    tbl_jobcarddetails.MakeAndModel, 
                    tbl_jobcarddetails.RegNo, 
                    tbl_invoices.Date, 
                    tbl_jobcarddetails.ChassisNo, 
                    tbl_jobcarddetails.EngineCode, 
                    tbl_jobcarddetails.Mileage, 
                    tbl_invoices.Item, 
                    tbl_invoices.QuantityIssued, 
                    tbl_invoices.Amount, 
                    tbl_invoices.AssignedJobID,
                    tbl_invoices.Part_No,
                    tbl_carpartnames.ID AS CarPartID   -- ✅ new column
                FROM 
                    (tbl_jobcarddetails 
                    INNER JOIN tbl_clientcontacts 
                        ON tbl_jobcarddetails.ClientDetails = tbl_clientcontacts.ID) 
                INNER JOIN tbl_invoices 
                    ON tbl_jobcarddetails.ID = tbl_invoices.AssignedJobID
                LEFT JOIN tbl_carpartnames   -- ✅ join to match PartNo
                    ON tbl_invoices.Part_No = tbl_carpartnames.PartNo
                WHERE 
                    tbl_invoices.AssignedJobID = %s
                ORDER BY 
                    tbl_invoices.ID

        """
        cursor.execute(query, (job_id,))
        rows = cursor.fetchall()

        result = []
        for r in rows:
            # Helper function to ensure we get a float from various types
            def safe_float_convert(value):
                if value is None:
                    return None  # Or 0.0, depending on desired behavior for NULLs
                if isinstance(value, (int, float, decimal.Decimal)):
                    return float(value)
                # If it's a string, try to clean it and convert
                if isinstance(value, str):
                    try:
                        return float(value.replace(",", ""))
                    except ValueError:
                        # Handle cases where string cannot be converted (e.g., non-numeric string)
                        print(f"Warning: Could not convert string '{value}' to float.")
                        return None  # Or raise an error
                return None  # Default if type is unexpected

            quantity_issued_val = safe_float_convert(r[8])  # QuantityIssued
            amount_val = safe_float_convert(r[9])  # Amount

            # Ensure 'QuantityIssued' for the dictionary can be "" if originally None or non-numeric string
            # If your front-end expects an empty string, keep this logic
            display_quantity_issued = (
                ""
                if r[8] is None or not isinstance(r[8], (int, float, decimal.Decimal))
                else quantity_issued_val
            )

            # Calculate Total - ensuring we use the float versions for calculation
            total_calc = None
            if quantity_issued_val is None:  # If QuantityIssued was NULL/None from DB
                total_calc = amount_val
            elif amount_val is not None:  # Ensure Amount is not None for multiplication
                total_calc = round(quantity_issued_val * amount_val, 2)
            # If amount_val is None and quantity_issued_val is not None, total_calc would remain None

            result.append(
                {
                    "Fullname": r[0],
                    "MakeAndModel": r[1],
                    "RegNo": r[2],
                    "Date": r[3],
                    "ChassisNo": r[4],
                    "EngineCode": r[5],
                    "Mileage": r[6],
                    "Item": r[7],
                    "QuantityIssued": display_quantity_issued,  # Use the potentially "" value for display
                    "Amount": amount_val,  # This is now always a float or None
                    "AssignedJobID": r[10],
                    "PartNo": r[11],
                    "CarPartID": r[12],
                    "Total": total_calc,  # This is now always a float or None
                }
            )
        return result


@anvil.server.callable()
def getPendingInvoices():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT tbl_invoices.AssignedJobID, tbl_jobcarddetails.JobCardRef FROM tbl_invoices 
            INNER JOIN tbl_jobcarddetails ON tbl_invoices.AssignedJobID = tbl_jobcarddetails.ID 
            WHERE (((tbl_invoices.Status)="Pending")) 
            GROUP BY tbl_invoices.AssignedJobID
        """
        cursor.execute(query)
        result = cursor.fetchall()
        return [{"ID": r[0], "JobCardRef": r[1]} for r in result]


@anvil.server.callable()
def get_invoice_total_by_job_id(assigned_job_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT SUM(
                COALESCE(QuantityIssued, 1) * Amount
            ) AS GrandTotal
            FROM tbl_invoices
            WHERE AssignedJobID = %s
        """
        cursor.execute(query, (assigned_job_id,))
        result = cursor.fetchone()
        return f"{float(result[0]):,.2f}" if result and result[0] is not None else 0.0


@anvil.server.callable()
def update_invoice_status(jobCardRefID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_invoices
            SET Status = 'Paid'
            WHERE AssignedJobID = %s
        """
        cursor.execute(query, (jobCardRefID,))


@anvil.server.callable()
def get_invoice_totals_and_counts_by_date(start_date, end_date):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Status,
                   COUNT(DISTINCT AssignedJobID) AS count,
                   SUM(
                       CASE
                           WHEN QuantityIssued IS NULL THEN 1 * Amount
                           ELSE QuantityIssued * Amount
                       END
                   ) AS total
            FROM tbl_invoices
            WHERE Status IN ('Pending', 'Paid')
              AND Date BETWEEN %s AND %s
            GROUP BY Status
        """
        cursor.execute(query, (start_date, end_date))
        results = cursor.fetchall()

        # Structure: { "Pending": {"count": X, "total": Y}, "Paid": {...} }
        summary = {
            "Pending": {"count": 0, "total": 0.0},
            "Paid": {"count": 0, "total": 0.0},
        }

        for status, count, total in results:
            summary[status]["count"] = count
            summary[status]["total"] = float(total) if total else 0.0

        return summary


@anvil.server.callable()
def getJobCardRefInvoiceDetails(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT DISTINCT 
                tbl_jobcarddetails.ID, tbl_jobcarddetails.JobCardRef 
            FROM 
                tbl_jobcarddetails 
            JOIN 
                tbl_invoices ON tbl_jobcarddetails.ID = tbl_invoices.AssignedJobID 
            WHERE 
                tbl_invoices.Status = 'Pending' 
            AND 
                tbl_jobcarddetails.JobCardRef LIKE %s
            ORDER BY 
                tbl_jobcarddetails.JobCardRef DESC;
        """

        like_value = f"%{value}%"
        cursor.execute(query, (like_value,))
        result = cursor.fetchall()

        # Return directly as list of tuples for a dropdown [(display_text, value), ...]
        return [(row[1], row[0]) for row in result]  # (JobCardRef, AssignedJobID)


@anvil.server.callable()
def updateInvoice(invoicedate, job_card_id, items):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            DELETE FROM tbl_invoices WHERE AssignedJobID = %s
        """
        cursor.execute(query, (job_card_id,))

        query2 = """
            INSERT INTO tbl_invoices (Date,AssignedJobID, Item, Part_No, QuantityIssued, Amount, Status)
            VALUES (%s, %s, %s, %s, %s, %s, "Pending")
        """

        for item in items:
            cursor.execute(
                query2,
                (
                    invoicedate,
                    job_card_id,
                    item.get("name"),
                    item.get("number"),
                    item.get("quantity"),
                    item.get("amount"),
                ),
            )
        query3 = """
            DELETE FROM tbl_assignedcarparts WHERE AssignedJobID = %s
        """
        cursor.execute(query3, (job_card_id,))

        query4 = """
            INSERT INTO tbl_assignedcarparts ( Date, AssignedJobID, CarPartID, QuantityIssued )
            VALUES (%s, %s, %s, %s)
        """
        for item2 in items:
            if item2["CarPartID"] is not None:
                cursor.execute(
                    query4,
                    (invoicedate, job_card_id, item2["CarPartID"], item2["quantity"]),
                )


@anvil.server.callable()
def getInvoiceStatus(jobcardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = "SELECT DISTINCT Status FROM tbl_invoices WHERE AssignedJobID = %s"
        cursor.execute(query, (jobcardID,))
        result = cursor.fetchone()
        return result[0] if result else None


# ***************************************************Payment Details Section ************************************


@anvil.server.callable()
def get_previous_payment(invoice_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT SUM(AmountPaid)
            FROM tbl_payments
            WHERE JobCardRefID = %s
        """
        cursor.execute(query, (invoice_id,))
        result = cursor.fetchone()
        return f"{float(result[0]):,.2f}" if result[0] is not None else 0.0

@anvil.server.callable()
def get_discount_payment(invoice_id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT SUM(Discount)
            FROM tbl_payments
            WHERE JobCardRefID = %s
        """
        cursor.execute(query, (invoice_id,))
        result = cursor.fetchone()
        return f"{float(result[0]):,.2f}" if result[0] is not None else 0.0
    
@anvil.server.callable()
def save_payment_details(
    paymentDate, jobCardRefID, paymentMode, amountPaid, discount, bal
):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_payments (Date, JobCardRefID, PaymentMode, AmountPaid, Discount, Balance)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query, (paymentDate, jobCardRefID, paymentMode, amountPaid, discount, bal)
        )


@anvil.server.callable()
def getPaidPendingInvoices(status, start_date, end_date):
    _get_current_user()
    count = 0
    with db_cursor() as cursor:
        query = """
        SELECT tbl_invoices.Date,  tbl_jobcarddetails.JobCardRef,  tbl_invoices.Item,  tbl_invoices.Part_No, tbl_invoices.QuantityIssued, tbl_invoices.Amount
        FROM tbl_invoices 
        JOIN tbl_jobcarddetails ON tbl_invoices.AssignedJobID = tbl_jobcarddetails.ID 
        WHERE tbl_invoices.STATUS = %s AND tbl_invoices.DATE BETWEEN %s AND %s 
        ORDER BY tbl_invoices.DATE DESC;
        """
        # Pass all parameters to the execute method in the correct order
        cursor.execute(query, (status, start_date, end_date))
        rows = cursor.fetchall()
        result = []
        for count, row in enumerate(rows, start=1):
            # Helper function to ensure we get a float from various types
            def safe_float_convert(value):
                if value is None:
                    return None  # Or 0.0, depending on desired behavior for NULLs
                if isinstance(value, (int, float, decimal.Decimal)):
                    return float(value)
                # If it's a string, try to clean it and convert
                if isinstance(value, str):
                    try:
                        return float(value.replace(",", ""))
                    except ValueError:
                        # Handle cases where string cannot be converted (e.g., non-numeric string)
                        print(f"Warning: Could not convert string '{value}' to float.")
                        return None  # Or raise an error
                return None  # Default if type is unexpected

            quantity_issued_val = safe_float_convert(row[4])  # QuantityIssued
            amount_val = safe_float_convert(row[5])  # Amount

            # Ensure 'QuantityIssued' for the dictionary can be "" if originally None or non-numeric string
            # If your front-end expects an empty string, keep this logic
            display_quantity_issued = (
                ""
                if row[4] is None
                or not isinstance(row[4], (int, float, decimal.Decimal))
                else quantity_issued_val
            )

            # Calculate Total - ensuring we use the float versions for calculation
            total_calc = None
            if quantity_issued_val is None:  # If QuantityIssued was NULL/None from DB
                total_calc = amount_val
            elif amount_val is not None:  # Ensure Amount is not None for multiplication
                total_calc = round(quantity_issued_val * amount_val, 2)
            # If amount_val is None and quantity_issued_val is not None, total_calc would remain None

            result.append(
                {
                    "No": count,
                    "Date": row[0],
                    "JobCardRef": row[1],
                    "Item": row[2],
                    "PartNo": row[3],
                    "Quantity": display_quantity_issued,
                    "Amount": f"{float(amount_val):,.2f}",
                    "Total": f"{float(total_calc):,.2f}",
                }
            )

        return result


@anvil.server.callable()
def getPaymentsDetails(paymentID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                            SELECT
                                p.Date,
                                j.JobCardRef,
                                p.PaymentMode,
                                SUM(
                                    COALESCE(i.QuantityIssued, 1) * i.Amount
                                ) AS InvoiceAmount,
                                p.AmountPaid,
                                p.Discount,
                                p.Balance
                            FROM
                                tbl_payments AS p
                            JOIN tbl_jobcarddetails AS j
                              ON p.JobCardRefID = j.ID
                            JOIN tbl_invoices AS i
                              ON j.ID = i.AssignedJobID
                            WHERE p.JobCardRefID = %s
                            GROUP BY
                                p.ID,
                                p.Date,
                                j.JobCardRef,
                                p.PaymentMode,
                                p.AmountPaid,
                                p.Discount,
                                p.Balance
                            ORDER BY
                                p.Date DESC
                        """
        cursor.execute(query, (paymentID,))
        rows = cursor.fetchall()
        # Convert rows to a list of dictionaries
        result = [
            {
                "No": index + 1,
                "Date": row[0],
                "JobCardRef": row[1],
                "PaymentMode": row[2],
                "InvoiceAmount": f"{float(row[3]):,.2f}" if row[3] is not None else 0.0,
                "AmountPaid": f"{float(row[4]):,.2f}" if row[4] is not None else 0.0,
                "Discount": f"{float(row[5]):,.2f}" if row[5] is not None else 0.0,
                "Balance": f"{float(row[6]):,.2f}" if row[6] is not None else 0.0,
            }
            for index, row in enumerate(rows)
        ]

        return result


@anvil.server.callable()
def get_payment_ref():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT
              DISTINCT j.JobCardRef,
              p.JobCardRefID
            FROM
              tbl_payments AS p
            JOIN
              tbl_jobcarddetails AS j ON p.JobCardRefID = j.ID
            ORDER BY
              j.JobCardRef ASC
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        result = [{"JobCardRef": row[0], "JobCardRefID": row[1]} for row in rows]

        return result


@anvil.server.callable()
def fillReportData(jobCardID, docType, logo_path: str = os.getenv("LOGO"), font_path: str = os.getenv("FONT_PATH")) -> str:
    _get_current_user()
    if docType == "Payment":
        docTitle = "Payment Details"
        reportdetails = getPaymentsDetails(jobCardID)
        vehicledetails = get_invoice_details_by_job_id(jobCardID)

    # === Embed MozillaHeadline font as base64 ===
    font_base64 = ""
    if font_path and os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # === Handle company logo ===
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
        logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" alt="Company Logo" style="width: 100%; height: 100%; border-radius: 2px;">'
    else:
        logo_img_tag = "LOGO"

    # Generate items table rows
    items_html = ""
    counter = 0
    for item in reportdetails:
        counter = counter + 1
        items_html += f"""
                    <tr class="item-row">
                        <td>{counter}</td>
                        <td>{html_module.escape(str(item['Date']))}</td>
                        <td>{html_module.escape(str(item['JobCardRef']))}</td>
                        <td>{html_module.escape(str(item['PaymentMode']))}</td>
                        <td>{html_module.escape(str(item['InvoiceAmount']))}</td>
                        <td>{html_module.escape(str(item['AmountPaid']))}</td>
                        <td>{html_module.escape(str(item['Discount']))}</td>
                        <td>{html_module.escape(str(item['Balance']))}</td>
                    </tr>"""

        # Complete HTML template with fixed structure
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{docTitle}</title>
    <style>
        /* --- Embed Mozilla Headline font --- */
        @font-face {{
            font-family: 'Mozilla Headline';
            src: url(data:font/ttf;base64,{font_base64}) format('truetype');
        }}
        body {{
            font-family: Roboto, Noto, Arial, sans-serif;
            font-size: 14px;
            line-height: 1.4286;
            background-color: #fafafa;
            margin: 0;
            padding: 16px;
        }}

        .quotation-container {{
            background-color: white;
            border-radius: 2px;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
            max-width: 800px;
            margin: 0 auto;
            overflow: hidden;
        }}

        .logo-section {{
            text-align: center;
            padding: 24px;
            background-color: white;
            border-bottom: 1px solid #e0e0e0;
        }}

        .logo-container {{
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 20px;
            margin-bottom: 16px;
        }}

        .logo-image {{
            width: 725px;
            height: 100px;
            background: linear-gradient(135deg, #228B22, #90EE90, #FFD700, #FF6347);
            border-radius: 2px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
            color: white;
            font-weight: 500;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}

        .header {{
            background-color: #000;
            color: white;
            text-align: center;
            padding: 16px 24px;
            font-size: 16px;
            font-weight: 300;
            letter-spacing: .5px;
            box-shadow: 0 4px 5px 0 rgba(0, 0, 0, 0.14),
                        0 1px 10px 0 rgba(0, 0, 0, 0.12),
                        0 2px 4px -1px rgba(0, 0, 0, 0.2);
        }}

        .detail-row {{
            display: grid;
            grid-template-columns: 140px 1fr; /* label column, value column */
            column-gap: 8px;
            margin-bottom: 12px;
        }}

        .detail-label {{
                font-weight: bold;
                font-size: 16px;
                color: rgba(0,0,0,0.87);
                
        }}

        .detail-value {{
            font-size: 16px;
            color: rgba(0,0,0,0.87);
            text-align: left;
            
        }}

        .items-table {{
            border-collapse: collapse;
            width: 100%;
            margin: 0 24px 24px 0;
            background-color: white;
            border-radius: 2px;
            overflow: hidden;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}

        .items-table th {{
            background-color: #f5f5f5;
            border-bottom: 1px solid #e0e0e0;
            padding: 16px;
            text-align: left;
            font-weight: bold;
            font-size: 14px;
            color: rgba(0,0,0,0.87);
            text-transform: uppercase;
            letter-spacing: .5px;
            
        }}

        .items-table td {{
            border-bottom: 1px solid rgba(0,0,0,0.12);
            padding: 16px;
            font-size: 14px;
            color: rgba(0,0,0,0.87);
           
        }}

        .items-table .item-row:hover {{
            background-color: rgba(0,0,0,0.04);
        }}

        .total-row {{
            background-color: #000 !important;
            color: white !important;
        }}

        .total-row td {{
            border-bottom: none !important;
            font-weight: 300;
            font-size: 16px;
            color: white !important;
            padding: 16px;
           
        }}

        .notes-section {{
            padding: 24px;
            background-color: #f5f5f5;
            margin-top: 16px;
        }}

        .notes-title {{
            margin-bottom: 16px;
            font-weight: 500;
            font-size: 16px;
            color: rgba(0,0,0,0.87);
            font-family: 'Mozilla Headline';
        }}

        .notes-list {{
            margin: 0;
            padding-left: 24px;
            color: rgba(0,0,0,0.74);
        }}

        .notes-list li {{
            margin-bottom: 8px;
            line-height: 1.5;
            font-family: 'Mozilla Headline';
        }}
        #footer  div {{
                width: 80%;
                margin: 0 auto;
                text-align: center;
                font-size: 12px;
                font-family: 'Mozilla Headline';
            }}
    </style>
</head>
<body>
    <div class="quotation-container">
        <div class="logo-section">
            <div class="logo-container">
                <div class="logo-image">
                    {logo_img_tag}
                </div>
            </div>
        </div>

        <div class="header">
            {docTitle.upper()}
        </div>
         <!-- UPDATED: Replaced details-section div with table layout -->
        <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
            <tr>
                <!-- Left Column -->
                <td style="width: 50%; vertical-align: top; padding-left: 24px; padding-right: 32px;">
                    <div class="detail-row">
                        <span class="detail-label">Customer Name:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['Fullname']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Make And Model:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['MakeAndModel']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Reg No:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['RegNo']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Date:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['Date']}</span>
                        </div>
                    </div>
                </td>

                <!-- Right Column -->
                <td style="width: 50%; vertical-align: top; padding-left: 32px;">
                    <div class="detail-row">
                        <span class="detail-label">Chassis:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['ChassisNo']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Engine:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['EngineCode']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Mileage:</span>
                        <div>
                        <span class="detail-value">{vehicledetails[0]['Mileage']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">&nbsp;</span>
                        <span class="detail-value">&nbsp;</span>
                    </div>
                </td>
            </tr>
        </table>
        <!-- END UPDATED -->

        <table class="items-table">
            <thead>
                <tr>
                    <th>No.</th>
                    <th>Date</th>
                    <th>Ref</th>
                    <th>Mode</th>
                    <th>Invoiced</th>
                    <th>Paid</th>
                    <th>Discount</th>
                    <th>Balance</th>
                </tr>
            </thead>
            <tbody>
                {items_html}
            </tbody>
        </table>
        <footer id="footer">
        <div> 
            <p>Joy Is The Feeling Of Being Looked After By The Best - BMW CENTER For Your BMW.</p>
        </div>
    </footer>    
    </div>
    </body>
    </html>"""

    return html_content


@anvil.server.callable()
def createReportPdf(jobCardID, docType):
    _get_current_user()
    try:
        docName = anvil.server.call("getQuotationInvoiceName", jobCardID)
        if docType == "Payment":
            fileName = str(docName) + " Payment"

        setting_options = {
            "encoding": "UTF-8",
            "custom-header": [("Accept-Encoding", "gzip")],
            "page-size": "A4",
            "orientation": "Portrait",
            "margin-top": "0.75in",
            "margin-right": "0.75in",
            "margin-bottom": "0.75in",
            "margin-left": "0.75in",
            "no-outline": False,
            "enable-local-file-access": None,
        }

        html_string = fillReportData(jobCardID, docType)
        pdfkit.from_string(
            html_string, fileName, options=setting_options, configuration=config
        )
        media_object = anvil.media.from_file(fileName, "application/pdf", name=fileName)
        return media_object

    except Exception as e:
        print("PDF generation failed:", str(e))
        raise


# ***************************************************Duplicate Phone Number Details Section ************************************


# Check for existing contacts
@anvil.server.callable()
def check_duplicate_contact(contacttype, phone):
    _get_current_user()
    with db_cursor() as cursor:
        if contacttype == "Client":
            query = """
                SELECT ID FROM tbl_clientcontacts 
                WHERE Phone = %s
            """
            cursor.execute(query, (phone,))
        elif contacttype == "Technician":
            query2 = """
                            SELECT ID FROM tbl_technicians 
                            WHERE Phone = %s
                        """
            cursor.execute(query2, (phone,))
        elif contacttype == "Staff":
            query3 = """
                            SELECT ID FROM tbl_checkstaff 
                            WHERE Phone = %s
                        """
            cursor.execute(query3, (phone,))
        elif contacttype == "Supplier":
            query4 = """
                            SELECT ID FROM tbl_carpartssupplier 
                            WHERE Phone = %s
                        """
            cursor.execute(query4, (phone,))
        result = cursor.fetchone()
        return result is not None


# *************************************************** Interim Quotation Details Section ************************************


@anvil.server.callable()
def saveJobCardDetailsFromInterimQuotation(
    customer_ID,
    job_card_ref,
    received_date,
    due_date,
    check_in_staff,
    reg_no,
    make_and_model,
    chassis,
    engine_code,
    mileage,
    oldJobCardID,
):
    _get_current_user()
    with db_cursor() as cursor:
        # Check if old job card ref ends with IQ
        result = anvil.server.call("getJobCardRef", oldJobCardID)
        # If it does, delete old quotation and job card for a fresh start
        if result and "JobCardRef" in result[0]:
            if result[0]["JobCardRef"].endswith("IQ"):
                # Delete old quotation if exists
                cursor.execute(
                    "DELETE FROM tbl_quotation WHERE AssignedJobID = %s",
                    (oldJobCardID,),
                )
                # Delete old job card if exists
                cursor.execute(
                    "DELETE FROM tbl_jobcarddetails WHERE ID = %s", (oldJobCardID,)
                )

        # Insert new job card details
        query = """
            INSERT INTO tbl_jobcarddetails
            (ClientDetails, JobCardRef, ReceivedDate, DueDate, CheckedInBy, 
             RegNo, MakeAndModel, ChassisNo, EngineCode, Mileage, 
             Ins, ClientInstruction, Notes, IsComplete) 
            VALUES (%s, %s, %s, %s, %s, 
                    %s, %s, %s, %s, %s, 
                    %s, %s, %s, %s)
        """
        cursor.execute(
            query,
            (
                customer_ID,
                job_card_ref,
                received_date,
                due_date,
                check_in_staff,
                reg_no,
                make_and_model,
                chassis,
                engine_code,
                mileage,
                1,
                "None",
                "None",
                1,
            ),
        )
        return cursor.lastrowid


@anvil.server.callable()
def saveInterimQuotationPartsAndServices(assignedDate, jobCardID, items):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                INSERT INTO tbl_quotation (Date, AssignedJobID, Item, Part_No, QuantityIssued, Amount)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
        for item in items:
            name = item.get("name", "")
            number = item.get("number", "")
            quantity = item.get("quantity")
            amount = item.get("amount")
            cursor.execute(
                query, (assignedDate, jobCardID, name, number, quantity, amount)
            )


@anvil.server.callable()
def getInterimQuoteAndAmendedInvoiceStaff():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Staff
            FROM tbl_checkstaff
            WHERE IsArchived = FALSE
            ORDER BY Staff ASC
        """
        cursor.execute(query)
        result = cursor.fetchall()
        # Return list of tuples: (label shown, value stored)
        return [(r[1], r[0]) for r in result]


@anvil.server.callable()
def getPreviousInterimQuoteDetails(customerID, makeAndModel, chassis, receivedDate):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                j.ID,
                j.JobCardRef,
                j.ReceivedDate,
                j.RegNo,
                j.MakeAndModel,
                j.ChassisNo,
                j.EngineCode,
                j.Mileage,
                q.Item,
                q.Part_No,
                q.QuantityIssued,
                q.Amount
            FROM 
                tbl_jobcarddetails j
            LEFT JOIN 
                tbl_quotation q ON j.ID = q.AssignedJobID
            WHERE 
                j.ClientDetails = %s
                AND j.MakeAndModel = %s
                AND j.ChassisNo = %s
                AND j.ReceivedDate = %s
                AND q.Item IS NOT NULL  -- Only include rows with parts data
            ORDER BY 
                j.ReceivedDate DESC, j.ID DESC;
        """
        cursor.execute(query, (customerID, makeAndModel, chassis, receivedDate))
        rows = cursor.fetchall()  # Get all rows instead of just one

        if rows:
            # Get job details from first row
            job_details = {
                "JobCardID": rows[0][0],
                "JobCardRef": rows[0][1],
                "ReceivedDate": rows[0][2],
                "RegNo": rows[0][3],
                "MakeAndModel": rows[0][4],
                "ChassisNo": rows[0][5],
                "EngineCode": rows[0][6],
                "Mileage": rows[0][7],
            }

            # Create parts list
            parts_list = []
            for row in rows:
                if row[8] is not None:  # If Item exists
                    parts_list.append(
                        {
                            "Name": row[8],
                            "Number": row[9],
                            "Quantity": row[10],
                            "Amount": (
                                f"{float(row[11]):,.2f}"
                                if row[11] is not None
                                else "0.00"
                            ),
                        }
                    )

            return {"job_details": job_details, "parts": parts_list}
        else:
            return None


@anvil.server.callable()
def getInterimQuote(value):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, JobCardRef
            FROM tbl_jobcarddetails
            WHERE LOWER(JobCardRef) LIKE %s
              AND LOWER(JobCardRef) LIKE '%%-iq';
        """
        like_value = f"%{value.lower()}%"
        cursor.execute(query, (like_value,))
        results = cursor.fetchall()

        # Return in format suitable for dropdown: [(display_text, value), ...]
        return [(row[1], row[0]) for row in results]  # (JobCardRef, ID)


@anvil.server.callable()
def transitionInterimQuoteToInvoice(jocardID):
    _get_current_user()
    with db_cursor() as cursor:
        # Delete old invoice details
        cursor.execute("DELETE FROM tbl_invoices WHERE AssignedJobID = %s", (jocardID,))

        # Insert new invoice details from interim quotation
        query = """
            INSERT INTO tbl_invoices (Date, AssignedJobID, Item, Part_No, QuantityIssued, Amount, Status)
            SELECT Date, AssignedJobID, Item, Part_No, QuantityIssued, Amount, 'Pending'
            FROM tbl_quotation
            WHERE AssignedJobID = %s   
        """
        cursor.execute(query, (jocardID,))

        # Update jobcard status to Ready for Pickup
        cursor.execute(
            "UPDATE tbl_jobcarddetails SET Status = 'Ready for Pickup' WHERE ID = %s",
            (jocardID,),
        )


# *************************************************** Inventory Details Section ************************************


@anvil.server.callable()
def get_filtered_parts(part_filter=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            WITH latest_stock AS (
                SELECT sp.CarPart, sp.UnitCost, sp.Date,
                       ROW_NUMBER() OVER (PARTITION BY sp.CarPart ORDER BY sp.ID DESC) AS rn
                FROM tbl_stockparts sp
            ),
            issued AS (
                SELECT CarPartID, SUM(QuantityIssued) AS TotalIssued
                FROM tbl_assignedcarparts
                GROUP BY CarPartID
            ),
            latest_stocktake AS (
                SELECT CarPartNameID, HarmornizedValue
                FROM (
                    SELECT 
                        CarPartNameID,
                        HarmornizedValue,
                        ROW_NUMBER() OVER (PARTITION BY CarPartNameID ORDER BY StockTakeDate DESC) AS rn
                    FROM tbl_stocktakeharmonized
                ) ranked
                WHERE rn = 1
            )
            SELECT 
                cpn.Name,
                cpn.PartNo,
                cpn.OrderLevel AS `Reorder Level`,
                ls.UnitCost AS Cost,
                psp.Amount AS Selling,
                psp.SaleDiscount AS Discount,
                MAX(sp.Date) AS MaxOfDate,
                (COALESCE(SUM(sp.NoOfUnits), 0) 
                    - COALESCE(i.TotalIssued, 0)
                    + COALESCE(lst.HarmornizedValue, 0)) AS StockBalance
            FROM tbl_carpartnames cpn
            INNER JOIN tbl_partssellingprice psp 
                ON cpn.ID = psp.CarPartsNamesID
            INNER JOIN tbl_stockparts sp 
                ON cpn.ID = sp.CarPart
            INNER JOIN latest_stock ls 
                ON ls.CarPart = cpn.ID AND ls.rn = 1
            LEFT JOIN issued i 
                ON cpn.ID = i.CarPartID
            LEFT JOIN latest_stocktake lst 
                ON cpn.ID = lst.CarPartNameID
            WHERE cpn.Name LIKE %s OR cpn.PartNo LIKE %s
            GROUP BY 
                cpn.Name,
                cpn.PartNo,
                cpn.OrderLevel,
                psp.Amount,
                psp.SaleDiscount,
                cpn.ID,
                ls.UnitCost,
                i.TotalIssued,
                lst.HarmornizedValue
            ORDER BY cpn.Name;
        """

        like_filter = f"%{part_filter}%"
        cursor.execute(query, (like_filter, like_filter))
        rows = cursor.fetchall()

        result = []
        for count, row in enumerate(rows, start=1):
            result.append(
                {
                    "No": count,
                    "Name": row[0],
                    "PartNo": row[1],
                    "ReorderLevel": row[2],
                    "Cost": f"{float(row[3]):,.2f}" if row[3] is not None else None,
                    "Selling": f"{float(row[4]):,.2f}" if row[4] is not None else None,
                    "Discount": row[5],
                    "MaxOfDate": row[6],
                    "StockBalance": row[7],
                }
            )

        return result

@anvil.server.callable()
def search_car_parts_location(search_term=""):
    _get_current_user()
    with db_cursor() as cursor:
        # If no search term is provided, avoid unnecessary LIKE scans
        if search_term:
            query = """
                SELECT 
                    n.ID,
                    n.Name,
                    n.PartNo,
                    l.Location
                FROM tbl_carpartnames n
                INNER JOIN tbl_carpartslocation l ON l.ID = n.Location
                WHERE n.Name LIKE %s
                   OR n.PartNo LIKE %s
                   OR l.Location LIKE %s
                ORDER BY n.Name
            """
            like_pattern = f"%{search_term}%"
            params = (like_pattern, like_pattern, like_pattern)
        else:
            query = """
                SELECT 
                    n.ID,
                    n.Name,
                    n.PartNo,
                    l.Location
                FROM tbl_carpartnames n
                INNER JOIN tbl_carpartslocation l ON l.ID = n.Location
                ORDER BY n.Name
            """
            params = ()

        cursor.execute(query, params)
        rows = cursor.fetchall()

        return [
            {"No": count, "Name": row[1], "PartNo": row[2], "Location": row[3]}
            for count, row in enumerate(rows, start=1)
        ]


@anvil.server.callable()
def missing_buying_prices(search_term=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT
                tbl_carpartssupplier.Name, 
                tbl_carpartnames.Name, 
                tbl_carpartnames.PartNo, 
                tbl_stockparts.UnitCost
            FROM 
                tbl_carpartnames 
            INNER JOIN 
                tbl_stockparts 
            ON 
                tbl_carpartnames.ID = tbl_stockparts.CarPart
            INNER JOIN 
                tbl_carpartssupplier 
            ON 
                tbl_stockparts.CarPartsSupplierID = tbl_carpartssupplier.ID
            WHERE 
                tbl_stockparts.UnitCost <1
                AND 
                    (tbl_carpartnames.Name LIKE %s 
                        OR 
                     tbl_carpartnames.PartNo LIKE %s )
            ORDER BY 
                tbl_carpartnames.Name
        """

        like_pattern = f"%{search_term}%"

        cursor.execute(query, (like_pattern, like_pattern))
        rows = cursor.fetchall()

        result = []
        for count, row in enumerate(rows, start=1):
            result.append(
                {
                    "No": count,
                    "Supplier": row[0],
                    "Name": row[1],
                    "PartNo": row[2],
                    "Cost": f"{row[3]:,.2f}",
                }
            )

        return result


@anvil.server.callable()
def get_buying_prices(search_term=""):
    _get_current_user()
    with db_cursor() as cursor:
        base_query = """
            SELECT
                sp.Date,
                s.Name AS Supplier,
                cp.Name AS PartName,
                cp.PartNo,
                sp.UnitCost,
                sp.NoOfUnits
            FROM tbl_stockparts sp
            INNER JOIN tbl_carpartnames cp ON cp.ID = sp.CarPart
            INNER JOIN tbl_carpartssupplier s ON s.ID = sp.CarPartsSupplierID
            WHERE sp.UnitCost > 0
        """

        params = []

        # Add search filter only if search_term is provided
        if search_term and search_term.strip():
            base_query += " AND (cp.Name LIKE %s OR cp.PartNo LIKE %s)"
            like_pattern = f"%{search_term.strip()}%"
            params.extend([like_pattern, like_pattern])

        # Always order results
        base_query += " ORDER BY sp.Date DESC, cp.Name"

        cursor.execute(base_query, tuple(params))
        rows = cursor.fetchall()

        result = [
            {
                "No": idx,
                "Date": row[0],
                "Supplier": row[1],
                "Name": row[2],
                "PartNo": row[3],
                "Cost": f"{row[4]:,.2f}",
                "NoOfUnits": row[5],
            }
            for idx, row in enumerate(rows, start=1)
        ]

        return result


@anvil.server.callable()
def get_buying_prices_by_partID(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT
                tbl_stockparts.Date,
                tbl_carpartssupplier.Name, 
                tbl_carpartnames.Name, 
                tbl_carpartnames.PartNo, 
                tbl_stockparts.UnitCost
            FROM 
                tbl_carpartnames 
            INNER JOIN 
                tbl_stockparts 
            ON 
                tbl_carpartnames.ID = tbl_stockparts.CarPart
            INNER JOIN 
                tbl_carpartssupplier 
            ON 
                tbl_stockparts.CarPartsSupplierID = tbl_carpartssupplier.ID
            WHERE 
                tbl_stockparts.UnitCost >0
                AND 
                    tbl_carpartnames.ID = %s
            ORDER BY 
                tbl_stockparts.Date DESC, tbl_carpartnames.Name 
        """

        cursor.execute(query, (valueID,))
        rows = cursor.fetchall()

        result = []
        for count, row in enumerate(rows, start=1):
            result.append(
                {
                    "No": count,
                    "Date": row[0],
                    "Supplier": row[1],
                    "Name": row[2],
                    "PartNo": row[3],
                    "Cost": f"{row[4]:,.2f}",
                }
            )

        return result


@anvil.server.callable()
def missing_selling_prices(search_term=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT
                      tbl_carpartssupplier.Name,  
                      tbl_carpartnames.Name, 
                      tbl_carpartnames.PartNo, 
                      tbl_partssellingprice.Amount,
                      tbl_partssellingprice.SaleDiscount
                    FROM 
                      tbl_carpartnames
                    INNER JOIN tbl_stockparts 
                      ON tbl_carpartnames.ID = tbl_stockparts.CarPart
                    INNER JOIN 
                        tbl_carpartssupplier 
                    ON 
                        tbl_stockparts.CarPartsSupplierID = tbl_carpartssupplier.ID
                    LEFT JOIN tbl_partssellingprice 
                      ON tbl_stockparts.CarPart = tbl_partssellingprice.CarPartsNamesID
                    WHERE 
                      tbl_partssellingprice.Amount < 1
                    AND
                                            (tbl_carpartnames.Name LIKE %s 
                                             OR tbl_carpartnames.PartNo LIKE %s )
                    GROUP BY 
                      tbl_carpartnames.Name, 
                      tbl_carpartnames.PartNo, 
                      tbl_partssellingprice.Amount,
                      tbl_carpartssupplier.Name,
                      tbl_partssellingprice.SaleDiscount
                    ORDER BY 
                      tbl_carpartnames.Name   
        """

        like_pattern = f"%{search_term}%"

        cursor.execute(query, (like_pattern, like_pattern))
        rows = cursor.fetchall()

        result = []
        for count, row in enumerate(rows, start=1):
            result.append(
                {
                    "No": count,
                    "Supplier": row[0],
                    "Name": row[1],
                    "PartNo": row[2],
                    "Amount": f"{row[3]:,.2f}",
                    "Discount": None if row[4] is None else f"{float(row[4]):,.2f}",
                }
            )

        return result


@anvil.server.callable()
def get_selling_prices(search_term=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT
                      tbl_carpartssupplier.Name,  
                      tbl_carpartnames.Name, 
                      tbl_carpartnames.PartNo, 
                      tbl_partssellingprice.Amount,
                      tbl_partssellingprice.SaleDiscount
                    FROM 
                      tbl_carpartnames
                    INNER JOIN tbl_stockparts 
                      ON tbl_carpartnames.ID = tbl_stockparts.CarPart
                    INNER JOIN 
                        tbl_carpartssupplier 
                    ON 
                        tbl_stockparts.CarPartsSupplierID = tbl_carpartssupplier.ID
                    LEFT JOIN tbl_partssellingprice 
                      ON tbl_stockparts.CarPart = tbl_partssellingprice.CarPartsNamesID
                    WHERE 
                      tbl_partssellingprice.Amount > 0
                    AND
                                            (tbl_carpartnames.Name LIKE %s 
                                             OR tbl_carpartnames.PartNo LIKE %s )
                    GROUP BY 
                      tbl_carpartnames.Name, 
                      tbl_carpartnames.PartNo, 
                      tbl_partssellingprice.Amount,
                      tbl_carpartssupplier.Name,
                      tbl_partssellingprice.SaleDiscount
                    ORDER BY 
                      tbl_carpartnames.Name   
        """

        like_pattern = f"%{search_term}%"

        cursor.execute(query, (like_pattern, like_pattern))
        rows = cursor.fetchall()

        result = []
        for count, row in enumerate(rows, start=1):
            result.append(
                {
                    "No": count,
                    "Supplier": row[0],
                    "Name": row[1],
                    "PartNo": row[2],
                    "Amount": f"{row[3]:,.2f}",
                    "Discount": None if row[4] is None else f"{float(row[4]):,.2f}",
                }
            )

        return result


@anvil.server.callable()
def updatePrice(priceType, newPrice, discount, partNo):
    _get_current_user()
    with db_cursor() as cursor:
        if priceType == "Selling":
            query = """
                UPDATE tbl_partssellingprice
                SET 
                    SetPriceDate = NOW(),
                    Amount = %s,
                    SaleDiscount = %s 
                WHERE CarPartsNamesID = (
                    SELECT ID FROM tbl_carpartnames WHERE PartNo = %s
                )
            """
            cursor.execute(query, (newPrice, discount, partNo))

        elif priceType == "Buying":
            query = """
                UPDATE tbl_stockparts
                SET 
                    Date = NOW(),
                    UnitCost = %s 
                WHERE CarPart = (
                    SELECT ID FROM tbl_carpartnames WHERE PartNo = %s
                )
            """
            cursor.execute(query, (newPrice, partNo))


# *************************************************** Repair Priorities Details Section ************************************


@anvil.server.callable()
def getCarRegNo(search_text=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT DISTINCT RegNo
            FROM tbl_jobcarddetails
            WHERE RegNo LIKE %s
            ORDER BY RegNo ASC
        """
        like_pattern = f"%{search_text}%"
        cursor.execute(query, (like_pattern,))
        results = [row[0] for row in cursor.fetchall()]
        return results


@anvil.server.callable()
def getPriorityList(regNo=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT RegNo, PartName, PartNumber, Quantity, Amount, Priority
            FROM tbl_repairpriorities
            WHERE RegNo = %s
        """
        cursor.execute(query, (regNo,))
        rows = cursor.fetchall()

        results = []
        for row in rows:
            results.append(
                {
                    "Name": row[1],
                    "Number": row[2],
                    "Quantity": row[3],
                    "Amount": f"{float(row[4]):,.2f}",
                    "Priority": row[5],
                }
            )

        return results


@anvil.server.callable()
def deleteCurrentPriority(regNo):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            DELETE FROM tbl_repairpriorities WHERE RegNo = %s
        """
        cursor.execute(query, (regNo,))


@anvil.server.callable()
def savePriority(assignedDate, regNo, name, number, quantity, amount, priority):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO 
                tbl_repairpriorities(Date,RegNo,PartName,PartNumber,Quantity,Amount,Priority) 
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query, (assignedDate, regNo, name, number, quantity, amount, priority)
        )


@anvil.server.callable()
def get_priority_list(regNo):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Date, RegNo, PartName, PartNumber, Quantity, Amount, Priority
            FROM tbl_repairpriorities
            WHERE RegNo = %s 
            ORDER BY Priority DESC
        """
        cursor.execute(query, (regNo,))
        rows = cursor.fetchall()

        urgent = []
        can_wait = []

        for count, r in enumerate(rows, start=1):

            def safe_float_convert(value):
                if value is None:
                    return None
                if isinstance(value, (int, float, decimal.Decimal)):
                    return float(value)
                if isinstance(value, str):
                    try:
                        return float(value.replace(",", ""))
                    except ValueError:
                        print(f"Warning: Could not convert string '{value}' to float.")
                        return None
                return None

            quantity_issued_val = safe_float_convert(r[4])
            amount_val = safe_float_convert(r[5])
            display_quantity_issued = (
                ""
                if r[4] is None or not isinstance(r[4], (int, float, decimal.Decimal))
                else quantity_issued_val
            )

            total_calc = None
            if quantity_issued_val is None:
                total_calc = amount_val
            elif amount_val is not None:
                total_calc = round(quantity_issued_val * amount_val, 2)

            entry = {
                "No": count,
                "Date": r[0],
                "RegNo": r[1],
                "PartName": r[2],
                "PartNumber": r[3],
                "Quantity": display_quantity_issued,
                "Amount": amount_val,
                "Total": total_calc,
                "Priority": r[6],
            }

            if str(r[6]).strip().lower() == "urgent":
                urgent.append(entry)
            elif str(r[6]).strip().lower() == "can wait":
                can_wait.append(entry)

        return {"Urgent": urgent, "CanWait": can_wait}


@anvil.server.callable()
def fillFormData(regNo, docType, logo_path: str = os.getenv("LOGO"), font_path: str = os.getenv("FONT_PATH")) -> str:
    _get_current_user()
    if docType == "Priority":
        docTitle = "Repair Priority List"
        priorityDetails = get_priority_list(regNo)
        # Get Date and RegNo
        if priorityDetails["Urgent"]:
            currentDate = priorityDetails["Urgent"][0]["Date"]
        else:
            currentDate = priorityDetails["CanWait"][0]["Date"]

    # === Embed MozillaHeadline font as base64 ===
    font_base64 = ""
    if font_path and os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # === Handle company logo ===
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
        logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" alt="Company Logo" style="width: 100%; height: 100%; border-radius: 2px;">'
    else:
        logo_img_tag = "LOGO"

    # Generate items table rows
    urgent_header = "Urgent"
    items_html_urgent = ""
    # Calculate sub total
    sub_total_urgent = sum(float(item["Total"]) for item in priorityDetails["Urgent"])
    canwait_header = "Can Wait"
    items_html_canwait = ""
    # Calculate sub total
    sub_total_canwait = sum(float(item["Total"]) for item in priorityDetails["CanWait"])

    for item in priorityDetails["Urgent"]:
        items_html_urgent += f"""
                    <tr class="item-row">
                        <td>{item['No']}</td>
                        <td>{item['PartName']}</td>
                        <td>{item['Quantity']}</td>
                        <td>{f"{item['Amount']:,.2f}"}</td>
                        <td>{f"{item['Total']:,.2f}"}</td>
                    </tr>"""

    for item in priorityDetails["CanWait"]:
        items_html_canwait += f"""
                    <tr class="item-row">
                        <td>{item['No']}</td>
                        <td>{item['PartName']}</td>
                        <td>{item['Quantity']}</td>
                        <td>{item['Amount']}</td>
                        <td>{item['Total']}</td>
                    </tr>"""

    if docType in ("Priority"):
        reportNotes = """
                    <div class="notes-section">
                        <div class="notes-title">NOTE: THE ABOVE ESTIMATE IS SUBJECT TO REVIEW DUE TO:</div>
                        <ol class="notes-list">
                            <li>Price change at the time of actual repair</li>
                            <li>Further damages found during repairs</li>
                            <li>100% Deposit on imported parts</li>
                            <li>70% deposit on local parts on commencement</li>
                        </ol>
                    </div>
            """
    else:
        reportNotes = ""

        # Complete HTML template with fixed structure
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{docTitle}</title>
    <style>
        /* --- Embed Mozilla Headline font --- */
        @font-face {{
            font-family: 'Mozilla Headline';
            src: url(data:font/ttf;base64,{font_base64}) format('truetype');
        }}
        body {{
            font-family: Roboto, Noto, Arial, sans-serif;
            font-size: 14px;
            line-height: 1.4286;
            background-color: #fafafa;
            margin: 0;
            padding: 16px;
        }}

        .quotation-container {{
            background-color: white;
            border-radius: 2px;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
            max-width: 800px;
            margin: 0 auto;
            overflow: hidden;
        }}

        .logo-section {{
            text-align: center;
            padding: 24px;
            background-color: white;
            border-bottom: 1px solid #e0e0e0;
        }}

        .logo-container {{
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 20px;
            margin-bottom: 16px;
        }}

        .logo-image {{
            width: 725px;
            height: 100px;
            background: linear-gradient(135deg, #228B22, #90EE90, #FFD700, #FF6347);
            border-radius: 2px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
            color: white;
            font-weight: 500;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}

        .header {{
            background-color: #000;
            color: white;
            text-align: center;
            padding: 16px 24px;
            font-size: 16px;
            font-weight: 300;
            letter-spacing: .5px;
            box-shadow: 0 4px 5px 0 rgba(0, 0, 0, 0.14),
                        0 1px 10px 0 rgba(0, 0, 0, 0.12),
                        0 2px 4px -1px rgba(0, 0, 0, 0.2);
        }}

        .detail-row {{
            display: grid;
            grid-template-columns: 140px 1fr; /* label column, value column */
            column-gap: 8px;
            margin-bottom: 12px;
        }}

        .detail-label {{
                font-weight: bold;
                font-size: 16px;
                color: rgba(0,0,0,0.87);
                
        }}

        .detail-value {{
            font-size: 16px;
            color: rgba(0,0,0,0.87);
            text-align: left;
        }}

        .items-table {{
            border-collapse: collapse;
            width: 100%;
            margin: 0 24px 24px 0;
            background-color: white;
            border-radius: 2px;
            overflow: hidden;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}

        .items-table th {{
            background-color: #f5f5f5;
            border-bottom: 1px solid #e0e0e0;
            padding: 16px;
            text-align: left;
            font-weight: bold;
            font-size: 14px;
            color: rgba(0,0,0,0.87);
            text-transform: uppercase;
            letter-spacing: .5px;
        }}

        .items-table td {{
            border-bottom: 1px solid rgba(0,0,0,0.12);
            padding: 16px;
            font-size: 14px;
            color: rgba(0,0,0,0.87);
        }}

        .items-table .item-row:hover {{
            background-color: rgba(0,0,0,0.04);
        }}

        .total-row {{
            background-color: #000 !important;
            color: white !important;
        }}

        .total-row td {{
            border-bottom: none !important;
            font-weight: 300;
            font-size: 16px;
            color: white !important;
            padding: 16px;
        }}

        .notes-section {{
            padding: 24px;
            background-color: #f5f5f5;
            margin-top: 16px;
        }}

        .notes-title {{
            margin-bottom: 16px;
            font-weight: 500;
            font-size: 16px;
            color: rgba(0,0,0,0.87);
            font-family: 'Mozilla Headline';
        }}

        .notes-list {{
            margin: 0;
            padding-left: 24px;
            color: rgba(0,0,0,0.74);
        }}

        .notes-list li {{
            margin-bottom: 8px;
            line-height: 1.5;
            font-family: 'Mozilla Headline';
        }}
        #footer  div {{
                width: 80%;
                margin: 0 auto;
                text-align: center;
                font-size: 12px;
            font-family: 'Mozilla Headline';
            }}
    </style>
</head>
<body>
    <div class="quotation-container">
        <div class="logo-section">
            <div class="logo-container">
                <div class="logo-image">
                    {logo_img_tag}
                </div>
            </div>
        </div>

        <div class="header">
            {docTitle.upper()}
        </div>

        <!-- UPDATED: Replaced details-section div with table layout -->
        <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
            <tr>
                <!-- Left Column -->
                <td style="width: 50%; vertical-align: top; padding-left: 24px; padding-right: 32px;">
                    <div class="detail-row">
                        <span class="detail-label">Date:</span>
                        <div>
                        <span class="detail-value">{currentDate}</span>
                        </div>
                    </div>
                </td>

                <!-- Right Column -->
                <td style="width: 50%; vertical-align: top; padding-left: 32px;">
                    <div class="detail-row">
                        <span class="detail-label">Reg No:</span>
                        <div>
                        <span class="detail-value">{regNo}</span>
                        </div>
                    </div>
                    
                </td>
            </tr>
        </table>
        <!-- END UPDATED -->
        
        <h3 style="text-align: center;"> {urgent_header}</h3>
        <table class="items-table">
            <thead>
                <tr>
                    <th>No.</th>
                    <th>Item</th>
                    <th>Quantity</th>
                    <th>Amount (Kshs)</th>
                    <th>Total (Kshs)</th>
                </tr>
            </thead>
            <tbody>
                {items_html_urgent}
                <tr class="total-row">
                    <td colspan="4" style="text-align: right; font-weight: 500;">Total</td>
                    <td style="font-weight: 500;">{sub_total_urgent:,.2f}</td>
                </tr>
            </tbody>
        </table>
        
        <h3 style="text-align: center;"> {canwait_header}</h3>
        <table class="items-table">
            <thead>
                <tr>
                    <th>No.</th>
                    <th>Item</th>
                    <th>Quantity</th>
                    <th>Amount (Kshs)</th>
                    <th>Total (Kshs)</th>
                </tr>
            </thead>
            <tbody>
                {items_html_canwait}
                <tr class="total-row">
                    <td colspan="4" style="text-align: right; font-weight: 500;">Total</td>
                    <td style="font-weight: 500;">{sub_total_canwait:,.2f}</td>
                </tr>
            </tbody>
        </table>
    {reportNotes}   
    <footer id="footer">
        <div> 
            <p>Joy Is The Feeling Of Being Looked After By The Best - BMW CENTER For Your BMW.</p>
        </div>
    </footer>      
    </div>
    </body>
    </html>"""

    return html_content


@anvil.server.callable()
def fillDefectsListFormData(
    jobCardID, docType, logo_path: str = os.getenv("LOGO"),font_path: str = os.getenv("FONT_PATH")
) -> str:
    _get_current_user()
    if docType == "DefectsList":
        docTitle = "Defects List"
        defectsdetails = get_defects_list_details_by_job_id(jobCardID)
    elif docType == "PricedDefectsList":
        docTitle = "Priced Defects List"
        defectsdetails = get_priced_defects_list_details_by_job_id(jobCardID)

    # === Embed MozillaHeadline font as base64 ===
    font_base64 = ""
    if font_path and os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # === Handle company logo ===
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
        logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" alt="Company Logo" style="width: 100%; height: 100%; border-radius: 2px;">'
    else:
        logo_img_tag = "LOGO"

    # === Generate defects rows — escape DB content to prevent XSS ===
    items_html = ""
    for item in defectsdetails["Defects"]:
        items_html += f"""
            <tr class="item-row">
                <td>{html_module.escape(str(item["No"]))}</td>
                <td>{html_module.escape(str(item["Defect"]))}</td>
            </tr>"""

    # === Signature ===
    signature_html = f"""
    <img src="data:image/png;base64,{defectsdetails["Signature"]}" 
         alt="Technician Signature" 
         style="max-width: 90%; height: auto; display: block; border: 1px solid #ccc; padding: 4px;"/>
    """

    # === Full HTML ===
    html_content = f"""<!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>{docTitle}</title>
            <style>
                /* --- Embed Mozilla Headline font --- */
                @font-face {{
                    font-family: 'Mozilla Headline';
                    src: url(data:font/ttf;base64,{font_base64}) format('truetype');
                }}

                body {{
                    font-family: 'Roboto', 'Noto', Arial, sans-serif;
                    font-size: 14px;
                    line-height: 1.4286;
                    background-color: #fafafa;
                    margin: 0;
                    padding: 16px;
                }}

                .quotation-container {{
                    background-color: white;
                    border-radius: 2px;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                    max-width: 800px;
                    margin: 0 auto;
                    overflow: hidden;
                }}

                .logo-section {{
                    text-align: center;
                    padding: 24px;
                    background-color: white;
                    border-bottom: 1px solid #e0e0e0;
                }}

                .logo-container {{
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    gap: 20px;
                    margin-bottom: 16px;
                }}

                .logo-image {{
                    width: 725px;
                    height: 100px;
                    background: linear-gradient(135deg, #228B22, #90EE90, #FFD700, #FF6347);
                    border-radius: 2px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-size: 24px;
                    color: white;
                    font-weight: 500;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                }}

                .header {{
                    background-color: #000;
                    color: white;
                    text-align: center;
                    padding: 16px 24px;
                    font-size: 16px;
                    font-weight: 300;
                    letter-spacing: .5px;
                    box-shadow: 0 4px 5px 0 rgba(0, 0, 0, 0.14),
                                0 1px 10px 0 rgba(0, 0, 0, 0.12),
                                0 2px 4px -1px rgba(0, 0, 0, 0.2);
                }}

                .detail-row {{
                    display: grid;
                    grid-template-columns: 140px 1fr;
                    column-gap: 8px;
                    margin-bottom: 12px;
                }}

                .detail-label {{
                    font-weight: bold;
                    font-size: 16px;
                    color: rgba(0,0,0,0.87);
                }}

                .detail-value {{
                    font-size: 16px;
                    color: rgba(0,0,0,0.87);
                    text-align: left;
                }}

                .items-table {{
                    border-collapse: collapse;
                    width: 100%;
                    margin: 0 24px 24px 0;
                    background-color: white;
                    border-radius: 2px;
                    overflow: hidden;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                }}

                .items-table th {{
                    background-color: #f5f5f5;
                    border-bottom: 1px solid #e0e0e0;
                    padding: 16px;
                    text-align: left;
                    font-weight: bold;
                    font-size: 14px;
                    color: rgba(0,0,0,0.87);
                    text-transform: uppercase;
                    letter-spacing: .5px;
                }}

                .items-table td {{
                    border-bottom: 1px solid rgba(0,0,0,0.12);
                    padding: 16px;
                    font-size: 14px;
                    color: rgba(0,0,0,0.87);
                }}

                .items-table .item-row:hover {{
                    background-color: rgba(0,0,0,0.04);
                }}

                .notes-title, .notes-list li, #footer div {{
                    font-family: 'Mozilla Headline', Arial, sans-serif;
                }}

                #footer div {{
                    width: 80%;
                    margin: 0 auto;
                    text-align: center;
                    font-size: 12px;
                }}
            </style>
        </head>
        <body>
            <div class="quotation-container">
                <div class="logo-section">
                    <div class="logo-container">
                        <div class="logo-image">
                            {logo_img_tag}
                        </div>
                    </div>
                </div>

                <div class="header">{docTitle.upper()}</div>

                <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
                    <tr>
                        <td style="width: 50%; vertical-align: top; padding-left: 24px; padding-right: 32px;">
                            <div class="detail-row"><span class="detail-label">Customer Name:</span><div><span class="detail-value">{defectsdetails['ClientName']}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Make And Model:</span><div><span class="detail-value">{defectsdetails['MakeAndModel']}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Reg No:</span><div><span class="detail-value">{defectsdetails['RegNo']}</span></div></div>
                        </td>
                        <td style="width: 50%; vertical-align: top; padding-left: 32px;">
                            <div class="detail-row"><span class="detail-label">Engine:</span><div><span class="detail-value">{defectsdetails['EngineCode']}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Chassis:</span><div><span class="detail-value">{defectsdetails['ChassisNo']}</span></div></div>
                            <div class="detail-row"><span class="detail-label">Date:</span><div><span class="detail-value">{defectsdetails['ReceivedDate']}</span></div></div>
                        </td>
                    </tr>
                </table>

                <table class="items-table">
                    <thead><tr><th>No.</th><th>Defects</th></tr></thead>
                    <tbody>{items_html}</tbody>
                </table>

                <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
                    <tr>
                        <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                            <div class="detail-row"><span class="detail-label">Prepared By:</span><div><span class="detail-value">{defectsdetails['Staff']}</span></div></div>
                        </td>
                        <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                            <div class="detail-row">
                                <span class="detail-label">Signature:</span>
                                <div style="width: 100%; overflow: hidden;">
                                    <span class="detail-value" style="display: block; width: 100%;">
                                        {signature_html}
                                    </span>
                                </div>
                            </div>
                        </td>
                    </tr>
                </table>

                <footer id="footer">
                    <div><p>Joy Is The Feeling Of Being Looked After By The Best - BMW CENTER For Your BMW.</p></div>
                </footer>
            </div>
        </body>
        </html>
    """

    return html_content


@anvil.server.callable()
def downloadDefectsPdfForm(jobcardID, docType):
    _get_current_user()
    try:
        result = anvil.server.call("getJobCardRow", jobcardID)
        docName = result["RegNo"]

        if docType == "DefectsList":
            fileName = str(docName) + " Defects List"
        elif docType == "PricedDefectsList":
            fileName = str(docName) + " Priced Defects List"

        setting_options = {
            "encoding": "UTF-8",
            "custom-header": [("Accept-Encoding", "gzip")],
            "page-size": "A4",
            "orientation": "Portrait",
            "margin-top": "0.75in",
            "margin-right": "0.75in",
            "margin-bottom": "0.75in",
            "margin-left": "0.75in",
            "no-outline": False,
            "enable-local-file-access": None,
        }

        if docType == "DefectsList" or docType == "PricedDefectsList":
            html_string = fillDefectsListFormData(jobcardID, docType)

        pdfkit.from_string(
            html_string,
            fileName,
            options={**setting_options, "debug-javascript": ""},
            configuration=config,
        )

        media_object = anvil.media.from_file(fileName, "application/pdf", name=fileName)
        return media_object

    except Exception as e:
        print("PDF generation failed:", str(e))
        raise


@anvil.server.callable()
def downloadRevisionPdfForm(regNo, docType):
    _get_current_user()
    try:
        docName = regNo
        if docType == "Priority":
            fileName = str(docName) + " Priority"
        elif docType == "Brand":
            fileName = str(docName) + " Brand"

        setting_options = {
            "encoding": "UTF-8",
            "custom-header": [("Accept-Encoding", "gzip")],
            "page-size": "A4",
            "orientation": "Portrait",
            "margin-top": "0.75in",
            "margin-right": "0.75in",
            "margin-bottom": "0.75in",
            "margin-left": "0.75in",
            "no-outline": False,
            "enable-local-file-access": None,
        }

        if docType == "Priority":
            html_string = fillFormData(regNo, docType)
        elif docType == "Brand":
            html_string = fillBrandComparisonFormData(regNo, docType)

        pdfkit.from_string(
            html_string, fileName, options=setting_options, configuration=config
        )
        media_object = anvil.media.from_file(fileName, "application/pdf", name=fileName)
        return media_object

    except Exception as e:
        print("PDF generation failed:", str(e))
        raise


@anvil.server.callable()
def deleteRevisionFile(regNo, docType):
    # Construct file name based on docType
    _get_current_user()
    if docType == "Priority":
        file_name = f"{regNo} Priority"
    elif docType == "Brand":
        file_name = f"{regNo} Brand"
    else:
        return f"Unsupported document type: {docType}"

    # Optional: define file path
    # file_path = os.path.join("/your/path", file_name)
    file_path = file_name  # if no custom directory

    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            return f"File '{file_name}' deleted successfully."
        else:
            return f"File '{file_name}' does not exist."
    except Exception as e:
        return f"Error deleting file: {e}"


# *************************************************** Brand Comparison Details Section ************************************


@anvil.server.callable()
def getBrandComparisonList(regNo=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT RegNo, PartName, PartNumber, Quantity, Amount, GroupID
            FROM tbl_brandcomparison
            WHERE RegNo = %s
        """
        cursor.execute(query, (regNo,))
        rows = cursor.fetchall()

        results = []
        for row in rows:
            results.append(
                {
                    "Name": row[1],
                    "Number": row[2],
                    "Quantity": row[3],
                    "Amount": (
                        "TO BE CONFIRMED"
                        if float(row[4]) == 0
                        else f"{float(row[4]):,.2f}"
                    ),
                    "GroupID": row[5],
                }
            )

        return results


@anvil.server.callable()
def get_brand_list(regNo):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Date, RegNo, PartName, PartNumber, Quantity, Amount, GroupID
            FROM tbl_brandcomparison
            WHERE RegNo = %s 
            ORDER BY GroupID ASC
        """
        cursor.execute(query, (regNo,))
        rows = cursor.fetchall()

        entry = []
        for count, r in enumerate(rows, start=1):

            def safe_float_convert(value):
                if value is None:
                    return None
                if isinstance(value, (int, float, decimal.Decimal)):
                    return float(value)
                if isinstance(value, str):
                    try:
                        return float(value.replace(",", ""))
                    except ValueError:
                        print(f"Warning: Could not convert string '{value}' to float.")
                        return None
                return None

            quantity_issued_val = safe_float_convert(r[4])
            amount_val = safe_float_convert(r[5])
            display_quantity_issued = (
                ""
                if r[4] is None or not isinstance(r[4], (int, float, decimal.Decimal))
                else quantity_issued_val
            )

            total_calc = None
            if quantity_issued_val is None:
                total_calc = amount_val
            elif amount_val is not None:
                total_calc = round(quantity_issued_val * amount_val, 2)

            entry.append(
                {
                    "No": count,
                    "Date": r[0],
                    "RegNo": r[1],
                    "PartName": r[2],
                    "PartNumber": r[3],
                    "Quantity": display_quantity_issued,
                    "Amount": amount_val,
                    "Total": total_calc,
                    "GroupID": r[6],
                }
            )

        return entry


@anvil.server.callable()
def fillBrandComparisonFormData(
    regNo, docType, logo_path: str = os.getenv("LOGO"),font_path: str = os.getenv("FONT_PATH")
) -> str:
    _get_current_user()
    if docType == "Brand":
        docTitle = "Brand Comparison List"
        comparisonDetails = get_brand_list(regNo)

    # === Embed MozillaHeadline font as base64 ===
    font_base64 = ""
    if font_path and os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # === Handle company logo ===
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
        logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" alt="Company Logo" style="width: 100%; height: 100%; border-radius: 2px;">'
    else:
        logo_img_tag = "LOGO"

    currentDate = comparisonDetails[0]["Date"]

    # Calculate sub total
    # sub_total = sum(float(item['Total']) for item in comparisonDetails)

    # Group items by GroupID
    grouped_data = defaultdict(list)
    for item in comparisonDetails:
        grouped_data[item["GroupID"]].append(item)

    items_html = ""
    grand_total = 0

    for group_id, group_items in grouped_data.items():
        group_total = sum(item["Total"] or 0 for item in group_items)
        grand_total += group_total

        for item in group_items:
            if item["Amount"] == 0:  # Implies To Be Confirmed
                textAmount = "TO BE CONFIRMED"
            else:
                textAmount = f"{item['Amount']:,.2f}"
            if item["Total"] == 0:  # Implies To Be Confirmed
                textTotal = "TO BE CONFIRMED"
            else:
                textTotal = f"{item['Total']:,.2f}"

            items_html += f"""
                <tr class="item-row">
                    <td>{html_module.escape(str(item['No']))}</td>
                    <td>{html_module.escape(str(item['PartName']))}</td>
                    <td>{html_module.escape(str(item['Quantity']))}</td>
                    <td>{textAmount}</td>
                    <td>{textTotal}</td>
                </tr>
            """

        items_html += f"""
            <tr class="total-row">
                <td colspan="4" style="text-align: right;">Sub Total</td>
                <td>{group_total:,.2f}</td>
            </tr>
            
        """

    # Replace sub_total with grand_total
    sub_total = grand_total

    if docType in ("Brand"):
        reportNotes = """
                    <div class="notes-section">
                        <div class="notes-title">NOTE: THE ABOVE ESTIMATE IS SUBJECT TO REVIEW DUE TO:</div>
                        <ol class="notes-list">
                            <li>Price change at the time of actual repair</li>
                            <li>Further damages found during repairs</li>
                            <li>100% Deposit on imported parts</li>
                            <li>70% deposit on local parts on commencement</li>
                        </ol>
                    </div>
            """
    else:
        reportNotes = ""

        # Complete HTML template with fixed structure
    html_content = f"""
    <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>{docTitle}</title>
            <style>
                /* --- Embed Mozilla Headline font --- */
                @font-face {{
                    font-family: 'Mozilla Headline';
                    src: url(data:font/ttf;base64,{font_base64}) format('truetype');
                }}
                body {{
                    font-family: Roboto, Noto, Arial, sans-serif;
                    font-size: 14px;
                    line-height: 1.4286;
                    background-color: #fafafa;
                    margin: 0;
                    padding: 16px;
                }}

                .quotation-container {{
                    background-color: white;
                    border-radius: 2px;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                    max-width: 800px;
                    margin: 0 auto;
                    overflow: hidden;
                }}

                .logo-section {{
                    text-align: center;
                    padding: 24px;
                    background-color: white;
                    border-bottom: 1px solid #e0e0e0;
                }}

                .logo-container {{
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    gap: 20px;
                    margin-bottom: 16px;
                }}

                .logo-image {{
                    width: 725px;
                    height: 100px;
                    background: linear-gradient(135deg, #228B22, #90EE90, #FFD700, #FF6347);
                    border-radius: 2px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-size: 24px;
                    color: white;
                    font-weight: 500;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                }}

                .header {{
                    background-color: #000;
                    color: white;
                    text-align: center;
                    padding: 16px 24px;
                    font-size: 16px;
                    font-weight: 300;
                    letter-spacing: .5px;
                    box-shadow: 0 4px 5px 0 rgba(0, 0, 0, 0.14),
                                0 1px 10px 0 rgba(0, 0, 0, 0.12),
                                0 2px 4px -1px rgba(0, 0, 0, 0.2);
                }}

                .detail-row {{
                    display: grid;
                    grid-template-columns: 140px 1fr; /* label column, value column */
                    column-gap: 8px;
                    margin-bottom: 12px;
                }}

                .detail-label {{
                        font-weight: bold;
                        font-size: 16px;
                        color: rgba(0,0,0,0.87);
                }}

                .detail-value {{
                    font-size: 16px;
                    color: rgba(0,0,0,0.87);
                    text-align: left;
                }}

                .items-table {{
                    border-collapse: collapse;
                    width: 100%;
                    margin: 0 24px 24px 0;
                    background-color: white;
                    border-radius: 2px;
                    overflow: hidden;
                    box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                                0 3px 1px -2px rgba(0, 0, 0, 0.2),
                                0 1px 5px 0 rgba(0, 0, 0, 0.12);
                }}

                .items-table th {{
                    background-color: #f5f5f5;
                    border-bottom: 1px solid #e0e0e0;
                    padding: 16px;
                    text-align: left;
                    font-weight: bold;
                    font-size: 14px;
                    color: rgba(0,0,0,0.87);
                    text-transform: uppercase;
                    letter-spacing: .5px;
                }}

                .items-table td {{
                    border-bottom: 1px solid rgba(0,0,0,0.12);
                    padding: 16px;
                    font-size: 14px;
                    color: rgba(0,0,0,0.87);
                }}

                .items-table .item-row:hover {{
                    background-color: rgba(0,0,0,0.04);
                }}

                .total-row {{
                    background-color: #000 !important;
                    color: white !important;
                }}

                .total-row td {{
                    border-bottom: none !important;
                    font-weight: 300;
                    font-size: 16px;
                    color: white !important;
                    padding: 16px;
                }}

                .notes-section {{
                    padding: 24px;
                    background-color: #f5f5f5;
                    margin-top: 16px;
                }}

                .notes-title {{
                    margin-bottom: 16px;
                    font-weight: 500;
                    font-size: 16px;
                    color: rgba(0,0,0,0.87);
                    font-family: 'Mozilla Headline';
                }}

                .notes-list {{
                    margin: 0;
                    padding-left: 24px;
                    color: rgba(0,0,0,0.74);
                }}

                .notes-list li {{
                    margin-bottom: 8px;
                    line-height: 1.5;
                    font-family: 'Mozilla Headline';
                }}
                #footer  div {{
                        width: 80%;
                        margin: 0 auto;
                        text-align: center;
                        font-size: 12px;
                        font-family: 'Mozilla Headline';
                    }}
            </style>
        </head>
        <body>
            <div class="quotation-container">
                <div class="logo-section">
                    <div class="logo-container">
                        <div class="logo-image">
                            {logo_img_tag}
                        </div>
                    </div>
                </div>

                <div class="header">
                    {docTitle.upper()}
                </div>

                <!-- UPDATED: Replaced details-section div with table layout -->
                <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
                    <tr>
                        <!-- Left Column -->
                        <td style="width: 50%; vertical-align: top; padding-left: 24px; padding-right: 32px;">
                            <div class="detail-row">
                                <span class="detail-label">Date:</span>
                                <div>
                                <span class="detail-value">{currentDate}</span>
                                </div>
                            </div>
                        </td>

                        <!-- Right Column -->
                        <td style="width: 50%; vertical-align: top; padding-left: 32px;">
                            <div class="detail-row">
                                <span class="detail-label">Reg No:</span>
                                <div>
                                <span class="detail-value">{regNo}</span>
                                </div>
                            </div>

                        </td>
                    </tr>
                </table>
                <!-- END UPDATED -->

                <table class="items-table">
                    <thead>
                        <tr>
                            <th>No.</th>
                            <th>Item</th>
                            <th>Quantity</th>
                            <th>Amount (Kshs)</th>
                            <th>Total (Kshs)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {items_html}
                    </tbody>
                </table>

            {reportNotes} 
            <footer id="footer">
                <div> 
                    <p>Joy Is The Feeling Of Being Looked After By The Best - BMW CENTER For Your BMW.</p>
                </div>
            </footer>        
            </div>
            </body>
            </html>"""

    return html_content


@anvil.server.callable()
def deleteCurrentBrandComparison(regNo):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            DELETE FROM tbl_brandcomparison WHERE RegNo = %s
        """
        cursor.execute(query, (regNo,))


@anvil.server.callable()
def saveBrand(assignedDate, regNo, name, number, quantity, amount, groupid):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO 
                tbl_brandcomparison(Date,RegNo,PartName,PartNumber,Quantity,Amount,GroupID) 
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query, (assignedDate, regNo, name, number, quantity, amount, groupid)
        )


# *************************************************** Location Details Section ************************************


@anvil.server.callable()
def getLocation():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Location
            FROM tbl_carpartslocation
            ORDER BY Location ASC
        """
        cursor.execute(query)
        results = cursor.fetchall()

        # Return as a list of dicts for Anvil Dropdown
        # label = what user sees, value = what is stored
        dropdown_items = [(row[1], row[0]) for row in results]

    return dropdown_items


@anvil.server.callable()
def addLocation(locationName):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_carpartslocation (Location) VALUES (%s)
        """
        cursor.execute(query, (locationName,))


@anvil.server.callable()
def getLocationName(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Location 
            FROM tbl_carpartslocation 
            WHERE ID = %s
        """
        cursor.execute(query, (valueID,))
        result = cursor.fetchone()
        if result:
            return result[1]
        return None


@anvil.server.callable()
def updateLocation(value, id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_carpartslocation
            SET Location = %s
            WHERE ID = %s
        """
        cursor.execute(query, (value, id))


# *************************************************** Supplier Details Section ************************************


@anvil.server.callable()
def getSupplier():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT ID, Name
            FROM tbl_carpartssupplier
            Order By Name ASC
        """
        cursor.execute(query)
        results = cursor.fetchall()

        # Return as a list of dicts for Anvil Dropdown
        # label = what user sees, value = what is stored
        dropdown_items = [(row[1], row[0]) for row in results]

    return dropdown_items


@anvil.server.callable()
def addSupplier(supplierName, supplierPhone):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_carpartssupplier (Name, Phone) VALUES (%s, %s)
        """
        cursor.execute(query, (supplierName, supplierPhone))


@anvil.server.callable()
def getSupplierDetails(supplierID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Name, Phone 
            FROM tbl_carpartssupplier 
            WHERE ID = %s
        """
        cursor.execute(query, (supplierID,))
        result = cursor.fetchone()  # Get one row
        if result:
            return {"Name": result[0], "Phone": result[1]}
        else:
            return None


@anvil.server.callable()
def updateSupplier(supplierName, supplierPhone, supplierID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_carpartssupplier
            SET Name = %s, Phone = %s
            WHERE ID = %s
        """
        cursor.execute(query, (supplierName, supplierPhone, supplierID))


# *************************************************** New Car Parts Details Section ************************************


@anvil.server.callable()
def addNewParts(
    purchaseDate,
    partName,
    partNumber,
    locationID,
    supplierID,
    units,
    buyingPrice,
    sellingPrice,
    discountPrice,
    reorderLevel,
):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO 
            tbl_carpartnames(Name,PartNo,OrderLevel,Location) 
            VALUES (%s, %s, %s, %s)
        """
        cursor.execute(query, (partName, partNumber, reorderLevel, locationID))

        partID = cursor.lastrowid

        query2 = """
            INSERT INTO 
            tbl_partssellingprice(SetPriceDate, CarPartsNamesID, Amount, SaleDiscount) 
            VALUES (%s,%s,%s,%s)
        """
        cursor.execute(query2, (purchaseDate, partID, sellingPrice, discountPrice))

        query3 = """
          INSERT INTO 
          tbl_stockparts(Date, CarPart, CarPartsSupplierId, NoOfUnits, UnitCost, Narration) 
          VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query3, (purchaseDate, partID, supplierID, units, buyingPrice, " ")
        )


@anvil.server.callable()
def check_duplicate_number(number):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
                SELECT ID FROM tbl_carpartnames 
                WHERE PartNo = %s
            """
        cursor.execute(query, (number,))

        result = cursor.fetchone()
        return result is not None


@anvil.server.callable()
def getPartsDetailsID(id):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_stockparts.Date,
                tbl_stockparts.NoOfUnits,
                tbl_stockparts.UnitCost,
                tbl_stockparts.CarPartsSupplierID,
                tbl_carpartnames.Name,
                tbl_carpartnames.PartNo,
                tbl_carpartnames.OrderLevel,
                tbl_carpartnames.Location
            FROM tbl_carpartnames
            INNER JOIN tbl_stockparts
                ON tbl_carpartnames.ID = tbl_stockparts.CarPart
            WHERE tbl_carpartnames.ID = %s
            ORDER BY tbl_stockparts.Date DESC
            LIMIT 1
        """
        cursor.execute(query, (id,))
        row = cursor.fetchone()

    if row:
        columns = [
            "Date",
            "NoOfUnits",
            "UnitCost",
            "CarPartsSupplierID",
            "Name",
            "PartNo",
            "OrderLevel",
            "Location",
        ]
        return dict(zip(columns, row))
    else:
        return None


@anvil.server.callable()
def getPartsSellingDetailsID(valueID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT SetPriceDate, Amount, SaleDiscount FROM tbl_partssellingprice 
            WHERE CarPartsNamesID = %s
        """
        cursor.execute(query, (valueID,))
        row = cursor.fetchone()

    if row:
        columns = ["SetPriceDate", "Amount", "SaleDiscount"]
        return dict(zip(columns, row))
    else:
        return None


@anvil.server.callable()
def updateNewParts(
    purchaseDate,
    partName,
    partNumber,
    locationID,
    supplierID,
    units,
    buyingPrice,
    sellingPrice,
    discountPrice,
    reorderLevel,
    valueID,
    oldSupplierID,
):
    _get_current_user()
    with db_cursor() as cursor:

        query = """
                    UPDATE tbl_carpartnames
                    SET 
                        Name = %s,
                        PartNo = %s,
                        OrderLevel = %s,
                        Location = %s 
                    WHERE ID = %s
                """
        cursor.execute(query, (partName, partNumber, reorderLevel, locationID, valueID))

        query2 = """
            INSERT INTO tbl_partssellingprice
                (SetPriceDate, CarPartsNamesID, Amount, SaleDiscount)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                SetPriceDate = VALUES(SetPriceDate),
                Amount = VALUES(Amount),
                SaleDiscount = VALUES(SaleDiscount)
        """
        cursor.execute(query2, (purchaseDate, valueID, sellingPrice, discountPrice))

        query3 = """
                  UPDATE tbl_stockparts
                  SET 
                    Date = %s, 
                    CarPart = %s,
                    CarPartsSupplierID = %s, 
                    NoOfUnits = %s, 
                    UnitCost = %s
                 WHERE CarPart = %s and CarPartsSupplierID = %s
                """
        cursor.execute(
            query3,
            (
                purchaseDate,
                valueID,
                supplierID,
                units,
                buyingPrice,
                valueID,
                oldSupplierID,
            ),
        )


@anvil.server.callable()
def addStock(additionDate, supplierID, partID, no_of_units, unit_cost):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO 
                tbl_stockparts(Date, CarPart, CarPartsSupplierID, NoOfUnits, UnitCost, Narration) 
            VALUES  
                (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query, (additionDate, partID, supplierID, no_of_units, unit_cost, "")
        )

# ***************************************************  Car Parts Where Buying Price Exceeds Selling Price Section ************************************
@anvil.server.callable()
def getPartsWhereBuyingPriceExceedsSelling(logged_in_user):
    _get_current_user()
    
    if not logged_in_user:
        return []

    with db_cursor() as cursor:
        query = """
            WITH latest_stock AS (
                SELECT sp.CarPart, sp.UnitCost
                FROM tbl_stockparts sp
                INNER JOIN (
                    SELECT CarPart, MAX(ID) AS max_id
                    FROM tbl_stockparts
                    GROUP BY CarPart
                ) latest
                ON sp.CarPart = latest.CarPart 
                AND sp.ID = latest.max_id
            )
            SELECT 
                cpn.Name,
                cpn.PartNo,
                ls.UnitCost AS Cost,
                psp.Amount AS Selling,
                psp.SaleDiscount AS Discount
            FROM tbl_carpartnames cpn
            INNER JOIN tbl_partssellingprice psp 
                ON cpn.ID = psp.CarPartsNamesID
                AND psp.Amount > -1
            INNER JOIN latest_stock ls 
                ON cpn.ID = ls.CarPart
            WHERE 
                ls.UnitCost IS NOT NULL
                AND psp.Amount IS NOT NULL
                AND ls.UnitCost > psp.Amount
            ORDER BY cpn.Name
        """

        cursor.execute(query)
        rows = cursor.fetchall()

        result = []
        for count, row in enumerate(rows, start=1):
            result.append({
                "No": count,
                "Name": row[0],
                "PartNo": row[1],
                "Cost": f"{float(row[2]):,.2f}" if row[2] is not None else None,
                "Selling": f"{float(row[3]):,.2f}" if row[3] is not None else None,
                "Discount": row[4],
                "BuyingPriceExceedsSelling": "Yes"
            })

        return result
    
# *************************************************** Stock Balance Details Section ************************************


@anvil.server.callable()
def get_car_parts_summary(search_term=""):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            WITH issued AS (
                SELECT CarPartID, SUM(QuantityIssued) AS TotalIssued
                FROM tbl_assignedcarparts
                GROUP BY CarPartID
            ),
            latest_stocktake AS (
                SELECT CarPartNameID, HarmornizedValue
                FROM (
                    SELECT 
                        CarPartNameID,
                        HarmornizedValue,
                        ROW_NUMBER() OVER (PARTITION BY CarPartNameID ORDER BY StockTakeDate DESC) AS rn
                    FROM tbl_stocktakeharmonized
                ) ranked
                WHERE rn = 1
            )
            SELECT 
                cp.ID,
                cp.Name,
                cp.PartNo,
                cp.OrderLevel,
                COALESCE(i.TotalIssued, 0) AS TotalIssued,
                COALESCE(l.HarmornizedValue, 0) AS LatestStockTake,
                MAX(sp.Date) AS LastStockDate,
                SUM(sp.NoOfUnits) AS TotalStock,
                (COALESCE(SUM(sp.NoOfUnits), 0) 
                    - COALESCE(i.TotalIssued, 0)
                    + COALESCE(l.HarmornizedValue, 0)) AS StockBalance,
                CASE 
                    WHEN (COALESCE(SUM(sp.NoOfUnits), 0) 
                        - COALESCE(i.TotalIssued, 0)
                        + COALESCE(l.HarmornizedValue, 0)) < cp.OrderLevel
                    THEN 'Yes'
                    ELSE 'No'
                END AS BelowOrderLevel
            FROM (
                SELECT *
                FROM tbl_carpartnames
                WHERE Name LIKE %s OR PartNo LIKE %s
            ) cp
            LEFT JOIN tbl_stockparts sp ON cp.ID = sp.CarPart
            LEFT JOIN issued i ON cp.ID = i.CarPartID
            LEFT JOIN latest_stocktake l ON cp.ID = l.CarPartNameID
            GROUP BY 
                cp.ID,
                cp.Name,
                cp.PartNo,
                cp.OrderLevel,
                i.TotalIssued,
                l.HarmornizedValue
            ORDER BY cp.Name
        """

        like_pattern = f"%{search_term}%"
        cursor.execute(query, (like_pattern, like_pattern))
        rows = cursor.fetchall()

        result = []
        for idx, row in enumerate(rows, start=1):
            row_data = {
                "No": idx,
                "Name": row[1],
                "PartNo": row[2],
                "OrderLevel": row[3],
                "TotalIssued": row[4],
                "LatestStockTake": row[5],
                "TotalStock": row[7],
                "StockBalance": row[8],
                "BelowOrderLevel": row[9],
            }
            result.append(row_data)

        return result


# ***************************************************Car Parts Used Section ************************************


@anvil.server.callable()
def get_car_parts_used(job_card_id):
    _get_current_user()
    with db_cursor() as cursor:
        if job_card_id is None:
            query = """
                SELECT 
                    tbl_assignedcarparts.Date, tbl_jobcarddetails.JobCardRef, tbl_carpartnames.Name, tbl_carpartnames.PartNo, 
                    tbl_assignedcarparts.QuantityIssued, tbl_jobcarddetails.MakeAndModel
                FROM 
                    tbl_jobcarddetails 
                INNER JOIN 
                    (tbl_carpartnames 
                INNER JOIN 
                    tbl_assignedcarparts ON tbl_carpartnames.ID = tbl_assignedcarparts.CarPartID) 
                ON 
                    tbl_jobcarddetails.ID = tbl_assignedcarparts.AssignedJobID
                ORDER BY 
                    tbl_assignedcarparts.Date DESC
                """
            cursor.execute(query)

        else:
            query2 = """
                SELECT 
                    tbl_assignedcarparts.Date, tbl_jobcarddetails.JobCardRef, tbl_carpartnames.Name, tbl_carpartnames.PartNo, 
                    tbl_assignedcarparts.QuantityIssued, tbl_jobcarddetails.MakeAndModel
                FROM 
                    tbl_jobcarddetails 
                INNER JOIN 
                    (tbl_carpartnames 
                INNER JOIN 
                    tbl_assignedcarparts ON tbl_carpartnames.ID = tbl_assignedcarparts.CarPartID) 
                ON 
                    tbl_jobcarddetails.ID = tbl_assignedcarparts.AssignedJobID
                WHERE 
                    tbl_jobcarddetails.ID = %s
                ORDER BY 
                    tbl_assignedcarparts.Date DESC

            """
            cursor.execute(query2, (job_card_id,))

        rows = cursor.fetchall()
        result = []
        for idx, row in enumerate(rows, start=1):
            row_data = {
                "No": idx,
                "AssignedDate": row[0],
                "JobCardRef": row[1],
                "PartName": row[2],
                "PartNo": row[3],
                "Quantity": row[4],
                "MakeAndModel": row[5],
            }
            result.append(row_data)
        return result


@anvil.server.callable()
def get_job_card_from_car_parts_used(search_term):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT
                tbl_assignedcarparts.Date,
                tbl_jobcarddetails.JobCardRef,
                tbl_carpartnames.Name,
                tbl_carpartnames.PartNo,
                tbl_assignedcarparts.QuantityIssued,
                tbl_jobcarddetails.MakeAndModel
            FROM
                tbl_jobcarddetails
            INNER JOIN(
                    tbl_carpartnames
                INNER JOIN tbl_assignedcarparts ON tbl_carpartnames.ID = tbl_assignedcarparts.CarPartID
                )
            ON
                tbl_jobcarddetails.ID = tbl_assignedcarparts.AssignedJobID
            WHERE
                tbl_carpartnames.PartNo LIKE %s
            ORDER BY
                tbl_assignedcarparts.Date
            DESC
        """
        cursor.execute(query, (f"%{search_term}%",))
        rows = cursor.fetchall()

        result = []
        for idx, row in enumerate(rows, start=1):
            result.append(
                {
                    "No": idx,
                    "AssignedDate": row[0],
                    "JobCardRef": row[1],
                    "PartName": row[2],
                    "PartNo": row[3],
                    "Quantity": row[4],
                    "MakeAndModel": row[5],
                }
            )
        return result


# ***************************************************Assigned JobCard Details Section ************************************


@anvil.server.callable()
def get_assigned_jobs(startDate, endDate, technicianID):
    _get_current_user()
    with db_cursor() as cursor:
        if startDate is None and endDate is None and technicianID is None:
            query = """
                    SELECT 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname, 
                        SUM(
                            CASE 
                                WHEN i.QuantityIssued IS NULL THEN i.Amount
                                ELSE i.Amount * i.QuantityIssued
                            END
                        ) AS InvoicedAmount,
                        MAX(
                            CASE 
                                WHEN i.Item LIKE '%Labour%' THEN i.Amount
                                ELSE 0
                            END
                        ) AS LabourAmount,
                        jd.MakeAndModel
                    FROM 
                        tbl_invoices AS i
                    INNER JOIN 
                        tbl_jobcarddetails AS jd ON i.AssignedJobID = jd.ID
                    INNER JOIN 
                        tbl_pendingassignedjobs AS pj ON jd.ID = pj.JobCardRefID
                    INNER JOIN 
                        tbl_technicians AS t ON t.ID = pj.TechnicianID
                    GROUP BY 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname,
                        jd.MakeAndModel
                    ORDER BY 
                        jd.ReceivedDate DESC

            """
            cursor.execute(query)

        elif startDate is not None and endDate is not None and technicianID is None:
            query2 = """
                    SELECT 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname, 
                        SUM(
                            CASE 
                                WHEN i.QuantityIssued IS NULL THEN i.Amount
                                ELSE i.Amount * i.QuantityIssued
                            END
                        ) AS InvoicedAmount,
                        MAX(
                            CASE 
                                WHEN i.Item LIKE '%Labour%' THEN i.Amount
                                ELSE 0
                            END
                        ) AS LabourAmount,
                        jd.MakeAndModel
                    FROM 
                        tbl_invoices AS i
                    INNER JOIN 
                        tbl_jobcarddetails AS jd ON i.AssignedJobID = jd.ID
                    INNER JOIN 
                        tbl_pendingassignedjobs AS pj ON jd.ID = pj.JobCardRefID
                    INNER JOIN 
                        tbl_technicians AS t ON t.ID = pj.TechnicianID
                    WHERE 
                        jd.ReceivedDate BETWEEN %s AND %s 
                    GROUP BY 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname,
                        jd.MakeAndModel
                    ORDER BY 
                        jd.ReceivedDate DESC

            """
            cursor.execute(query2, (startDate, endDate))

        elif startDate is None and endDate is None and technicianID is not None:
            query3 = """
                    SELECT 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname, 
                        SUM(
                            CASE 
                                WHEN i.QuantityIssued IS NULL THEN i.Amount
                                ELSE i.Amount * i.QuantityIssued
                            END
                        ) AS InvoicedAmount,
                        MAX(
                            CASE 
                                WHEN i.Item LIKE '%Labour%' THEN i.Amount
                                ELSE 0
                            END
                        ) AS LabourAmount,
                        jd.MakeAndModel
                    FROM 
                        tbl_invoices AS i
                    INNER JOIN 
                        tbl_jobcarddetails AS jd ON i.AssignedJobID = jd.ID
                    INNER JOIN 
                        tbl_pendingassignedjobs AS pj ON jd.ID = pj.JobCardRefID
                    INNER JOIN 
                        tbl_technicians AS t ON t.ID = pj.TechnicianID
                    WHERE 
                        pj.TechnicianID = %s
                    GROUP BY 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname,
                        jd.MakeAndModel
                    ORDER BY 
                        jd.ReceivedDate DESC

            """
            cursor.execute(query3, (technicianID,))

        elif startDate is not None and endDate is not None and technicianID is not None:
            query4 = """
                    SELECT 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname, 
                        SUM(
                            CASE 
                                WHEN i.QuantityIssued IS NULL THEN i.Amount
                                ELSE i.Amount * i.QuantityIssued
                            END
                        ) AS InvoicedAmount,
                        MAX(
                            CASE 
                                WHEN i.Item LIKE '%Labour%' THEN i.Amount
                                ELSE 0
                            END
                        ) AS LabourAmount,
                        jd.MakeAndModel
                    FROM 
                        tbl_invoices AS i
                    INNER JOIN 
                        tbl_jobcarddetails AS jd ON i.AssignedJobID = jd.ID
                    INNER JOIN 
                        tbl_pendingassignedjobs AS pj ON jd.ID = pj.JobCardRefID
                    INNER JOIN 
                        tbl_technicians AS t ON t.ID = pj.TechnicianID
                    WHERE 
                        (jd.ReceivedDate BETWEEN %s AND %s) AND pj.TechnicianID = %s 
                    GROUP BY 
                        jd.ReceivedDate, 
                        jd.JobCardRef, 
                        t.Fullname,
                        jd.MakeAndModel
                    ORDER BY 
                        jd.ReceivedDate DESC

            """
            cursor.execute(query4, (startDate, endDate, technicianID))

        rows = cursor.fetchall()
        result = []
        for idx, row in enumerate(rows, start=1):
            result.append(
                {
                    "No": idx,
                    "ReceivedDate": row[0],
                    "JobCardRef": row[1],
                    "Technician": row[2],
                    "InvoicedAmount": f"{row[3]:,.2f}",
                    "LabourAmount": f"{row[4]:,.2f}",
                    "MakeAndModel": row[5],
                }
            )
        return result



# *************************************************** Export To Excel Section ************************************
@anvil.server.callable()
def make_excel(headers, rows, filename="report.xlsx", sheet_name="Data"):
    """
    Generic Excel export utility.

    :param headers: list[str] - Header titles in order
    :param rows: list[list] or list[dict] - Data rows
    :param filename: str - Filename for download
    :param sheet_name: str - Sheet name
    :return: anvil.BlobMedia
    """
    _get_current_user()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Add headers
    ws.append(headers)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Add rows
    for row in rows:
        if isinstance(row, dict):
            values = [row.get(h, "") for h in headers]
        else:
            values = row
        ws.append(values)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Freeze header row + enable auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return anvil.BlobMedia(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        buffer.read(),
        name=filename,
    )


@anvil.server.callable()
def export_assigned_jobcards(rows):
    _get_current_user()
    headers = [
        "No",
        "ReceivedDate",
        "JobCardRef",
        "Technician",
        "InvoicedAmount",
        "LabourAmount",
        "MakeAndModel",
    ]

    # Transform into list-of-dicts or list-of-lists
    processed_rows = []
    for i, row in enumerate(rows, start=1):
        processed_rows.append(
            {
                "No": i,
                "ReceivedDate": row.get("ReceivedDate", ""),
                "JobCardRef": row.get("JobCardRef", ""),
                "Technician": row.get("Technician", ""),
                "InvoicedAmount": row.get("InvoicedAmount", ""),
                "LabourAmount": row.get("LabourAmount", ""),
                "MakeAndModel": row.get("MakeAndModel", ""),
            }
        )
    return make_excel(
        headers,
        processed_rows,
        filename="assigned_jobcards.xlsx",
        sheet_name="Jobcards",
    )


@anvil.server.callable()
def export_missing_buying_prices(rows):
    _get_current_user()
    headers = ["No", "Supplier", "Name", "PartNo", "BuyingPrice"]

    # Transform into list-of-dicts or list-of-lists
    processed_rows = []
    for i, row in enumerate(rows, start=1):
        processed_rows.append(
            {
                "No": i,
                "Supplier": row.get("Supplier", ""),
                "Name": row.get("Name", ""),
                "PartNo": row.get("PartNo", ""),
                "BuyingPrice": row.get("BuyingPrice", ""),
            }
        )

    return make_excel(
        headers,
        processed_rows,
        filename="missing_buying_prices.xlsx",
        sheet_name="Buying Prices",
    )


@anvil.server.callable()
def export_missing_selling_prices(rows):
    _get_current_user()
    headers = ["No", "Supplier", "Name", "PartNo", "SellingPrice", "DiscountPrice"]

    # Transform into list-of-dicts or list-of-lists
    processed_rows = []
    for i, row in enumerate(rows, start=1):
        processed_rows.append(
            {
                "No": i,
                "Supplier": row.get("Supplier", ""),
                "Name": row.get("Name", ""),
                "PartNo": row.get("PartNo", ""),
                "SellingPrice": row.get("Amount", ""),
                "DiscountPrice": row.get("Discount", ""),
            }
        )

    return make_excel(
        headers,
        processed_rows,
        filename="missing_selling_prices.xlsx",
        sheet_name="Selling Prices",
    )


@anvil.server.callable()
def export_client_payment_details(rows):
    _get_current_user()
    headers = [
        "No",
        "JobReceivedDate",
        "JobCardRef",
        "Fullname",
        "Phone",
        "TotalPaid",
        "Discount",
        "Balance",
    ]

    # Transform into list-of-dicts or list-of-lists
    processed_rows = []
    for i, row in enumerate(rows, start=1):
        processed_rows.append(
            {
                "No": i,
                "JobReceivedDate": row.get("JobReceivedDate", ""),
                "JobCardRef": row.get("JobCardRef", ""),
                "Fullname": row.get("Fullname", ""),
                "Phone": row.get("Phone", ""),
                "TotalPaid": row.get("TotalPaid", 0),
                "Discount": row.get("Discount", 0),
                "Balance": row.get("Balance", 0),
            }
        )

    return make_excel(
        headers,
        processed_rows,
        filename="client_payment_details.xlsx",
        sheet_name="Client Payments",
    )


@anvil.server.callable()
def export_stock_balance(rows):
    _get_current_user()
    headers = [
        "No",
        "Name",
        "PartNo",
        "Reorder Level",
        "Total Stock",
        "Total Issued",
        "Variance",
        "Stock Balance",
        "Below Reorder Level",
    ]

    # Transform into list-of-dicts or list-of-lists
    processed_rows = []
    for i, row in enumerate(rows, start=1):
        processed_rows.append(
            {
                "No": i,
                "Name": row.get("Name", ""),
                "PartNo": row.get("PartNo", ""),
                "Reorder Level": row.get("OrderLevel", ""),
                "Total Stock": row.get("TotalStock", ""),
                "Total Issued": row.get("TotalIssued", ""),
                "Variance": row.get("LatestStockTake", ""),
                "Stock Balance": row.get("StockBalance", ""),
                "Below Reorder Level": row.get("BelowOrderLevel", ""),
            }
        )

    return make_excel(
        headers,
        processed_rows,
        filename="export_stock_balance.xlsx",
        sheet_name="Stock Balance",
    )


@anvil.server.callable()
def export_monthly_schedule(rows):
    _get_current_user()
    headers = [
        "No",
        "Date",
        "Name",
        "Total Amount",
        "Car Parts",
        "Engine Oil",
        "Consumable",
        "Diagnosis",
        "Labour",
        "Engine Wash",
        "Alignment",
        "T/Gearbox Oil",
        "Others",
        "Discount",
        "Balance",
    ]

    # Transform into list-of-dicts or list-of-lists
    processed_rows = []
    for i, row in enumerate(rows, start=1):
        processed_rows.append(
            {
                "No": i,
                "Date": row.get("Date", ""),
                "Name": row.get("ClientName", ""),
                "Total Amount": row.get("TotalInvoiceAmount", ""),
                "Car Parts": row.get("CarParts", ""),
                "Engine Oil": row.get("EngineOil", ""),
                "Consumable": row.get("Consumable", ""),
                "Diagnosis": row.get("Diagnosis", ""),
                "Labour": row.get("Labour", ""),
                "Engine Wash": row.get("EngineWash", ""),
                "Alignment": row.get("Alignment", ""),
                "T/Gearbox Oil": row.get("TransferGearboxOil", ""),
                "Others": row.get("Others", ""),
                "Discount": row.get("TotalDiscount", ""),
                "Balance": row.get("Balance", ""),
            }
        )

    return make_excel(
        headers,
        processed_rows,
        filename="export_monthly_schedule.xlsx",
        sheet_name="Monthly Schedule",
    )

@anvil.server.callable()
def export_current_selling_prices_and_reorder_levels():
    _get_current_user()
    headers = ["No", "Name", "PartNo", "OrderLevel", "SellingPrice", "DiscountPrice"]

    # Execute the SQL query
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_carpartnames.Name, 
                tbl_carpartnames.PartNo, 
                tbl_carpartnames.OrderLevel,
                tbl_partssellingprice.Amount, 
                tbl_partssellingprice.SaleDiscount  
            FROM tbl_carpartnames 
            INNER JOIN tbl_partssellingprice 
                ON tbl_carpartnames.ID = tbl_partssellingprice.CarPartsNamesID
            ORDER BY tbl_carpartnames.Name ASC
        """
        cursor.execute(query)
        rows = cursor.fetchall()

    # Transform into list-of-dicts
    processed_rows = []
    for i, row in enumerate(rows, start=1):
        processed_rows.append(
            {
                "No": i,
                "Name": row[0],
                "PartNo": row[1],
                "OrderLevel": row[2],
                "SellingPrice": row[3],
                "DiscountPrice": row[4],
            }
        )

    return make_excel(
        headers,
        processed_rows,
        filename="selling_prices_and_reorder_levels.xlsx",
        sheet_name="Selling Prices & Reorder Lvl",
    )

# *************************************************** Import From Excel Section ************************************

@anvil.server.callable()
def import_selling_prices_and_reorder_levels(file):
    _get_current_user()
    
    # Load the uploaded Excel file from Anvil media object
    file_bytes = BytesIO(file.get_bytes())
    wb = load_workbook(file_bytes, data_only=True)
    ws = wb.active

    # Validate headers
    expected_headers = ["No", "Name", "PartNo", "OrderLevel", "SellingPrice", "DiscountPrice"]
    actual_headers = [cell.value for cell in ws[1]]
    if actual_headers != expected_headers:
        raise Exception(
            f"Invalid file format. Expected headers: {expected_headers}, but got: {actual_headers}"
        )

    def is_valid_number(value):
        """Check if value is a non-negative number."""
        if value is None:
            return False, "missing"
        try:
            num = float(value)
            if isinstance(value, str) and value.strip().replace('.', '', 1).lstrip('-').isdigit() is False:
                return False, "not a number"
            if num < 0:
                return False, "negative"
            return True, num
        except (ValueError, TypeError):
            return False, "not a number"

    updated_count = 0
    skipped_rows = []

    with db_cursor() as cursor:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            no, name, part_no, order_level, selling_price, discount_price = row

            # Skip completely empty rows
            if all(v is None for v in row):
                continue

            # Validate required fields
            if not name or not part_no:
                skipped_rows.append(f"Row {row_idx}: Missing Name or PartNo — skipped.")
                continue

            # Validate OrderLevel
            valid, order_level_result = is_valid_number(order_level)
            if not valid:
                if order_level_result == "negative":
                    skipped_rows.append(f"Row {row_idx} ({name}): OrderLevel cannot be negative ({order_level}) — skipped.")
                elif order_level_result == "not a number":
                    skipped_rows.append(f"Row {row_idx} ({name}): OrderLevel must be a number, got '{order_level}' — skipped.")
                else:
                    skipped_rows.append(f"Row {row_idx} ({name}): OrderLevel is missing — skipped.")
                continue

            # Validate SellingPrice
            valid, selling_price_result = is_valid_number(selling_price)
            if not valid:
                if selling_price_result == "negative":
                    skipped_rows.append(f"Row {row_idx} ({name}): SellingPrice cannot be negative ({selling_price}) — skipped.")
                elif selling_price_result == "not a number":
                    skipped_rows.append(f"Row {row_idx} ({name}): SellingPrice must be a number, got '{selling_price}' — skipped.")
                else:
                    skipped_rows.append(f"Row {row_idx} ({name}): SellingPrice is missing — skipped.")
                continue

            # Validate DiscountPrice (optional but must be non-negative number if present)
            if discount_price is not None:
                valid, discount_price_result = is_valid_number(discount_price)
                if not valid:
                    if discount_price_result == "negative":
                        skipped_rows.append(f"Row {row_idx} ({name}): DiscountPrice cannot be negative ({discount_price}) — skipped.")
                    elif discount_price_result == "not a number":
                        skipped_rows.append(f"Row {row_idx} ({name}): DiscountPrice must be a number, got '{discount_price}' — skipped.")
                    continue
            else:
                discount_price_result = 0.0

            # Fetch the CarPartNames ID using Name and PartNo
            cursor.execute("""
                SELECT ID FROM tbl_carpartnames
                WHERE Name = %s AND PartNo = %s
                LIMIT 1
            """, (str(name).strip(), str(part_no).strip()))
            part_row = cursor.fetchone()

            if not part_row:
                skipped_rows.append(f"Row {row_idx} ({name} / {part_no}): No matching part found in database — skipped.")
                continue

            car_part_id = part_row[0]

            # Update OrderLevel in tbl_carpartnames
            cursor.execute("""
                UPDATE tbl_carpartnames
                SET OrderLevel = %s
                WHERE ID = %s
            """, (int(order_level_result), car_part_id))

            # Update Amount and SaleDiscount in tbl_partssellingprice
            cursor.execute("""
                UPDATE tbl_partssellingprice
                SET Amount = %s, SaleDiscount = %s, SetPriceDate = %s
                WHERE CarPartsNamesID = %s
            """, (
                float(selling_price_result),
                float(discount_price_result),
                datetime.date.today(),
                car_part_id
            ))

            updated_count += 1

    # Build result summary
    summary = f"{updated_count} part(s) updated successfully."
    if skipped_rows:
        summary += f"\n{len(skipped_rows)} row(s) skipped:\n" + "\n".join(skipped_rows)

    return summary

# *************************************************** Users Section ************************************


@anvil.server.callable()
def getUsers():
    _require_role("Administrator")
    with db_cursor() as cursor:
        rows = app_tables.users.search()
        result = []
        for i, row in enumerate(rows, start=1):

            def fmt_date(val):
                if val:
                    if (
                        hasattr(val, "tzinfo") and val.tzinfo
                    ):  # remove timezone if exists
                        val = val.replace(tzinfo=None)
                    return val.strftime("%Y-%m-%d %H:%M:%S")
                return None

            cursor.execute(
                "SELECT Roles FROM tbl_roles WHERE ID = %s", (row["role_id"],)
            )
            role = cursor.fetchone()

            result.append(
                {
                    "No": i,
                    "email": row["email"],
                    "last_login": fmt_date(row["last_login"]),
                    "signed_up": fmt_date(row["signed_up"]),
                    "role": role[0] if role else None,
                    "enabled": "Yes" if row["enabled"] else "No",
                    "loginattempts": row["n_password_failures"],
                }
            )
        return result


# *************************************************** Periodic Quotation And Invoice Totals Section ************************************


@anvil.server.callable()
def get_jobcard_quote_invoice_totals(jobcard_ref=None, start_date=None, end_date=None):
    """
    Returns jobcards with BOTH quotations and invoices,
    filtered by JobCardRef (LIKE) and ReceivedDate range.

    Parameters:
        jobcard_ref (str)   -> optional, partial match on JobCardRef
        start_date (date)   -> optional, filter from this date (inclusive)
        end_date (date)     -> optional, filter up to this date (inclusive)
    """
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT
                j.ID              AS ID,
                j.ReceivedDate    AS ReceivedDate,
                j.JobCardRef      AS JobCardRef,
                COALESCE(q.QuotationTotal, 0) AS QuotationTotal,
                COALESCE(i.InvoiceTotal, 0)   AS InvoiceTotal
            FROM tbl_jobcarddetails j
            INNER JOIN (
                SELECT AssignedJobID,
                       SUM(
                         CASE
                           WHEN QuantityIssued IS NULL THEN Amount
                           ELSE QuantityIssued * Amount
                         END
                       ) AS QuotationTotal
                FROM tbl_quotation
                GROUP BY AssignedJobID
            ) q ON q.AssignedJobID = j.ID
            INNER JOIN (
                SELECT AssignedJobID,
                       SUM(
                         CASE
                           WHEN QuantityIssued IS NULL THEN Amount
                           ELSE QuantityIssued * Amount
                         END
                       ) AS InvoiceTotal
                FROM tbl_invoices
                GROUP BY AssignedJobID
            ) i ON i.AssignedJobID = j.ID
            WHERE 1=1
        """

        params = []

        if jobcard_ref:
            query += " AND j.JobCardRef LIKE %s"
            params.append(f"%{jobcard_ref}%")  # <-- partial match

        if start_date and end_date:
            query += " AND j.ReceivedDate BETWEEN %s AND %s"
            params.extend([start_date, end_date])
        elif start_date:
            query += " AND j.ReceivedDate >= %s"
            params.append(start_date)
        elif end_date:
            query += " AND j.ReceivedDate <= %s"
            params.append(end_date)

        query += " ORDER BY j.ReceivedDate DESC"

        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cols = [c[0] for c in cursor.description]

    # Convert tuples to dicts, add No
    output = []
    for no, tup in enumerate(rows, start=1):
        r = dict(zip(cols, tup))
        output.append(
            {
                "No": no,
                "ID": r["ID"],
                "ReceivedDate": r["ReceivedDate"],
                "JobCardRef": r["JobCardRef"],
                "QuotationTotal": f"{float(r['QuotationTotal'] or 0):,.2f}",
                "InvoiceTotal": f"{float(r['InvoiceTotal'] or 0):,.2f}",
            }
        )

    return output


# *************************************************** Client Payment Summary Section ************************************
@anvil.server.callable()
def get_client_payment_summary(search_term=None, start_date=None, end_date=None):
    """
    Returns payment summary per JobCard with optional search term and date range.
    - No (enumerate)
    - Fullname
    - Phone
    - JobCardRef
    - ReceivedDate
    - TotalPaid (SUM of AmountPaid for the JobCardRefID)
    - Discount (latest Discount value for the JobCardRefID)
    - Balance (latest Balance value for the JobCardRefID)
    """

    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_clientcontacts.Fullname,
                tbl_clientcontacts.Phone,
                tbl_jobcarddetails.JobCardRef,
                tbl_jobcarddetails.ReceivedDate,
                SUM(tbl_payments.AmountPaid) AS TotalPaid,
                MAX(tbl_payments.Discount)   AS Discount,
                (
                    SELECT tbl_payments.Balance
                    FROM tbl_payments
                    WHERE tbl_payments.JobCardRefID = tbl_jobcarddetails.ID
                    ORDER BY tbl_payments.ID DESC
                    LIMIT 1
                ) AS LatestBalance
            FROM tbl_payments 
            INNER JOIN tbl_jobcarddetails  
                ON tbl_jobcarddetails.ID = tbl_payments.JobCardRefID
            INNER JOIN tbl_clientcontacts  
                ON tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails
            WHERE 1=1
        """

        params = []

        # Optional search filter
        if search_term:
            query += " AND (tbl_clientcontacts.Fullname LIKE %s OR tbl_clientcontacts.Phone LIKE %s OR tbl_jobcarddetails.RegNo LIKE %s)"
            params.extend([f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"])

        # Optional date filters
        if start_date and end_date:
            query += " AND tbl_jobcarddetails.ReceivedDate BETWEEN %s AND %s"
            params.extend([start_date, end_date])
        elif start_date:
            query += " AND tbl_jobcarddetails.ReceivedDate >= %s"
            params.append(start_date)
        elif end_date:
            query += " AND tbl_jobcarddetails.ReceivedDate <= %s"
            params.append(end_date)

        query += """
            GROUP BY tbl_clientcontacts.Fullname, tbl_clientcontacts.Phone, tbl_jobcarddetails.JobCardRef, tbl_jobcarddetails.ReceivedDate, tbl_jobcarddetails.ID
            ORDER BY tbl_jobcarddetails.ReceivedDate DESC
        """

        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cols = [col[0] for col in cursor.description]

    # Convert rows to dicts and add No
    results = []
    for no, tup in enumerate(rows, start=1):
        r = dict(zip(cols, tup))
        results.append(
            {
                "No": no,
                "JobReceivedDate": (
                    r["ReceivedDate"].strftime("%Y-%m-%d")
                    if r["ReceivedDate"]
                    else None
                ),
                "JobCardRef": r["JobCardRef"],
                "Fullname": r["Fullname"],
                "Phone": r["Phone"],
                "TotalPaid": f"{float(r['TotalPaid'] or 0):,.2f}",
                "Discount": f"{float(r['Discount'] or 0):,.2f}",
                "Balance": f"{float(r['LatestBalance'] or 0):,.2f}",
            }
        )

    return results


# *************************************************** Roles Section ************************************


@anvil.server.callable()
def listRoles():
    _get_current_user()
    with db_cursor() as cursor:
        query = "SELECT Roles, Description FROM tbl_roles ORDER BY Roles"
        cursor.execute(query)
        rows = cursor.fetchall()
        result = [{"Roles": row[0], "Description": row[1]} for row in rows]
        return result


@anvil.server.callable()
def duplicateRole(role, description):
    _get_current_user()
    with db_cursor() as cursor:
        # Check if a role with the same name already exists
        cursor.execute(
            "SELECT ID, Roles, Description FROM tbl_roles WHERE Roles = %s", (role,)
        )
        existing = cursor.fetchone()

        if existing:
            # Duplicate exists → don’t insert
            return {
                "status": "duplicate",
                "id": existing[0],
                "role": existing[1],
                "description": existing[2],
            }

        # If not found, insert new
        query = "INSERT INTO tbl_roles (Roles, Description) VALUES (%s, %s)"
        cursor.execute(query, (role, description))
        return {
            "status": "inserted",
            "id": cursor.lastrowid,
            "role": role,
            "description": description,
        }


@anvil.server.callable()
def getRoles():
    _get_current_user()
    with db_cursor() as cursor:
        query = "SELECT ID, Roles FROM tbl_roles ORDER BY Roles"
        cursor.execute(query)
        rows = cursor.fetchall()
        result = [(row[1], row[0]) for row in rows]
        return result


@anvil.server.callable()
def save_user_permissions(role_id, permissions_dict):
    _require_role("Administrator")
    with db_cursor() as cursor:
        # Delete old permissions first
        cursor.execute("DELETE FROM tbl_userpermissions WHERE RoleID = %s", (role_id,))

        # Insert new permissions
        for section, data in permissions_dict.items():
            # Save main section permission
            cursor.execute(
                """
                INSERT INTO tbl_userpermissions (RoleID, Section, SubSection, Allowed)
                VALUES (%s, %s, %s, %s)
                """,
                (role_id, section, None, int(data["main"])),
            )

            # Save each sub permission
            for sub, allowed in data["subs"].items():
                cursor.execute(
                    """
                    INSERT INTO tbl_userpermissions (RoleID, Section, SubSection, Allowed)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (role_id, section, sub, int(allowed)),
                )


# *************************************************** Get User Permissions Section ************************************
@anvil.server.callable()
def get_user_permissions(role_id):
    with db_cursor() as cursor:
        query = """
            SELECT Section, SubSection, Allowed
            FROM tbl_userpermissions
            WHERE RoleID = %s
        """
        cursor.execute(query, (role_id,))
        rows = cursor.fetchall()

        permissions = {}

        for section, subsection, allowed in rows:
            if section not in permissions:
                permissions[section] = {"main": False, "subs": {}}

            if subsection is None or subsection == "":
                # This is a main section permission
                permissions[section]["main"] = bool(allowed)
            else:
                # This is a sub-section permission
                permissions[section]["subs"][subsection] = bool(allowed)

        return permissions


# *************************************************** Edit User Account Section ************************************
@anvil.server.callable()
def get_account_roles():
    _get_current_user()
    with db_cursor() as cursor:
        query = "SELECT ID, Roles FROM tbl_roles ORDER BY Roles ASC"
        cursor.execute(query)
        rows = cursor.fetchall()

    # Return as list of tuples (label, value) for dropdown
    return [(row[1], row[0]) for row in rows]


@anvil.server.callable()
def get_role_id(role_name: str):
    _get_current_user()
    with db_cursor() as cursor:
        query = "SELECT ID FROM tbl_roles WHERE Roles = %s"
        cursor.execute(query, (role_name,))
        row = cursor.fetchone()

    return row[0] if row else None


@anvil.server.callable()
def getFailedAttempts(email):
    _get_current_user()
    user_row = app_tables.users.get(email=email)

    if user_row:
        # Safely return the value from the 'n_password_failures' column
        return user_row["n_password_failures"]
    else:
        # No user was found with that email
        return None


@anvil.server.callable()
def update_user(email, new_email, enabled, role_id):
    _require_role("Administrator")

    user = app_tables.users.get(email=email)
    if not user:
        raise Exception("Invalid request.")

    updates = {}
    if new_email is not None:
        updates["email"] = new_email
    if enabled is not None:
        updates["enabled"] = enabled
    if role_id is not None:
        updates["role_id"] = role_id

    # Reset login attempt counter on account edit
    updates["n_password_failures"] = None

    if updates:
        user.update(**updates)

    return "User updated successfully."


# *************************************************** Reset User Password Section ************************************
def hash_password(password, salt):
    """Hash a password with bcrypt. salt must be bytes from bcrypt.gensalt()."""
    if not isinstance(password, bytes):
        password = password.encode()
    # salt from bcrypt.gensalt() is already bytes; do NOT re-encode it
    if not isinstance(salt, bytes):
        raise TypeError("salt must be bytes produced by bcrypt.gensalt()")

    result = bcrypt.hashpw(password, salt)

    if isinstance(result, bytes):
        return result.decode("utf-8")


def _validate_password_strength(password: str):
    """Raise Exception if the password does not meet minimum policy requirements."""
    if len(password) < 8:
        raise Exception("Password must be at least 8 characters long.")
    if not re.search(r"[A-Z]", password):
        raise Exception("Password must contain at least one uppercase letter.")
    if not re.search(r"[0-9]", password):
        raise Exception("Password must contain at least one digit.")


@anvil.server.callable()
def reset_password(email, new_password):
    """Reset a user's password. Requires Admin role."""
    _require_role("Administrator")

    # Validate strength before touching the DB
    _validate_password_strength(new_password)

    user = app_tables.users.get(email=email)
    # Use a generic message to avoid user-enumeration
    if not user:
        raise Exception("Invalid request.")

    user["password_hash"] = hash_password(new_password, bcrypt.gensalt())
    return "Password reset successfully."


# *************************************************** Get Roles And Permissions In HTML Format Section ************************************


@anvil.server.callable()
def get_roles_permissions_html():
    _get_current_user()
    html = """
    <style>
      body {
        font-family: "Mozilla Headline", Arial, Helvetica, sans-serif;
        font-size: 16px;
      }
      table {
        border-collapse: collapse;
        width: 100%;
        margin-top: 10px;
      }
      td {
        padding: 6px 12px;
        vertical-align: top;
        border: 1px solid #ccc;
      }
      tr:nth-child(even) {
        background-color: #f9f9f9;
      }
      .role {
        font-weight: 700;
        font-size: 18px;
        color: #2c3e50;
        background-color: #e6f0ff;
      }
      .section {
        font-weight: 600;
        color: #34495e;
        padding-left: 20px;
        background-color: #f2f6fc;
      }
      .subsection {
        color: #555;
        padding-left: 40px;
      }
      .allowed {
        padding-left: 60px;
        font-style: italic;
      }
    </style>
    <table>
    """

    sections_without_subs = {"JOB CARD", "TRACKER", "PAYMENT", "RESET", "TECHNICIAN PORTAL"}

    with db_cursor() as cursor:
        query = """
            SELECT 
                r.Roles,
                u.Section,
                u.SubSection,
                u.Allowed,
                CASE u.Section
                    WHEN 'CONTACT'   THEN 1
                    WHEN 'JOB CARD'  THEN 2
                    WHEN 'WORKFLOW'  THEN 3
                    WHEN 'TRACKER'   THEN 4
                    WHEN 'REVISION'  THEN 5
                    WHEN 'PAYMENT'   THEN 6
                    WHEN 'INVENTORY' THEN 7
                    WHEN 'REPORTS'   THEN 8
                    WHEN 'SETTINGS'  THEN 9
                    WHEN 'RESET'  THEN 10
                    ELSE 999
                END AS section_order
            FROM tbl_roles r
            JOIN tbl_userpermissions u
              ON r.ID = u.RoleID
            ORDER BY
                r.Roles,
                section_order,
                u.SubSection
        """
        cursor.execute(query)
        rows = cursor.fetchall()

    # Build HTML hierarchy
    current_role = None
    current_section = None

    for role, section, subsection, allowed, _ in rows:
        if role != current_role:
            html += f'<tr><td class="role" colspan="2">Role: {role}</td></tr>'
            current_role = role
            current_section = None

        if section != current_section:
            html += f'<tr><td class="section" colspan="2">{section}</td></tr>'
            current_section = section

        # Handle sections without subsections
        if section in sections_without_subs:
            if int(allowed) == 1:
                html += '<tr><td class="allowed">Allowed</td><td><span style="color:green;">&#10004;</span></td></tr>'
            else:
                html += '<tr><td class="allowed">Allowed</td><td><span style="color:red;">&#10008;</span></td></tr>'
        else:
            # Normal subsection logic
            if subsection and subsection != "None":
                html += f'<tr><td class="subsection">{subsection}</td>'
                if int(allowed) == 1:
                    html += '<td><span style="color:green;">&#10004;</span></td></tr>'
                else:
                    html += '<td><span style="color:red;">&#10008;</span></td></tr>'

    html += "</table>"
    return html


# *************************************************** Resolve Part Number from Barcode Section ************************************
@anvil.server.callable()
def resolve_part(barcode_or_partno):
    _get_current_user()
    with db_cursor() as cursor:
        # 1. Check in tbl_carpartnames.PartNo
        cursor.execute(
            "SELECT PartNo, Name FROM tbl_carpartnames WHERE PartNo = %s",
            (barcode_or_partno,),
        )
        row = cursor.fetchone()
        if row:
            return {"PartNo": row[0], "Name": row[1]}

        # 2. Check in tbl_barcodepartnomapping.Barcode
        cursor.execute(
            "SELECT PartNo FROM tbl_barcodepartnomapping WHERE Barcode = %s",
            (barcode_or_partno,),
        )
        row = cursor.fetchone()
        if row:
            partno = row[0]
            # lookup in carpartnames
            cursor.execute(
                "SELECT PartNo, Name FROM tbl_carpartnames WHERE PartNo = %s", (partno,)
            )
            carpart = cursor.fetchone()
            if carpart:
                return {"PartNo": carpart[0], "Name": carpart[1]}

        # Not found
        return None


@anvil.server.callable()
def saveBarcodePartNo(barcode, partno):
    _get_current_user()
    with db_cursor() as cursor:
        # Insert new mapping
        cursor.execute(
            "INSERT INTO tbl_barcodepartnomapping (Barcode, PartNo) VALUES (%s, %s)",
            (barcode, partno),
        )
        return f"Barcode '{barcode}' is mapped to PartNo '{partno}'."


# *************************************************** Stocktake Harmonization Section ************************************


@anvil.server.callable()
def save_stocktake(repeating_panel_data):
    """
    repeating_panel_data is expected to be a list of dicts:
    [
        {"PartNo": "ABC123", "Name": "Brake Pad", "Quantity": 20},
        {"PartNo": "XYZ456", "Name": "Oil Filter", "Quantity": 15},
        ...
    ]
    """
    _get_current_user()
    with db_cursor() as cursor:
        for item in repeating_panel_data:
            part_no = item["PartNo"]
            quantity = int(item["Quantity"])

            # 1. Get CarPartName info
            cursor.execute(
                """
                SELECT cp.ID, cp.Name, cp.PartNo, cp.OrderLevel, cl.Location
                FROM tbl_carpartnames cp
                LEFT JOIN tbl_carpartslocation cl ON cp.Location = cl.ID
                WHERE cp.PartNo = %s
            """,
                (part_no,),
            )
            result = cursor.fetchone()
            if not result:
                continue
            carpart_id, carpart_name, part_number, reorder_level, location = result

            # 2. Get TotalStock and TotalIssued
            cursor.execute(
                """
                SELECT 
                    COALESCE(SUM(sp.NoOfUnits), 0) AS TotalStock,
                    COALESCE(i.TotalIssued, 0) AS TotalIssued
                FROM tbl_carpartnames cp
                LEFT JOIN tbl_stockparts sp ON cp.ID = sp.CarPart
                LEFT JOIN (
                    SELECT CarPartID, SUM(QuantityIssued) AS TotalIssued
                    FROM tbl_assignedcarparts
                    GROUP BY CarPartID
                ) i ON cp.ID = i.CarPartID
                WHERE cp.ID = %s
                GROUP BY cp.ID
            """,
                (carpart_id,),
            )
            stock_info = cursor.fetchone()
            total_stock, total_issued = stock_info if stock_info else (0, 0)

            # 3. Balance and Variance
            balance = total_stock - total_issued
            variance = quantity - balance

            # 4. Comment logic
            if quantity < balance:
                comment = "System Count > Actual Count"
            elif quantity > balance:
                comment = "Actual Stock > System Stock"
            else:
                comment = "Stock Balanced"

            # 5. Insert/Update tbl_stocktakeharmonized (harmonized values)
            cursor.execute(
                """
                INSERT INTO tbl_stocktakeharmonized
                    (StockTakeDate, CarPartNameID, HarmornizedValue)
                VALUES (CURDATE(), %s, %s)
                ON DUPLICATE KEY UPDATE
                    HarmornizedValue = VALUES(HarmornizedValue)
            """,
                (carpart_id, variance),
            )

            # 6. Insert/Update tbl_finalcapturedstocktaking (daily unique)
            cursor.execute(
                """
                INSERT INTO tbl_finalcapturedstocktaking
                    (CountingDate, CarPart, PartNumber, ReorderLevel, Location, Balance, CapturedQuantity, Variance, Comment)
                VALUES (CURDATE(), %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    CarPart = VALUES(CarPart),
                    ReorderLevel = VALUES(ReorderLevel),
                    Location = VALUES(Location),
                    Balance = VALUES(Balance),
                    CapturedQuantity = VALUES(CapturedQuantity),
                    Variance = VALUES(Variance),
                    Comment = VALUES(Comment)
            """,
                (
                    carpart_name,
                    part_number,
                    reorder_level,
                    location,
                    balance,
                    quantity,
                    variance,
                    comment,
                ),
            )

    return "Stocktake saved/updated successfully"


# *************************************************** Stock Analysis Report Section ************************************


@anvil.server.callable()
def get_stock_analysis_report(start_date=None, end_date=None, partnumber=None):

    _get_current_user()
    now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3)))
    header_date = now.strftime("%A, %d %B %Y")
    header_time = now.strftime("%H:%M:%S")

    # --- Styles & Header ---
    html_parts = [
        f""" 
        <style> 
            body {{ 
                  font-family: "Mozilla Headline", Arial, Helvetica, sans-serif; font-size: 15px; margin: 0; padding: 10px; 
                  }} 
            .header {{ 
                  display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; flex-wrap: wrap; 
                  }} 
            .report-title {{
                   font-size: 20px; font-weight: bold; 
                  }}
             .datetime {{ 
                  text-align: right; font-size: 14px; color: #333; 
                  }}
            .table-note {{ 
                  border: 2px solid red; 
                  color: red; 
                  font-weight: bold; 
                  padding: 8px; 
                  margin: 10px 0; 
                  text-align: center; 
                  font-size: 16px;
                  border-radius: 6px; 
                  background-color: #ffe6e6; /* light red background for emphasis */
                   }} 
            .table-container {{
                   overflow-x: auto; -webkit-overflow-scrolling: touch; max-height: 70vh; 
                  }}
            table {{ 
                  border-collapse: collapse; width: 100%; min-width: 900px; margin-top: 10px; 
                  }}
            th, td {{
                   padding: 6px 10px; border: 1px solid #999; text-align: left; white-space: normal;
                   }} 
            th {{
                   background-color: #0074D9; color: #fff; position: sticky; top: 0; z-index: 2; 
                  }} 
            .comment-red {{ 
                  background-color: #ff4d4d; color: #fff; font-weight: bold; 
                  }}
            .comment-yellow {{ 
                  background-color: #ffd633; font-weight: bold; 
                  }}
            .comment-green {{ 
                  background-color: #00A300; color: #fff; font-weight: bold; 
                  }}
            </style> 
                  <div class="header"> 
                    <div class="report-title">Stock Analysis Report</div> 
                    <div class="datetime"> {header_date}<br>{header_time} </div> 
                  </div> 
                 <div class="table-note"> 
                    <div>System Stock = Total Stock Bought - Total Stock Issued</div>
                    <div>Variance = Actual Stock - System Stock</div>
                 </div> 
                 <div class="table-container"> 
            <table> 
             <tr> 
               <th>Date</th> 
               <th>Location</th> 
               <th>Name</th> 
               <th>PartNo</th> 
               <th>Reorder Level</th> 
               <th>System Stock</th> 
               <th>Actual Stock</th> 
               <th>Variance</th> 
               <th>Comment</th> 
               </tr> 
        """
    ]

    # --- Query ---
    query = """
        SELECT 
          CountingDate,
          Location,
          CarPart,
          PartNumber,
          ReorderLevel,
          Balance AS SystemStock,
          CapturedQuantity AS ActualStock,
          Variance,
          Comment
        FROM tbl_finalcapturedstocktaking
        WHERE 1=1
    """
    params = []

    if start_date and end_date:
        query += " AND CountingDate BETWEEN %s AND %s"
        params.extend([start_date, end_date])
    if partnumber:
        query += " AND PartNumber = %s"
        params.append(partnumber)

    query += " ORDER BY CountingDate DESC, Location, CarPart"

    with db_cursor() as cursor:
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()

    # --- Build Rows ---
    current_date = None
    current_location = None
    row_parts = []

    for (
        counting_date,
        location,
        carpart,
        partnumber,
        reorder,
        system_stock,
        actual_stock,
        variance,
        comment,
    ) in rows:

        date_cell = ""
        location_cell = ""

        if counting_date != current_date:
            date_cell = counting_date.strftime("%d/%m/%Y")
            current_date = counting_date
            current_location = None

        if location != current_location:
            location_cell = location
            current_location = location

        if comment and "System Count > Actual Count" in comment:
            comment_class = "comment-red"
        elif comment and "Actual Stock > System Stock" in comment:
            comment_class = "comment-yellow"
        elif comment and "Stock Balanced" in comment:
            comment_class = "comment-green"
        else:
            comment_class = ""

        row_parts.append(
            f"""
        <tr>
          <td>{date_cell}</td>
          <td>{location_cell}</td>
          <td>{carpart}</td>
          <td>{partnumber}</td>
          <td>{reorder}</td>
          <td>{system_stock}</td>
          <td>{actual_stock}</td>
          <td>{variance}</td>
          <td class="{comment_class}">{comment or ''}</td>
        </tr>
        """
        )

    html_parts.append("".join(row_parts))
    html_parts.append("</table></div>")
    return "".join(html_parts)


# *************************************************** Download Job Card Report Section ************************************


@anvil.server.callable()
def get_jobcard_details(jobCardID):
    """Fetch full job card details for a given JobCardID and return as list of dicts."""
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT
                tbl_jobcarddetails.ID,
                tbl_clientcontacts.Fullname AS ClientName,
                tbl_clientcontacts.Phone,
                tbl_clientcontacts.Address,
                tbl_clientcontacts.Email,
                tbl_technicians.Fullname AS TechnicianName,
                tbl_jobcarddetails.JobCardRef,
                tbl_jobcarddetails.ReceivedDate,
                tbl_jobcarddetails.DueDate,
                tbl_jobcarddetails.ExpDate,
                tbl_checkstaff.Staff,
                tbl_jobcarddetails.Ins,
                tbl_jobcarddetails.Comp,
                tbl_jobcarddetails.TPO,
                tbl_jobcarddetails.Spare,
                tbl_jobcarddetails.Jack,
                tbl_jobcarddetails.Brace,
                tbl_jobcarddetails.RegNo,
                tbl_jobcarddetails.MakeAndModel,
                tbl_jobcarddetails.ChassisNo,
                tbl_jobcarddetails.EngineCC,
                tbl_jobcarddetails.Mileage,
                tbl_jobcarddetails.EngineNo,
                tbl_jobcarddetails.EngineCode,
                tbl_jobcarddetails.Manual,
                tbl_jobcarddetails.Auto,
                tbl_jobcarddetails.Empty,
                tbl_jobcarddetails.Quarter,
                tbl_jobcarddetails.Half,
                tbl_jobcarddetails.ThreeQuarter,
                tbl_jobcarddetails.Full,
                tbl_jobcarddetails.PaintCode,
                tbl_jobcarddetails.ClientInstruction,
                tbl_jobcarddetails.Notes
            FROM
                tbl_technicians
            INNER JOIN(
                    tbl_pendingassignedjobs
                INNER JOIN(
                        tbl_checkstaff
                    INNER JOIN(
                            tbl_clientcontacts
                        INNER JOIN tbl_jobcarddetails ON tbl_clientcontacts.ID = tbl_jobcarddetails.ClientDetails
                        )
                    ON
                        tbl_checkstaff.ID = tbl_jobcarddetails.CheckedInBy
                    )
                ON
                    tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
                )
            ON
                tbl_technicians.ID = tbl_pendingassignedjobs.TechnicianID
            WHERE tbl_jobcarddetails.ID = %s
        """
        cursor.execute(query, (jobCardID,))
        row = cursor.fetchone()

    if not row:
        return []

    # Map tuple → dict
    result = {
        "JobCardId": row[0],
        "ClientName": str(row[1]) if row[1] else "",
        "ClientPhone": str(row[2]) if row[2] else "",
        "ClientAddress": str(row[3]) if row[3] else "",
        "ClientEmail": str(row[4]) if row[4] else "",
        "TechnicianName": str(row[5]) if row[5] else "",
        "JobCardRef": str(row[6]) if row[6] else "",
        "ReceivedDate": str(row[7]) if row[7] else "",
        "DueDate": str(row[8]) if row[8] else "",
        "ExpDate": str(row[9]) if row[9] else "",
        "CheckedInBy": str(row[10]) if row[10] else "",
        "Ins": str(row[11]) if row[11] else "",
        "Comp": str(row[12]) if row[12] else "",
        "TPO": str(row[13]) if row[13] else "",
        "Spare": str(row[14]) if row[14] else "",
        "Jack": str(row[15]) if row[15] else "",
        "Brace": str(row[16]) if row[16] else "",
        "RegNo": str(row[17]) if row[17] else "",
        "MakeAndModel": str(row[18]) if row[18] else "",
        "ChassisNo": str(row[19]) if row[19] else "",
        "EngineCC": str(row[20]) if row[20] else "",
        "Mileage": str(row[21]) if row[21] else "",
        "EngineNo": str(row[22]) if row[22] else "",
        "EngineCode": str(row[23]) if row[23] else "",
        "Manual": str(row[24]) if row[24] else "",
        "Auto": str(row[25]) if row[25] else "",
        "Empty": str(row[26]) if row[26] else "",
        "Quarter": str(row[27]) if row[27] else "",
        "Half": str(row[28]) if row[28] else "",
        "ThreeQuarter": str(row[29]) if row[29] else "",
        "Full": str(row[30]) if row[30] else "",
        "PaintCode": str(row[31]) if row[31] else "",
        "ClientInstruction": str(row[32]) if row[32] else "",
        "Notes": str(row[33]) if row[33] else "",
    }

    return result


@anvil.server.callable()
def get_signature(jobCardID: int) -> str | None:
    """
    Fetch the Base64-encoded client signature for a given job card.
    Returns None if no signature is found.
    """
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Signature
            FROM tbl_signedjobcards
            WHERE AssignedJobID = %s
            ORDER BY ID DESC
            LIMIT 1
        """
        cursor.execute(query, (jobCardID,))
        row = cursor.fetchone()

    if row:
        # If Signature is stored as BLOB → convert to base64
        if isinstance(row[0], (bytes, bytearray)):
            return base64.b64encode(row[0]).decode("utf-8")
        # If already base64/text stored in DB → return as is
        return row[0]

    return None


@anvil.server.callable()
def fillJobCardReport(jobCardID, logo_path: str = os.getenv("LOGO"),font_path: str = os.getenv("FONT_PATH")) -> str:
    _get_current_user()
    data = get_jobcard_details(jobCardID)

    signature_b64 = anvil.server.call("get_signature", jobCardID)
    if signature_b64:
        signature_html = f"""
            <img src="data:image/png;base64,{signature_b64}" 
                 alt="Client Signature" 
                 style="max-width: 90%; height: auto; display: block; border: 1px solid #ccc; padding: 4px;"/>
        """
    else:
        signature_html = "<em>No signature available</em>"

    docTitle = "Jobcard Details"

    # === Embed MozillaHeadline font as base64 ===
    font_base64 = ""
    if font_path and os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # === Handle company logo ===
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
        logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" alt="Company Logo" style="width: 100%; height: 100%; border-radius: 2px;">'
    else:
        logo_img_tag = "LOGO"

    insurance_html = ""
    if data["Comp"]:
        insurance_html += "Comprehensive"
    if data["TPO"]:
        insurance_html += "Third Party Only"

    items_list = []
    if data["Spare"]:
        items_list.append("Spare")
    if data["Jack"]:
        items_list.append("Jack")
    if data["Brace"]:
        items_list.append("Brace")

    transmission_html = []
    if data["Manual"]:
        transmission_html.append("Manual")
    if data["Auto"]:
        transmission_html.append("Auto")

    tank_html = []
    if data["Empty"]:
        tank_html.append("Empty")
    if data["Quarter"]:
        tank_html.append("1/4")
    if data["Half"]:
        tank_html.append("1/2")
    if data["ThreeQuarter"]:
        tank_html.append("3/4")
    if data["Full"]:
        tank_html.append("Full")

    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="UTF-8">
      <title>{docTitle}</title>
      <style>
        /* --- Embed Mozilla Headline font --- */
        @font-face {{
            font-family: 'Mozilla Headline';
            src: url(data:font/ttf;base64,{font_base64}) format('truetype');
        }}
        @page {{
          size: A4;
          margin: 20mm;
        }}
        body {{
          font-family: Roboto, Noto, Arial, sans-serif;
          font-size: 12pt;
          line-height: 1.5;
          background: #fff;
          margin: 0;
          color: #000;
        }}
        .container {{
          width: 100%;
          margin: auto;
        }}
        .logo-section {{
            text-align: center;
            padding: 24px;
            background-color: white;
            border-bottom: 1px solid #e0e0e0;
        }}

        .logo-container {{
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 20px;
            margin-bottom: 16px;
        }}

        .logo-image {{
            width: 725px;
            height: 100px;
            background: linear-gradient(135deg, #228B22, #90EE90, #FFD700, #FF6347);
            border-radius: 2px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
            color: white;
            font-weight: 500;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}
        .doc-title {{
          background-color: #000;
            color: white;
            text-align: center;
            padding: 16px 24px;
            font-size: 16px;
            font-weight: 300;
            letter-spacing: .5px;
            box-shadow: 0 4px 5px 0 rgba(0, 0, 0, 0.14),
                        0 1px 10px 0 rgba(0, 0, 0, 0.12),
                        0 2px 4px -1px rgba(0, 0, 0, 0.2);
          text-transform: uppercase;
          font-family: 'Mozilla Headline';
        }}
        .section-title {{
          font-size: 12pt;
          font-weight: bold;
          border-bottom: 1px solid #000;
          margin: 16px 0 8px;
          padding-bottom: 3px;
        }}
        .detail-row {{
            display: grid;
            grid-template-columns: 120px 1fr; /* Slightly smaller label column for 3-col layout */
            column-gap: 8px;
            margin-bottom: 12px;
            padding: 6px 10px;
            border: 1px solid #eee; /* subtle separation */
            border-radius: 4px;
            background: #fafafa;
        }}
        .detail-label {{
            font-weight: bold;
            font-size: 12px; /* Slightly smaller for 3-col layout */
            color: rgba(0,0,0,0.87);
            font-family: 'Mozilla Headline';      
        }}
        .detail-value {{
            font-size: 12px; /* Slightly smaller for 3-col layout */
            color: rgba(0,0,0,0.87);
            text-align: left;   
        }}        
        .signature {{
          margin-top: 20px;
          padding: 10px;
          text-align: left;
          page-break-inside: avoid;
          font-family: 'Mozilla Headline';
        }}
        footer {{
          text-align: center;
          font-size: 10pt;
          color: #444;
          margin-top: 5px;
          padding-top: 2px;
          border-top: 1px solid #000;
          page-break-inside: avoid;
          font-family: 'Mozilla Headline';
        }}
      </style>
    </head>
    <body>
      <div class="container">
        <div class="logo-section">
            <div class="logo-container">
                <div class="logo-image">
                    {logo_img_tag}
                </div>
            </div>
        </div>
        <div class="doc-title">{docTitle}</div>


        <!-- UPDATED: Changed to 3-column table layout -->
        <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
            <tr>
                <!-- First Column - Client Details -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Name:</span>
                        <div>
                        <span class="detail-value">{data['ClientName']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Phone:</span>
                        <div>
                        <span class="detail-value">{data['ClientPhone']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Address:</span>
                        <div>
                        <span class="detail-value">{data['ClientAddress']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Email:</span>
                        <div>
                        <span class="detail-value">{data['ClientEmail']}</span>
                        </div>
                    </div>
                </td>

                <!-- Second Column - Vehicle Basic Details -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Reg No:</span>
                        <div>
                        <span class="detail-value">{data['RegNo']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Make & Model:</span>
                        <div>
                        <span class="detail-value">{data['MakeAndModel']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Chassis No:</span>
                        <div>
                        <span class="detail-value">{data['ChassisNo']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Engine CC:</span>
                        <div>
                        <span class="detail-value">{data['EngineCC'] if data['EngineCC'] else "Not Captured"}</span>
                        </div>
                    </div>
                </td>

                <!-- Third Column - Vehicle Technical Details -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Mileage:</span>
                        <div>
                        <span class="detail-value">{data['Mileage']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Engine No:</span>
                        <div>
                        <span class="detail-value">{data['EngineNo'] if data['EngineNo'] else "Not Captured"}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Engine Code:</span>
                        <div>
                        <span class="detail-value">{data['EngineCode'] if data['EngineCode'] else "Not Captured"}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Paint Code:</span>
                        <div>
                        <span class="detail-value">{data['PaintCode'] if data['PaintCode'] else "Not Captured"}</span>
                        </div>
                    </div>
                </td>
            </tr>
        </table>
        <!-- END UPDATED -->

        
        <!-- UPDATED: Changed to 3-column table layout -->
        <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
            <tr>
                <div class="section-title" style="margin-top: 5px;"></div>
                <!-- First Column  -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Ref:</span>
                        <div>
                        <span class="detail-value">{data['JobCardRef']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Received:</span>
                        <div>
                        <span class="detail-value">{data['ReceivedDate']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Due:</span>
                        <div>
                        <span class="detail-value">{data['DueDate']}</span>
                        </div>
                    </div>
                </td>

                <!-- Second Column  -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Insurance Expiry:</span>
                        <div>
                        <span class="detail-value">{data['ExpDate'] if data['ExpDate'] else "Not Captured"}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Checked In By:</span>
                        <div>
                        <span class="detail-value">{data['CheckedInBy']}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Insurance:</span>
                        <div>
                        <span class="detail-value">{insurance_html}</span>
                        </div>
                    </div>
                </td>

                <!-- Third Column -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Available Items:</span>
                        <div>
                        <span class="detail-value">{", ".join(items_list) if items_list else "No Spare/Jack/Brace"}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Transmission:</span>
                        <div>
                        <span class="detail-value">{",".join(transmission_html) if transmission_html else "Not specified"}</span>
                        </div>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Tank Level:</span>
                        <div>
                        <span class="detail-value">{",".join(tank_html) if tank_html else "Not specified"}</span>
                        </div>
                    </div>
                </td>
            </tr>
        </table>
        <!-- END UPDATED -->

        <!-- UPDATED: Changed to 3-column table layout -->
        <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
            <tr>
                <div class="section-title" style="margin-top: 2px;"></div>
                <!-- First Column -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Instructions:</span>
                        <div>
                        <span class="detail-value">{data['ClientInstruction']}</span>
                        </div>
                    </div>
                </td>

                <!-- Second Column -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Tech Notes:</span>
                        <div>
                        <span class="detail-value">{data['Notes'] if data['Notes'] else "None"}</span>
                        </div>
                    </div>
                </td>

                <!-- Third Column -->
                <td style="width: 33.33%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Technician:</span>
                        <div>
                        <span class="detail-value">{data['TechnicianName']}</span>
                        </div>
                    </div>
                </td>
            </tr>
        </table>
        <!-- END UPDATED -->

        <!-- UPDATED: Changed to 1-column table layout -->
        <table style="width: 100%; table-layout: fixed; margin: 24px 0;">
            <tr>
                <div class="section-title" style="margin-top: 5px;"></div>
                <!-- First Column -->
                <td style="width: 100%; vertical-align: top; padding: 0 16px;">
                    <div class="detail-row">
                        <span class="detail-label">Signature:</span>
                        <div>
                        <span class="detail-value">{signature_html}</span>
                        </div>
                    </div>
                </td>
            </tr>
        </table>
        <!-- END UPDATED -->

        <footer>
          Joy Is The Feeling Of Being Looked After By The Best - BMW CENTER For Your BMW.
        </footer>
      </div>
    </body>
    </html>
    """
    return html_content


@anvil.server.callable()
def createSignedJobcardPdf(jobCardID):
    _get_current_user()
    try:
        docName = anvil.server.call("getJobCardRef", jobCardID)
        fileName = str(docName[0]["JobCardRef"]) + " Jobcard"

        setting_options = {
            "encoding": "UTF-8",
            "custom-header": [("Accept-Encoding", "gzip")],
            "page-size": "A4",
            "orientation": "Portrait",
            "margin-top": "0.75in",
            "margin-right": "0.75in",
            "margin-bottom": "0.75in",
            "margin-left": "0.75in",
            "no-outline": False,
            "enable-local-file-access": None,
        }

        html_string = fillJobCardReport(jobCardID)
        pdfkit.from_string(
            html_string, fileName, options=setting_options, configuration=config
        )
        media_object = anvil.media.from_file(fileName, "application/pdf", name=fileName)
        return media_object

    except Exception as e:
        print("PDF generation failed:", str(e))
        raise


# *************************************************** Import Order Tracking Section ************************************
@anvil.server.callable()
def saveImportOrderTracking(
    dateValue, clientID, partname, partnumber, quantity, status
):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_importordertracking
                (OrderDate, ClientID, PartName, PartNumber, Quantity, Status)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query, (dateValue, clientID, partname, partnumber, quantity, status)
        )
    return "Import order tracking saved successfully"


@anvil.server.callable()
def search_client_import_orders(search_text: str):
    """
    Search client contacts by name or phone.
    Return unique (Fullname - OrderDate) options for dropdown,
    storing {"client_id": ..., "order_date": ...} as the value.
    """
    _get_current_user()
    with db_cursor() as cursor:
        sql = """
            SELECT DISTINCT
                   tbl_clientcontacts.ID AS ClientID,
                   tbl_clientcontacts.Fullname,
                   tbl_importordertracking.OrderDate
            FROM tbl_clientcontacts
            INNER JOIN tbl_importordertracking
              ON tbl_clientcontacts.ID = tbl_importordertracking.ClientID
            WHERE tbl_clientcontacts.Fullname LIKE %s
               OR tbl_clientcontacts.Phone LIKE %s
            ORDER BY tbl_importordertracking.OrderDate DESC
        """
        like_pattern = f"%{search_text}%"
        cursor.execute(sql, (like_pattern, like_pattern))
        rows = cursor.fetchall()

        results = [
            (f"{row[1]} - {row[2]}", {"client_id": row[0], "order_date": row[2]})
            for row in rows
        ]
        return results


@anvil.server.callable()
def get_import_orders_for_selection(client_id: int, order_date: str):
    """
    Return all rows from tbl_importordertracking for a given client_id and order_date.
    """
    _get_current_user()
    with db_cursor() as cursor:
        sql = """
            SELECT 
                   tbl_importordertracking.PartName AS Name,
                   tbl_importordertracking.PartNumber AS Part_No,
                   tbl_importordertracking.Quantity,
                   tbl_importordertracking.Status,
                   tbl_importordertracking.OrderDate,
                   tbl_clientcontacts.Fullname,
                     tbl_importordertracking.Amount 
            FROM tbl_importordertracking
            INNER JOIN tbl_clientcontacts
              ON tbl_clientcontacts.ID = tbl_importordertracking.ClientID
            WHERE tbl_importordertracking.ClientID = %s
              AND tbl_importordertracking.OrderDate = %s
            ORDER BY tbl_importordertracking.ID
        """
        cursor.execute(sql, (client_id, order_date))
        rows = cursor.fetchall()

        # rows is a list of tuples, convert to list of dicts
        results = [
            {
                "No": idx + 1,
                "Name": r[0],
                "Part_No": r[1],
                "Quantity": r[2],
                "Status": r[3],
                "OrderDate": r[4],
                "Client": r[5],
                "Amount": f"{float(r[6]):,.2f}" if r[6] is not None else "0.00",
            }
            for idx, r in enumerate(rows)
        ]
        return results


@anvil.server.callable()
def updateImportOrderTracking(dateID, clientId, items):
    _get_current_user()
    with db_cursor() as cursor:
        # Delete existing records for the client and date
        cursor.execute(
            "DELETE FROM tbl_importordertracking WHERE ClientID = %s AND OrderDate = %s",
            (clientId, dateID),
        )

        # Append new records
        for item in items:
            query = """
                INSERT INTO tbl_importordertracking
                (OrderDate, ClientID, PartName, PartNumber, Quantity, Amount, Status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(
                query,
                (
                    dateID,
                    clientId,
                    item["name"],
                    item["number"],
                    item["quantity"],
                    item["amount"],
                    item["status"],
                ),
            )
    return "Import order tracking updated successfully"


# *************************************************** Monthly Performance Schedule Section ************************************


@anvil.server.callable()
def getMonthlyJobcardRef(startDate, endDate):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT tbl_jobcarddetails.JobCardRef,
                   tbl_invoices.Date, tbl_invoices.AssignedJobID
            FROM tbl_jobcarddetails
            INNER JOIN tbl_invoices
              ON tbl_jobcarddetails.ID = tbl_invoices.AssignedJobID
            WHERE tbl_invoices.Date BETWEEN %s AND %s
            GROUP BY tbl_jobcarddetails.JobCardRef, tbl_invoices.Date, tbl_invoices.AssignedJobID
            ORDER BY tbl_invoices.Date DESC
        """
        cursor.execute(query, (startDate, endDate))
        rows = cursor.fetchall()

        # Return list of tuples (label, value)
        return [(f"{r[0]} - {r[1].strftime('%Y-%m-%d')}", r[2]) for r in rows]


@anvil.server.callable()
def getPeriodicInvoices(startDate, endDate, jobcardref):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_invoices.Date, 
                tbl_invoices.Item, 
                tbl_invoices.Part_No, 
                tbl_invoices.QuantityIssued,
                tbl_invoices.Amount
            FROM tbl_invoices
            WHERE tbl_invoices.Date BETWEEN %s AND %s 
              AND tbl_invoices.AssignedJobID = %s
            ORDER BY tbl_invoices.Date DESC
        """
        cursor.execute(query, (startDate, endDate, jobcardref))
        rows = cursor.fetchall()

        # Add row number with enumerate
        result = [
            {
                "No": idx,  # row number
                "Date": r[0].strftime("%Y-%m-%d") if r[0] else None,
                "Item": r[1],
                "Part_No": r[2],
                "QuantityIssued": r[3],
                "Amount": f"{float(r[4]):,.2f}" if r[4] is not None else "0.00",
            }
            for idx, r in enumerate(rows, start=1)
        ]

        return result


@anvil.server.callable()
def getFullnameInvoiceAmountAndBalance(jobcard_id):
    # --- Get invoice details ---
    _get_current_user()
    invoice_rows = get_invoice_details_by_job_id(jobcard_id)
    if not invoice_rows:
        return {
            "Fullname": "",
            "TotalInvoiceAmount": 0.0,
            "TotalAmountPaid": 0.0,
            "TotalDiscount": 0.0,
            "MaxBalance": 0.0,
        }

    # Get the client fullname from the first row
    fullname = invoice_rows[0].get("Fullname", "")

    # Calculate the total invoice amount (sum of all item totals)
    total_invoice = sum(float(row.get("Total") or 0) for row in invoice_rows)

    # --- Get payment details ---
    payment_rows = getPaymentsDetails(jobcard_id)

    def safe_to_float(value):
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(",", ""))
            except ValueError:
                return 0.0
        return 0.0

    total_amount_paid = sum(
        safe_to_float(row.get("AmountPaid")) for row in payment_rows
    )

    total_discount = sum(safe_to_float(row.get("Discount")) for row in payment_rows)

    # Calculate balance
    balance = total_invoice - total_amount_paid - total_discount

    # Return formatted dictionary
    return {
        "Fullname": fullname,
        "TotalInvoiceAmount": round(total_invoice, 2),
        "TotalAmountPaid": round(total_amount_paid, 2),
        "TotalDiscount": round(total_discount, 2),
        "Balance": round(balance, 2),
    }


@anvil.server.callable()
def deleteMonthlySchedule(jobcardrefID):
    _get_current_user()
    with db_cursor() as cursor:
        cursor.execute(
            "DELETE FROM tbl_monthlyschedule WHERE AssignedJobID = %s", (jobcardrefID,)
        )
    return "Monthly schedule deleted successfully"


@anvil.server.callable()
def saveMonthlySchedule(
    invoiceDate,
    jobcardrefID,
    fullname,
    invoiceTotal,
    totalPaid,
    totalDiscount,
    balance,
    item,
    partNo,
    quantity,
    amount,
    category,
):
    _get_current_user()
    """
    Inserts a record into tbl_monthlyschedule table.
    Called for each row from the client repeating panel.
    """

    # Helper function to safely convert strings to float
    def safe_float(value):
        if value is None or value == "":
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        # Remove commas and whitespace before conversion
        return float(str(value).replace(",", "").strip())

    with db_cursor() as cursor:
        # Insert new record
        query = """
            INSERT INTO tbl_monthlyschedule
            (Date, AssignedJobID, ClientName, TotalInvoiceAmount,
             TotalAmountPaid, TotalDiscount, PaymentBalance, 
             Item, Part_No, QuantityIssued, Amount, Category)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        values = (
            invoiceDate,
            int(jobcardrefID),
            fullname,
            safe_float(invoiceTotal),
            safe_float(totalPaid),
            safe_float(totalDiscount),
            safe_float(balance),
            item,
            partNo,
            safe_float(quantity) if quantity is not None else None,
            safe_float(amount),
            category,
        )

        cursor.execute(query, values)
    return "Monthly schedule saved successfully"


# *************************************************** Monthly Schedule Pivot Section ************************************
@anvil.server.callable()
def get_monthly_schedule_pivot(start_date=None, end_date=None):
    """
    Returns a pivoted summary of monthly schedule data between start_date and end_date.
    Each category becomes its own column, with totals based on QuantityIssued * Amount if QuantityIssued is not NULL.
    Computes Balance = PaymentBalance + PrevBal safely after SQL aggregation.
    """
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                tbl_monthlyschedule.Date,
                tbl_monthlyschedule.AssignedJobID,
                tbl_monthlyschedule.ClientName,
                tbl_monthlyschedule.TotalInvoiceAmount,
                tbl_monthlyschedule.TotalAmountPaid,
                tbl_monthlyschedule.TotalDiscount,
                tbl_monthlyschedule.PaymentBalance,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Car Parts'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS CarParts,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Engine Oil'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS EngineOil,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Consumable'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS Consumable,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Diagnosis'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS Diagnosis,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Labour'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS Labour,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Engine Wash'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS EngineWash,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Alignment'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS Alignment,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Transfer / Gearbox Oil'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS TransferGearboxOil,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Others'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS Others,

                SUM(CASE WHEN tbl_monthlyschedule.Category = 'Prev Bal'
                         THEN IF(tbl_monthlyschedule.QuantityIssued IS NOT NULL,
                                 tbl_monthlyschedule.QuantityIssued * tbl_monthlyschedule.Amount,
                                 tbl_monthlyschedule.Amount)
                    ELSE 0 END) AS PrevBal

            FROM tbl_monthlyschedule
            WHERE (%s IS NULL OR tbl_monthlyschedule.Date >= %s)
              AND (%s IS NULL OR tbl_monthlyschedule.Date <= %s)
            GROUP BY 
                tbl_monthlyschedule.Date,
                tbl_monthlyschedule.AssignedJobID,
                tbl_monthlyschedule.ClientName,
                tbl_monthlyschedule.TotalInvoiceAmount,
                tbl_monthlyschedule.TotalAmountPaid,
                tbl_monthlyschedule.TotalDiscount,
                tbl_monthlyschedule.PaymentBalance
            ORDER BY tbl_monthlyschedule.Date DESC;
        """

        cursor.execute(query, (start_date, start_date, end_date, end_date))
        rows = cursor.fetchall()
        column_names = [desc[0] for desc in cursor.description]
        results = [dict(zip(column_names, row)) for row in rows]

        # Compute Balance = PaymentBalance + PrevBal safely in Python
        for r in results:
            payment_balance = r.get("PaymentBalance") or 0
            #prev_bal = r.get("PrevBal") or 0
            r["Balance"] = round(payment_balance, 2)

        return results

# ***************************************************Transition jobcards from 'Ready for Pickup' to 'Complete' Section ************************************
@anvil.server.callable()
def transitionreadyforpickuptocomplete():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_jobcarddetails
            SET Status = 'Complete'
            WHERE Status = 'Ready for Pickup'
        """
        cursor.execute(query)
    return "All 'Ready for Pickup' jobcards have been updated to 'Complete'."

# *************************************************** Get Notification Job Card Reference Section ************************************
@anvil.server.callable()
def getNotificationJobCardRef(jobCardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT JobCardRef
            FROM tbl_jobcarddetails
            WHERE ID = %s
        """

        cursor.execute(query, (jobCardID,))
        result = cursor.fetchone()
        return result[0]
    
# *************************************************** Deactivate Individual Notifications Section ************************************
@anvil.server.callable()
def updateNotifications(jobcard, createdat, notification):
    _get_current_user()
    with db_cursor() as cursor:
        if notification == "InvoiceAlerts":
            query = """
                Update tbl_notifications SET active = 0 WHERE jobcard = %s AND created_at = %s
            """
        elif notification == "IncompleteDefects":
            query = """
                Update tbl_incomplete_defects SET active = 0 WHERE jobcard = %s AND created_at = %s
            """
        elif notification == "TechnicianPortal":
            query = """
                Update tbl_technician_portal_notifications SET active = 0 WHERE jobcard = %s AND created_at = %s
            """
        cursor.execute(query, (jobcard, createdat))
        

# *************************************************** Role-Based Notification Section ************************************
@anvil.server.callable()
def publish_role_notification(JobCardID, message):
    _get_current_user()
    jobcard = anvil.server.call("getNotificationJobCardRef", JobCardID)
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_notifications
            (jobcard, message, role, created_at, active)
            VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (jobcard, message, 1, datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))), True))

@anvil.server.callable()
def fetch_role_notifications(logged_in_user):
    _get_current_user()
    if not logged_in_user:
        return []

    role_id = logged_in_user['role_id']
    notifications = []

    with db_cursor() as cursor:
        query = """
            SELECT jobcard, message, created_at
            FROM tbl_notifications
            WHERE active = 1
            AND role = %s
            ORDER BY created_at DESC
            LIMIT 1
        """
        cursor.execute(query, (role_id,))
        rows = cursor.fetchone()
        if rows:
            notifications.append({
                "jobcard": rows[0],
                "message": rows[1]
            })

    return notifications

@anvil.server.callable()
def fetch_active_notifications(logged_in_user):
    _get_current_user()
    if not logged_in_user:
        return []

    role_id = logged_in_user['role_id']
    notifications = []

    with db_cursor() as cursor:
        query = """
            SELECT jobcard, message, created_at, active
            FROM tbl_notifications
            WHERE active = 1
              AND role = %s
            ORDER BY created_at DESC
        """
        cursor.execute(query, (role_id,))
        rows = cursor.fetchall()

        for row in rows:
            notifications.append({
                "jobcard": row[0],
                "message": row[1],
                "created_at": row[2].strftime("%Y-%m-%d %H:%M:%S"),
                "active": "Yes" if row[3] in (1, True) else "No"
            })

    return notifications

@anvil.server.callable()
def deactivate_notifications():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_notifications
            SET active = 0
            WHERE active = 1
        """
        cursor.execute(query,)

#  *************************************************** Incomplete Defects Notification Section ************************************
@anvil.server.callable()
def publish_defects_notification(JobCardID, message):
    _get_current_user()
    jobcard = anvil.server.call("getNotificationJobCardRef", JobCardID)
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_incomplete_defects
            (jobcard, message, role, created_at, active)
            VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (jobcard, message, 1, datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))), True))

@anvil.server.callable()
def fetch_active_incomplete_defects_info(logged_in_user):
    _get_current_user()
    if not logged_in_user:
        return []

    role_id = logged_in_user['role_id']
    info = []

    with db_cursor() as cursor:
        query = """
            SELECT jobcard, message, created_at, active
            FROM tbl_incomplete_defects
            WHERE active = 1
              AND role = %s
            ORDER BY created_at DESC
        """
        cursor.execute(query, (role_id,))
        rows = cursor.fetchall()

        for row in rows:
            info.append({
                "jobcard": row[0],
                "message": row[1],
                "created_at": row[2].strftime("%Y-%m-%d %H:%M:%S"),
                "active": "Yes" if row[3] in (1, True) else "No"
            })

    return info

@anvil.server.callable()
def deactivate_incomplete_defects_info():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_incomplete_defects
            SET active = 0
            WHERE active = 1
        """
        cursor.execute(query,)

# *************************************************** Main Dashboard Notifications Section ************************************
@anvil.server.callable()
def fetch_all_dashboard_notifications(logged_in_user):
    # Return empty structures if no user is logged in
    _get_current_user()
    if not logged_in_user:
        return {
            "notifications": [],
            "incomplete_defects": [],
            "technician_portal": []
        }

    role_id = logged_in_user['role_id']
    
    # Initialize the result dictionary
    result = {
        "notifications": [],
        "incomplete_defects": [],
        "technician_portal": [],
        "pricing_alert": []
    }

    # Use a single database connection for all three queries
    with db_cursor() as cursor:
        
        # 1. Fetch Role Notifications
        query_notif = """
            SELECT jobcard, message, created_at
            FROM tbl_notifications
            WHERE active = 1 AND role = %s
            ORDER BY created_at DESC
            LIMIT 1
        """
        cursor.execute(query_notif, (role_id,))
        notif_row = cursor.fetchone()
        if notif_row:
            result["notifications"].append({
                "jobcard": notif_row[0],
                "message": notif_row[1]
            })

        # 2. Fetch Incomplete Defects Info
        query_defects = """
            SELECT jobcard, message, created_at, active
            FROM tbl_incomplete_defects
            WHERE active = 1 AND role = %s
            ORDER BY created_at DESC
        """
        cursor.execute(query_defects, (role_id,))
        for row in cursor.fetchall():
            result["incomplete_defects"].append({
                "jobcard": row[0],
                "message": row[1],
                "created_at": row[2].strftime("%Y-%m-%d %H:%M:%S"),
                "active": "Yes" if row[3] in (1, True) else "No"
            })

        # 3. Fetch Technician Portal Info
        query_tech = """
            SELECT jobcard, message, created_at, active
            FROM tbl_technician_portal_notifications
            WHERE active = 1 AND role = %s
            ORDER BY created_at DESC
        """
        cursor.execute(query_tech, (role_id,))
        for row in cursor.fetchall():
            result["technician_portal"].append({
                "jobcard": row[0],
                "message": row[1],
                "created_at": row[2].strftime("%Y-%m-%d %H:%M:%S"),
                "active": "Yes" if row[3] in (1, True) else "No"
            })
        
        # 4. Fetch Pricing Alert Details
        query = """
            WITH latest_stock AS (
                SELECT sp.CarPart, sp.UnitCost
                FROM tbl_stockparts sp
                INNER JOIN (
                    SELECT CarPart, MAX(ID) AS max_id
                    FROM tbl_stockparts
                    GROUP BY CarPart
                ) latest
                ON sp.CarPart = latest.CarPart 
                AND sp.ID = latest.max_id
            )
            SELECT 
                cpn.Name,
                cpn.PartNo,
                ls.UnitCost AS Cost,
                psp.Amount AS Selling,
                psp.SaleDiscount AS Discount
            FROM tbl_carpartnames cpn
            INNER JOIN tbl_partssellingprice psp 
                ON cpn.ID = psp.CarPartsNamesID
                AND psp.Amount > -1
            INNER JOIN latest_stock ls 
                ON cpn.ID = ls.CarPart
            WHERE 
                ls.UnitCost IS NOT NULL
                AND psp.Amount IS NOT NULL
                AND ls.UnitCost > psp.Amount
            ORDER BY cpn.Name
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        for count, row in enumerate(rows, start=1):
            result["pricing_alert"].append(
                {
                    "No":                       count,
                    "Name":                     row[0],
                    "PartNo":                   row[1],
                    "Cost":                     f"{float(row[2]):,.2f}" if row[2] is not None else None,
                    "Selling":                  f"{float(row[3]):,.2f}" if row[3] is not None else None,
                    "Discount":                 row[4],
                    "BuyingPriceExceedsSelling": "Yes",
                }
            )

    return result

#  *************************************************** Self Service Portal - Technician Portal Notification Section ************************************
@anvil.server.callable()
def publish_technician_portal_notification(JobCardID, message):
    _get_current_user()
    jobcard = anvil.server.call("getNotificationJobCardRef", JobCardID)
    with db_cursor() as cursor:
        query = """
            INSERT INTO tbl_technician_portal_notifications
            (jobcard, message, role, created_at, active)
            VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (jobcard, message, 1,datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))), True))

@anvil.server.callable()
def fetch_active_technician_portal_info(logged_in_user):
    _get_current_user()
    if not logged_in_user:
        return []

    role_id = logged_in_user['role_id']
    info = []

    with db_cursor() as cursor:
        query = """
            SELECT jobcard, message, created_at, active
            FROM tbl_technician_portal_notifications
            WHERE active = 1
              AND role = %s
            ORDER BY created_at DESC
        """
        cursor.execute(query, (role_id,))
        rows = cursor.fetchall()

        for row in rows:
            info.append({
                "jobcard": row[0],
                "message": row[1],
                "created_at": row[2].strftime("%Y-%m-%d %H:%M:%S"),
                "active": "Yes" if row[3] in (1, True) else "No"
            })

    return info

@anvil.server.callable()
def deactivate_technician_portal_info():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_technician_portal_notifications
            SET active = 0
            WHERE active = 1
        """
        cursor.execute(query,)


# *************************************************** Self Service Portal - Technician Assigned Jobcards by Status Section ************************************

@anvil.server.callable()
def get_technician_jobcards_by_status():
    _get_current_user()
    with db_cursor() as cursor:
        jobs_query = """
            SELECT 
                tbl_jobcarddetails.ReceivedDate,
                tbl_technicians.Fullname,
                tbl_jobcarddetails.JobCardRef,
                tbl_jobcarddetails.ClientInstruction,
                tbl_jobcarddetails.Status
            FROM 
                tbl_pendingassignedjobs
            JOIN 
                tbl_jobcarddetails 
                ON tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
            JOIN 
                tbl_technicians
                ON tbl_pendingassignedjobs.TechnicianID = tbl_technicians.ID
            WHERE 
                tbl_jobcarddetails.Status IN ('Checked In', 'Create Quote', 'Confirm Quote', 'In Service')
            ORDER BY 
                tbl_jobcarddetails.ReceivedDate DESC;
        """
        cursor.execute(jobs_query)
        assigned_rows = cursor.fetchall()

        jobcards = []

        for i, row in enumerate(assigned_rows, start=1):
            received_date = row[0].strftime("%Y-%m-%d") if row[0] else ""
            technician = row[1]
            jobcarderef = row[2]
            raw_instruction = row[3] or ""
            status_display = row[4]   # direct pass-through

            # Clean instructions: strip HTML and normalise lines
            text_only = re.sub(r'<[^>]+>', '', raw_instruction)
            lines = [line.strip() for line in text_only.splitlines() if line.strip()]
            formatted_instruction = "\n".join(lines)

            card = {
                "id": i,
                "ReceivedDate": received_date,
                "Technician": technician,
                "JobCardRef": str(jobcarderef),
                "Instruction": formatted_instruction,
                "status": status_display,
            }

            jobcards.append(card)

        return jobcards

# *************************************************** Self Service Portal - Technician Defects and Requested Parts Details Section ************************************
@anvil.server.callable()
def getExistingJobCardInDefectsAndRequestedParts(JobCardID):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT 
                ID 
            FROM 
                tbl_techniciandefectsandrequestedparts
            WHERE 
                JobCardRefID = %s     
        """
        cursor.execute(query, (JobCardID,))
        result = cursor.fetchone()
        return result[0] if result else None

@anvil.server.callable()
def saveOrUpdateTechnicianPortalDefectsAndRequestedParts(
    jobcardID, tech_notes, defects, requiredparts, staff, signature
):
    _get_current_user()
    with db_cursor() as cursor:
        # 1. Get signature bytes safely
        signature_bytes = signature.get_bytes() if signature else None

        # 2. Update the main Job Card notes
        query_notes = "UPDATE tbl_jobcarddetails SET Notes = %s WHERE ID = %s"
        cursor.execute(query_notes, (tech_notes, jobcardID))

        # 3. Check if record exists (Direct SQL instead of anvil.server.call)
        check_query = "SELECT ID FROM tbl_techniciandefectsandrequestedparts WHERE JobCardRefID = %s"
        cursor.execute(check_query, (jobcardID,))
        existing_record = cursor.fetchone()

        if existing_record:
            # UPDATE existing findings
            # Note: Ensure these column names match your 'describe' output exactly
            query_update = """
                UPDATE tbl_techniciandefectsandrequestedparts   
                SET Defects = %s, TechnicianPortalRequestedParts = %s, PreparedByStaff = %s, Signature = %s
                WHERE JobCardRefID = %s
            """
            cursor.execute(query_update, (defects, requiredparts, staff, signature_bytes, jobcardID))
            #Update technician portal notification
            publish_technician_portal_notification(jobcardID, "Technician Portal updates made under Checked In. Please review the changes.")
        else:
            # INSERT new findings + Update Status
            query_status = "UPDATE tbl_jobcarddetails SET Status = 'Create Quote' WHERE ID = %s"
            cursor.execute(query_status, (jobcardID,))
            
            query_insert = """
                INSERT INTO tbl_techniciandefectsandrequestedparts 
                (JobCardRefID, Defects, PricedDefectsList, TechnicianPortalRequestedParts, RequestedParts, PreparedByStaff, Signature)
                VALUES (%s, %s, %s, %s, %s, %s, %s) 
            """
            cursor.execute(query_insert, (jobcardID, defects, defects, requiredparts, requiredparts, staff, signature_bytes))
            #New technician portal notification
            publish_technician_portal_notification(jobcardID, "New Technician Portal entry made under Checked In. Please review the entry.")
            
    return True

        

@anvil.server.callable()
def get_jobcard_and_defect_details(jobcard_ref_str):
    """
    Joins job card details with technician defects and returns 
    data including a base64 encoded signature.
    """
    _get_current_user()
    query = """
        SELECT 
            j.Notes, 
            t.Defects, 
            t.TechnicianPortalRequestedParts, 
            t.PreparedByStaff, 
            t.Signature
        FROM tbl_jobcarddetails j
        JOIN tbl_techniciandefectsandrequestedparts t ON t.JobCardRefID = j.ID
        WHERE j.JobCardRef = %s
        ORDER BY t.ID DESC
    """
    
    with db_cursor() as cursor:
        cursor.execute(query, (jobcard_ref_str,))
        results = cursor.fetchall()

    formatted_data = []
    
    for row in results:
        notes, defects, parts, staff, sig_bytes = row
        
        # Convert blob signature to base64 string
        sig_base64 = None
        if sig_bytes:
            # We encode the bytes and decode to utf-8 to get a clean string
            sig_base64 = base64.b64encode(sig_bytes).decode('utf-8')
        
        if parts == "None":
            parts = ""

        formatted_data.append({
            "Notes": notes,
            "Defects": defects,
            "RequestedParts": parts,
            "PreparedByStaff": staff,
            "Signature": sig_base64  # This is now a base64 string
        })

    return formatted_data
# *************************************************** Self Service Portal - Car Part Names and Categories Section ************************************
@anvil.server.callable()
def getCarPartNamesAndCategory():
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            SELECT Name, Category, PartNo
            FROM tbl_carpartnames
            ORDER BY Category ASC, Name ASC
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Use a list comprehension and direct indices for maximum speed
        return [
            {
                "Name": r[0], 
                "Category": r[1], 
                "PartNo": r[2]
            } 
            for r in rows
        ]
# *************************************************** Self Service Portal - List of Technicians for Dropdown Section ************************************
@anvil.server.callable()
def get_technicians_list():
    _get_current_user()
    with db_cursor() as cursor:
        # Use 'AS combined_name' to ensure both sides of the UNION have the same key
        query = """
                SELECT Staff AS combined_name FROM tbl_checkstaff WHERE IsArchived = FALSE
                UNION ALL
                SELECT Fullname AS combined_name FROM tbl_technicians WHERE IsArchived = FALSE
                ORDER BY combined_name ASC
            """
        cursor.execute(query)
        rows = cursor.fetchall()
        
        result = []
        for r in rows:
            # Handle both dictionary-style and tuple-style rows safely
            if isinstance(r, dict):
                name = r["combined_name"]
            else:
                # If you only SELECT combined_name, it will be at index 0
                name = r[0] 
                
            result.append({"Fullname": name})
        
        # Extract just the values into a flat list
        return [r["Fullname"] if isinstance(r, dict) else r[0] for r in rows]

# *************************************************** Self Service Portal - Customer Feedback And Approved Parts by Job Card Reference Section ************************************
@anvil.server.callable()
def get_parts_and_feedback_by_jobcardref(jobcard_ref):

    _get_current_user()
    with db_cursor() as cursor:
        query = """
        SELECT
            tbl_quotationpartsandservicesfeedback.Item,
            tbl_quotationpartsandservicesfeedback.QuantityIssued,
            tbl_clientquotationfeedback.Remarks
        FROM tbl_jobcarddetails
        INNER JOIN tbl_clientquotationfeedback
            ON tbl_jobcarddetails.ID = tbl_clientquotationfeedback.AssignedJobID
        INNER JOIN tbl_quotationpartsandservicesfeedback
            ON tbl_jobcarddetails.ID = tbl_quotationpartsandservicesfeedback.AssignedJobID
        WHERE tbl_jobcarddetails.JobCardRef = %s
          AND tbl_quotationpartsandservicesfeedback.Item <> 'Labour'
        ORDER BY tbl_quotationpartsandservicesfeedback.ID ASC
         """
        cursor.execute(query, (jobcard_ref,))
        rows = cursor.fetchall()

    # return as list of dicts (Anvil-friendly)
    return [
        {
            "Item": row[0],
            "QuantityIssued": row[1],
            "Remarks": row[2]
        }
        for row in rows
    ]

# *************************************************** Self Service Portal - Car Part Categories Update Section ************************************

# 1. Turn the actual work into a background task
@anvil.server.background_task
def update_carpart_taxonomy_bg(data_list):
    try:
        with db_cursor() as cursor:
            for row in data_list:
                category = row.get('Category')
                patterns = row.get('Patterns')
                
                sql = "UPDATE tbl_carpart_taxonomy SET Patterns = %s WHERE Category = %s"
                cursor.execute(sql, (patterns, category))
                
        # Re-run categorization
        updated_count = _update_carpart_categories(filter_mode="all")
        
        # You can optionally return data from a background task
        return f"Success: Updated taxonomy and categorized {updated_count} parts."
        
    except Exception as e:
        return f"Error saving data: {str(e)}"

# 2. Create a callable function that launches the background task
@anvil.server.callable()
def launch_taxonomy_update(data_list):
    # This launches the task and returns control to the client immediately
    task = anvil.server.launch_background_task('update_carpart_taxonomy_bg', data_list)
    return task
    
@anvil.server.callable()
def load_taxonomy():
    """Load taxonomy into a dict of {category: [patterns]}."""
    taxonomy = {}
    with db_cursor() as cursor:
        cursor.execute("SELECT Category, Patterns FROM tbl_carpart_taxonomy ORDER BY Category ASC")
        rows = cursor.fetchall()
        for category, pattern_str in rows:
            # split by comma, strip spaces
            patterns = [p.strip() for p in pattern_str.split(",") if p.strip()]
            taxonomy[category] = patterns
    
    return taxonomy


# -------------------------------
# Shared categorization function
# -------------------------------

# Pass the taxonomy in as a parameter so we don't fetch it every time
def strict_categorize(name: str, taxonomy: dict) -> str:
    name_l = str(name).lower()
    for category, patterns in taxonomy.items():
        for pat in patterns:
            if re.search(pat, name_l):  
                return category

    # fallback
    if "engine" in name_l:
        return "Engine Components"
    return "Engine Components"

def _update_carpart_categories(filter_mode: str = "all", params: tuple = ()):
    """Update car part categories based on a safe filter_mode.

    filter_mode: "all"  -> no WHERE clause (update every row)
                 "null" -> only rows where Category IS NULL
    """
    _ALLOWED_FILTERS = {
        "all":  "",
        "null": "WHERE tbl_carpartnames.Category IS NULL",
    }
    if filter_mode not in _ALLOWED_FILTERS:
        raise ValueError(f"Invalid filter_mode: {filter_mode!r}")

    where_clause = _ALLOWED_FILTERS[filter_mode]

    # 1. LOAD TAXONOMY EXACTLY ONCE
    taxonomy = load_taxonomy()

    with db_cursor() as cursor:
        select_sql = f"""
            SELECT ID, Name
            FROM tbl_carpartnames
            {where_clause}
        """
        cursor.execute(select_sql, params)
        rows = cursor.fetchall()

        # 2. CATEGORIZE IN MEMORY
        update_data = []
        for part_id, name in rows:
            category = strict_categorize(name, taxonomy)
            update_data.append((category, part_id)) # Add tuple to list

        # 3. BATCH UPDATE EVERYTHING AT ONCE
        if update_data:
            update_sql = """
                UPDATE tbl_carpartnames
                SET Category = %s
                WHERE ID = %s
            """
            cursor.executemany(update_sql, update_data)

        return len(update_data)
    
@anvil.server.callable()
def update_carpart_categories():
    updated_count = _update_carpart_categories(
        filter_mode="all"
    )

    return {
        "status": "ok",
        "updated_rows": updated_count
    }

@anvil.server.callable()
def update_null_carpart_categories():
    updated_count = _update_carpart_categories(
        filter_mode="null"
    )

    return {
        "status": "ok",
        "updated_null_rows": updated_count
    }


# *************************************************** Self Service Portal - Save Work Done by Technician Section ************************************
@anvil.server.callable()
def save_work_done_by_technician(jobcardref, work_done):
    _get_current_user()
    with db_cursor() as cursor:
        query = """
            UPDATE tbl_jobcarddetails
            SET Status ="Verify Task"
            WHERE JobCardRef = %s
        """
        cursor.execute(query, ( jobcardref,))

    #Update work done in job card details
    jobcardid = anvil.server.call('getJobCardID', jobcardref)
    anvil.server.call('saveWorkDoneInJobCard', jobcardid, work_done)

# *************************************************** Tv Cast Dashboard HTTP Endpoint Section ************************************
@anvil.server.http_endpoint("/tv", methods=["GET"], enable_cors=True)
def dashboard_tv_endpoint():
    html = _build_dashboard_html()
    full_html = f"""<!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>BMA Auto Accessories - Dashboard</title>
        <style>
            body {{
                margin: 0;
                padding: 0;
                background-color: #f0f0f0;
            }}
            #status-bar {{
                display: none;
                background-color: #cc0000;
                color: white;
                text-align: center;
                padding: 8px;
                font-family: sans-serif;
                font-size: 14px;
                position: fixed;
                top: 0;
                width: 100%;
                z-index: 999;
            }}
            #status-bar.visible {{
                display: block;
            }}
            #dashboard-content {{
                padding-top: 0;
            }}
            #dashboard-content.offset {{
                padding-top: 36px;
            }}
        </style>
    </head>
    <body>
        <div id="status-bar">⚠️ Connection lost — retrying... Last updated: <span id="last-updated">—</span></div>
        <div id="dashboard-content">
            {html}
        </div>

        <script>
            const REFRESH_INTERVAL_MS = 60000; // 60 seconds
            const statusBar = document.getElementById('status-bar');
            const content = document.getElementById('dashboard-content');
            const lastUpdated = document.getElementById('last-updated');

            function formatTime(date) {{
                return date.toLocaleTimeString([], {{ hour: '2-digit', minute: '2-digit', second: '2-digit' }});
            }}

            async function refreshDashboard() {{
                try {{
                    const response = await fetch(window.location.href);
                    if (!response.ok) throw new Error('Server returned ' + response.status);

                    const fullHtml = await response.text();

                    // Parse the returned full HTML and extract only the dashboard-content div
                    const parser = new DOMParser();
                    const doc = parser.parseFromString(fullHtml, 'text/html');
                    const newContent = doc.getElementById('dashboard-content');

                    if (newContent) {{
                        content.innerHTML = newContent.innerHTML;
                    }}

                    // Hide error bar on success
                    statusBar.classList.remove('visible');
                    content.classList.remove('offset');
                    lastUpdated.textContent = formatTime(new Date());

                }} catch (err) {{
                    // Show error bar but keep last good data visible
                    statusBar.classList.add('visible');
                    content.classList.add('offset');
                    console.warn('Dashboard refresh failed:', err);
                }}
            }}

            // Kick off the refresh loop
            setInterval(refreshDashboard, REFRESH_INTERVAL_MS);
        </script>
    </body>
    </html>"""
    return anvil.server.HttpResponse(200, full_html, {"Content-Type": "text/html"})


def _build_dashboard_html(font_path: str = os.getenv("FONT_PATH")):
    # === Embed MozillaHeadline font as base64 ===
    #font_path = r"D:\BMAAutoAccessories\venv\Lib\site-packages\BMALocal\theme\assets\fonts\MozillaHeadline.ttf"
    font_base64 = ""
    if os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_base64 = base64.b64encode(f.read()).decode("utf-8")

    # =============================
    # 1. DASHBOARD QUERY
    # =============================
    with db_cursor() as cursor:
        dashboard_query = """
        SELECT ID, MakeAndModel, JobCardRef, RegNo, DueDate, ClientInstruction, Status
        FROM tbl_jobcarddetails 
        WHERE Status IN (
            'Checked In', 'Create Quote', 'Confirm Quote',
            'In Service', 'Verify Task', 'Issue Invoice', 'Ready for Pickup'
        )
        ORDER BY DueDate DESC
        """
        cursor.execute(dashboard_query)
        dashboard_rows = cursor.fetchall()

    # Group jobcards by status
    counts = {
        "Checked In": 0,
        "Create Quote": 0,
        "Confirm Quote": 0,
        "In Service": 0,
        "Verify Task": 0,
        "Issue Invoice": 0,
        "Ready for Pickup": 0
    }

    for r in dashboard_rows:
        status = r[6]
        if status in counts:
            counts[status] += 1

    # =============================
    # 2. Build DASHBOARD HTML
    # =============================
    dashboard_html = f"""
    <style>
        @font-face {{
            font-family: 'Mozilla Headline';
            src: url(data:font/ttf;base64,{font_base64}) format('truetype');
        }}
      body {{
        font-family: 'Headline', sans-serif;
        font-size: 16px;
      }}
      .dashboard-container {{
        text-align: center;
        margin: 10px;
      }}
      .dashboard-title {{
        background-color: #0066cc;
        color: white;
        padding: 10px;
        font-weight: bold;
        font-size: 20px;
        font-family: 'Mozilla Headline';
      }}
      .dashboard-grid {{
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 10px;
        margin-top: 10px;
      }}
      .dashboard-card {{
        border: 1px solid #ccc;
        padding: 20px;
        background: #f9f9f9;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.2);
        border-radius: 5px;
        font-size: 20px;
        font-family: 'Mozilla Headline';
      }}
      .card-icon {{
        font-size: 40px;
        margin-bottom: 4px;
        display: block;
      }}
      .dashboard-footer {{
        grid-column: span 3;
        border: 1px solid #ccc;
        padding: 20px;
        background: #f9f9f9;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.2);
        border-radius: 5px;
        font-size: 20px;
        font-family: 'Mozilla Headline';
      }}
    </style>

    <div class="dashboard-container">
      <div class="dashboard-title">Workflow Status Snapshot</div>
      <div class="dashboard-grid">
        <div class="dashboard-card"><span class="card-icon">✔️</span> 1. Checked In: {counts["Checked In"]}</div>
        <div class="dashboard-card"><span class="card-icon">📝</span> 2. Create Quote: {counts["Create Quote"]}</div>
        <div class="dashboard-card"><span class="card-icon">📄</span> 3. Confirm Quote: {counts["Confirm Quote"]}</div>
        <div class="dashboard-card"><span class="card-icon">🔧</span> 4. In Service: {counts["In Service"]}</div>
        <div class="dashboard-card"><span class="card-icon">✅</span> 5. Verify Task: {counts["Verify Task"]}</div>
        <div class="dashboard-card"><span class="card-icon">💰</span> 6. Issue Invoice: {counts["Issue Invoice"]}</div>
        <div class="dashboard-footer"><span class="card-icon">🚚</span> 7. Ready for Pickup: {counts["Ready for Pickup"]}</div>
      </div>
    </div>
    """

    # =============================
    # 3. ASSIGNED JOBS QUERY
    # =============================
    with db_cursor() as cursor:
        jobs_query = """
            SELECT 
                tbl_jobcarddetails.ReceivedDate,
                tbl_technicians.Fullname,
                tbl_jobcarddetails.RegNo,
                tbl_jobcarddetails.ClientInstruction,
                tbl_jobcarddetails.Status
            FROM 
                tbl_pendingassignedjobs
            JOIN 
                tbl_jobcarddetails 
                ON tbl_pendingassignedjobs.JobCardRefID = tbl_jobcarddetails.ID
            JOIN 
                tbl_technicians
                ON tbl_pendingassignedjobs.TechnicianID = tbl_technicians.ID
            WHERE 
                tbl_jobcarddetails.Status IN (
            'Checked In', 'Create Quote', 'Confirm Quote',
            'In Service', 'Verify Task', 'Issue Invoice', 'Ready for Pickup'
        )
            ORDER BY 
                tbl_jobcarddetails.ReceivedDate DESC;
        """
        cursor.execute(jobs_query)
        assigned_rows = cursor.fetchall()

    # =============================
    # 4. Assigned Jobs Table CSS
    # =============================
    assigned_css = f"""
    <style>
        @font-face {{
            font-family: 'Mozilla Headline';
            src: url(data:font/ttf;base64,{font_base64}) format('truetype');
        }}
        .items-table {{
            font-family: 'Mozilla Headline';
            border-collapse: collapse;
            width: 100%;
            margin: 20px 24px 24px 0;
            background-color: white;
            border-radius: 2px;
            overflow: hidden;
            box-shadow: 0 2px 2px 0 rgba(0, 0, 0, 0.14),
                        0 3px 1px -2px rgba(0, 0, 0, 0.2),
                        0 1px 5px 0 rgba(0, 0, 0, 0.12);
        }}
        .items-table th {{
            font-family: 'Mozilla Headline';
            background-color: #f5f5f5;
            border-bottom: 1px solid #e0e0e0;
            padding: 16px;
            text-align: left;
            font-weight: bold;
            font-size: 12px;
            color: rgba(0,0,0,0.87);
            text-transform: uppercase;
            letter-spacing: .5px;
        }}
        .items-table td {{
            border-bottom: 1px solid rgba(0,0,0,0.12);
            padding: 16px;
            font-size: 12px;
            color: rgba(0,0,0,0.87);
        }}
        .items-table .item-row:hover {{
            background-color: rgba(0,0,0,0.04);
        }}
    </style>
    """

    # =============================
    # 5. Build Assigned Jobs HTML Table
    # =============================
    assigned_html = assigned_css + """
    <h2 style="margin-left:20px;font-family: 'Mozilla Headline';">Assigned Jobs</h2>
    <table class="items-table">
        <thead>
            <tr>
                <th>No</th>
                <th>Received Date</th>
                <th>Technician</th>
                <th>Reg No</th>
                <th>Client Instruction</th>
                <th>Status</th>
            </tr>
        </thead>
        <tbody>
    """

    for i, row in enumerate(assigned_rows, start=1):
        received_date = row[0].strftime("%Y-%m-%d") if row[0] else ""
        technician = row[1]
        reg_no = row[2]
        raw_instruction = row[3] or ""
        status_raw = row[4]

        # Clean instructions
        text_only = re.sub(r'<[^>]+>', '', raw_instruction)
        lines = [line.strip() for line in text_only.splitlines() if line.strip()]
        safe_lines = [html_module.escape(line) for line in lines]
        formatted_instruction = "<br>".join(safe_lines)

        # Status
        status_display = status_raw if status_raw in [
            'Checked In', 'Create Quote', 'Confirm Quote',
            'In Service', 'Verify Task', 'Issue Invoice', 'Ready for Pickup'
        ] else "Unknown Status"

        assigned_html += f"""
            <tr class="item-row">
                <td>{i}</td>
                <td>{received_date}</td>
                <td>{technician}</td>
                <td>{reg_no}</td>
                <td>{formatted_instruction}</td>
                <td>{status_display}</td>
            </tr>
        """

    assigned_html += "</tbody></table>"

    # =============================
    # 6. RETURN COMBINED HTML
    # =============================
    return dashboard_html + "<br>" + assigned_html

# ***************************************************** Update Payment Details After Invoice Changes Section ************************************
@anvil.server.callable()
def updatePaymentDetailsAfterInvoiceChanges(jobcard_id):
    _get_current_user()

    invoiceDetails = get_invoice_details_by_job_id(jobcard_id)
    invoiceTotal = sum(float(row['Total'] or 0) for row in invoiceDetails)

    with db_cursor() as cursor:

        # Step 1: Get payments ordered DESC (latest first)
        cursor.execute("""
            SELECT 
                ID, AmountPaid, Discount
            FROM tbl_payments
            WHERE JobCardRefID = %s
            ORDER BY ID
        """, (jobcard_id,))

        payments = cursor.fetchall()

        if not payments:
            return []

        # Step 2: Recalculate balances
        running_balance = invoiceTotal
        updated_results = []

        for row in payments:
            payment_id = row[0]
            amount_paid = float(row[1] or 0)
            discount = float(row[2] or 0)

            # Calculate new balance
            new_balance = running_balance - amount_paid - discount

            # Prevent negative balance (optional safety)
            if new_balance < 0:
                new_balance = 0

            # Update DB
            cursor.execute("""
                UPDATE tbl_payments
                SET Balance = %s
                WHERE ID = %s
            """, (new_balance, payment_id))

            # Update running balance for next row
            running_balance = new_balance
     
